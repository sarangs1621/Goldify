from fastapi import FastAPI, APIRouter, Depends, HTTPException, status, Response, Request, Cookie
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from dotenv import load_dotenv
from motor.motor_asyncio import AsyncIOMotorClient
from slowapi import Limiter, _rate_limit_exceeded_handler
from slowapi.util import get_remote_address
from slowapi.errors import RateLimitExceeded
import os
import re
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timezone, timedelta
from passlib.context import CryptContext
import jwt
from decimal import Decimal
from bson import Decimal128, ObjectId
import secrets

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# ============================================================================
# ACCOUNTING CONFIGURATION - STRICT TAXONOMY
# ============================================================================

# Valid account types (LOCKED - do not modify)
VALID_ACCOUNT_TYPES = {'asset', 'income', 'expense', 'liability', 'equity'}

# Protected standard accounts that cannot be deleted
PROTECTED_ACCOUNTS = {
    'Cash', 'Bank', 'Petty Cash', 'Sales Income', 'Sales', 
    'Gold Exchange Income', 'Gold Exchange'
}

def validate_account_type(account_type: str) -> bool:
    """Validate that account type is in the allowed taxonomy"""
    return account_type.lower() in VALID_ACCOUNT_TYPES

def calculate_balance_delta(account_type: str, transaction_type: str, amount: float) -> float:
    """
    Calculate balance change based on account type and transaction type.
    
    ACCOUNTING RULES (NON-NEGOTIABLE):
    - ASSET/EXPENSE: Debit increases (+), Credit decreases (-)
    - INCOME/LIABILITY/EQUITY: Credit increases (+), Debit decreases (-)
    
    Examples:
    - Cash (ASSET) + Debit $100 = +$100 balance
    - Cash (ASSET) + Credit $100 = -$100 balance
    - Sales Income (INCOME) + Credit $100 = +$100 balance
    - Sales Income (INCOME) + Debit $100 = -$100 balance
    """
    account_type = account_type.lower()
    
    if account_type in ['asset', 'expense']:
        # Debit increases, Credit decreases
        return amount if transaction_type == 'debit' else -amount
    else:  # income, liability, equity
        # Credit increases, Debit decreases
        return amount if transaction_type == 'credit' else -amount

def get_normal_balance(account_type: str) -> str:
    """
    Get the normal balance side for an account type.
    Returns: 'debit' or 'credit'
    """
    account_type = account_type.lower()
    if account_type in ['asset', 'expense']:
        return 'debit'
    else:  # income, liability, equity
        return 'credit'

# JWT Configuration
JWT_SECRET = os.environ.get('JWT_SECRET', 'your-secret-key-change-in-production')
JWT_ALGORITHM = 'HS256'
JWT_EXPIRATION_HOURS = 24

# ============================================================================
# RATE LIMITING CONFIGURATION
# ============================================================================

# Custom function to get user ID from request for rate limiting
def get_user_identifier(request: Request) -> str:
    """
    Get user identifier for rate limiting.
    Returns user_id for authenticated requests, IP address for unauthenticated.
    """
    try:
        # Try to get token from cookie first
        token = request.cookies.get('access_token')
        
        # Fallback to Authorization header
        if not token:
            auth_header = request.headers.get('Authorization')
            if auth_header and auth_header.startswith('Bearer '):
                token = auth_header.split(' ')[1]
        
        if token:
            try:
                payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
                user_id = payload.get('user_id')
                if user_id:
                    return f"user:{user_id}"
            except:
                pass
    except:
        pass
    
    # Fallback to IP address for unauthenticated requests
    return f"ip:{get_remote_address(request)}"

# Initialize rate limiter with custom key function
limiter = Limiter(key_func=get_user_identifier)

app = FastAPI()
api_router = APIRouter(prefix="/api")

# Add rate limiter to app state
app.state.limiter = limiter
app.add_exception_handler(RateLimitExceeded, _rate_limit_exceeded_handler)

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
security = HTTPBearer(auto_error=False)  # auto_error=False makes it optional

# ============================================================================
# SECURITY HEADERS MIDDLEWARE
# ============================================================================

from starlette.middleware.base import BaseHTTPMiddleware
from starlette.responses import Response as StarletteResponse

class SecurityHeadersMiddleware(BaseHTTPMiddleware):
    """
    Middleware to add comprehensive security headers to all HTTP responses.
    
    Security Headers Implemented:
    - Content-Security-Policy (CSP): Restricts resource loading to prevent XSS
    - X-Frame-Options: Prevents clickjacking attacks
    - X-Content-Type-Options: Prevents MIME type sniffing
    - Strict-Transport-Security (HSTS): Enforces HTTPS connections
    - X-XSS-Protection: Enables browser XSS filtering
    - Referrer-Policy: Controls referrer information
    - Permissions-Policy: Controls browser features/APIs
    """
    
    async def dispatch(self, request: Request, call_next):
        response = await call_next(request)
        
        # Content Security Policy (CSP)
        # Note: 'unsafe-inline' and 'unsafe-eval' are required for React apps
        # In production with build optimization, these can be replaced with nonces/hashes
        csp_directives = [
            "default-src 'self'",
            "script-src 'self' 'unsafe-inline' 'unsafe-eval'",
            "style-src 'self' 'unsafe-inline' https://fonts.googleapis.com",
            "img-src 'self' data: https: blob:",
            "font-src 'self' data: https://fonts.gstatic.com",
            "connect-src 'self' http://localhost:3000 http://localhost:8001 http://127.0.0.1:3000 http://127.0.0.1:8001",
            "frame-ancestors 'none'",
            "base-uri 'self'",
            "form-action 'self'",
            "object-src 'none'",
            "upgrade-insecure-requests"
        ]
        response.headers['Content-Security-Policy'] = "; ".join(csp_directives)
        
        # X-Frame-Options: Prevent clickjacking by denying iframe embedding
        response.headers['X-Frame-Options'] = 'DENY'
        
        # X-Content-Type-Options: Prevent MIME type sniffing
        response.headers['X-Content-Type-Options'] = 'nosniff'
        
        # Strict-Transport-Security (HSTS): Force HTTPS for 1 year including subdomains
        # preload: Allows inclusion in browser HSTS preload lists
        response.headers['Strict-Transport-Security'] = 'max-age=31536000; includeSubDomains; preload'
        
        # X-XSS-Protection: Enable browser XSS filtering
        # mode=block: Block page rendering if XSS detected
        response.headers['X-XSS-Protection'] = '1; mode=block'
        
        # Referrer-Policy: Control referrer information sent with requests
        response.headers['Referrer-Policy'] = 'strict-origin-when-cross-origin'
        
        # Permissions-Policy: Disable unnecessary browser features
        # Restricts access to geolocation, camera, microphone, etc.
        permissions_policy = [
            "geolocation=()",
            "camera=()",
            "microphone=()",
            "payment=()",
            "usb=()",
            "magnetometer=()",
            "gyroscope=()",
            "accelerometer=()"
        ]
        response.headers['Permissions-Policy'] = ", ".join(permissions_policy)
        
        return response


# ============================================================================
# INPUT SANITIZATION MIDDLEWARE
# ============================================================================

from validators import sanitize_html, sanitize_text_field, PartyValidator
import json

class InputSanitizationMiddleware(BaseHTTPMiddleware):
    """
    Middleware to automatically sanitize all incoming request data.
    
    Protection Features:
    - Removes HTML tags and scripts from text inputs
    - Escapes special characters
    - Validates data types
    - Prevents XSS attacks through input sanitization
    
    Applied to all POST, PUT, PATCH requests with JSON body.
    """
    
    def sanitize_value(self, value: Any) -> Any:
        """
        Recursively sanitize values in request data.
        """
        if isinstance(value, str):
            # Only sanitize strings that are not UUIDs, dates, or specific fields
            # Avoid sanitizing IDs, dates, and technical fields
            if len(value) > 0 and not self._is_technical_field(value):
                return sanitize_html(value)
            return value
        elif isinstance(value, dict):
            return {k: self.sanitize_value(v) for k, v in value.items()}
        elif isinstance(value, list):
            return [self.sanitize_value(item) for item in value]
        else:
            return value
    
    def _is_technical_field(self, value: str) -> bool:
        """
        Check if value is a technical field that shouldn't be sanitized.
        Returns True for UUIDs, ISO dates, email-like patterns.
        """
        # Check for UUID pattern
        if re.match(r'^[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}$', value, re.IGNORECASE):
            return True
        # Check for ISO date pattern
        if re.match(r'^\d{4}-\d{2}-\d{2}', value):
            return True
        # Short technical strings (likely IDs)
        if len(value) < 5 and value.replace('-', '').replace('_', '').isalnum():
            return True
        return False
    
    async def dispatch(self, request: Request, call_next):
        # Only sanitize body data on state-changing methods
        if request.method in ['POST', 'PUT', 'PATCH']:
            try:
                # Read the body
                body = await request.body()
                if body:
                    try:
                        # Parse JSON
                        data = json.loads(body)
                        # Sanitize the data
                        sanitized_data = self.sanitize_value(data)
                        # Update request body with sanitized data
                        request._body = json.dumps(sanitized_data).encode('utf-8')
                    except json.JSONDecodeError:
                        # Not JSON, skip sanitization
                        pass
            except Exception as e:
                # If any error in sanitization, log but don't break the request
                logging.warning(f"Input sanitization error: {str(e)}")
        
        response = await call_next(request)
        return response


# ============================================================================
# HTTPS ENFORCEMENT MIDDLEWARE (Phase 7)
# ============================================================================

class HTTPSRedirectMiddleware(BaseHTTPMiddleware):
    """
    Middleware to enforce HTTPS by redirecting all HTTP requests to HTTPS.
    
    Security Benefits:
    - Forces encrypted connections
    - Works with HSTS header for comprehensive HTTPS enforcement
    - Prevents downgrade attacks
    
    Note: In production, this should be combined with HSTS preloading
    """
    
    async def dispatch(self, request: Request, call_next):
        # Check if request is HTTP (not HTTPS)
        # In production with reverse proxy, check X-Forwarded-Proto header
        forwarded_proto = request.headers.get('X-Forwarded-Proto', '')
        
        # If explicitly HTTP or no secure connection
        if forwarded_proto == 'http' or (
            not forwarded_proto and 
            request.url.scheme == 'http' and 
            request.url.hostname not in ['localhost', '127.0.0.1', 'testserver']
        ):
            # Construct HTTPS URL
            https_url = request.url.replace(scheme='https')
            return StarletteResponse(
                status_code=301,  # Permanent redirect
                headers={'Location': str(https_url)}
            )
        
        response = await call_next(request)
        return response

# ============================================================================
# CSRF PROTECTION MIDDLEWARE
# ============================================================================

def generate_csrf_token() -> str:
    """
    Generate a cryptographically secure CSRF token.
    Uses secrets module for secure random generation (128-bit token).
    """
    return secrets.token_urlsafe(32)

class CSRFProtectionMiddleware(BaseHTTPMiddleware):
    """
    Middleware to validate CSRF tokens on state-changing HTTP methods.
    
    Implementation: Double-Submit Cookie Pattern
    - CSRF token is stored in a readable cookie (not httponly)
    - Client must send the token in X-CSRF-Token header
    - Server validates that cookie value matches header value
    
    Protected Methods: POST, PUT, PATCH, DELETE
    Exempt Endpoints: /api/auth/login, /api/auth/register, /api/auth/request-password-reset,
                      /api/auth/reset-password, /api/health
    
    Security Benefits:
    - CSRF attacks can't read cookies (same-origin policy)
    - Requires both cookie AND custom header to match
    - Only state-changing operations are protected
    """
    
    # Endpoints that are exempt from CSRF validation
    EXEMPT_PATHS = {
        '/api/auth/login',
        '/api/auth/register',
        '/api/auth/request-password-reset',
        '/api/auth/reset-password',
        '/api/health'
    }
    
    async def dispatch(self, request: Request, call_next):
        # Only validate CSRF on state-changing methods
        if request.method in ['POST', 'PUT', 'PATCH', 'DELETE']:
            # Skip CSRF validation for exempt endpoints
            if request.url.path not in self.EXEMPT_PATHS:
                # Get CSRF token from cookie
                csrf_cookie = request.cookies.get('csrf_token')
                
                # Get CSRF token from header
                csrf_header = request.headers.get('X-CSRF-Token')
                
                # Validate that both exist and match
                if not csrf_cookie or not csrf_header:
                    return StarletteResponse(
                        content='{"detail": "CSRF token missing"}',
                        status_code=403,
                        media_type='application/json'
                    )
                
                if csrf_cookie != csrf_header:
                    return StarletteResponse(
                        content='{"detail": "CSRF token validation failed"}',
                        status_code=403,
                        media_type='application/json'
                    )
        
        response = await call_next(request)
        return response

# ============================================================================
# PERMISSION SYSTEM - RBAC Configuration
# ============================================================================

# Define all permissions in the system
PERMISSIONS = {
    # User Management
    'users.view': 'View users',
    'users.create': 'Create users',
    'users.update': 'Update users',
    'users.delete': 'Delete users',
    
    # Party Management
    'parties.view': 'View parties',
    'parties.create': 'Create parties',
    'parties.update': 'Update parties',
    'parties.delete': 'Delete parties',
    
    # Invoice Management
    'invoices.view': 'View invoices',
    'invoices.create': 'Create invoices',
    'invoices.finalize': 'Finalize invoices',
    'invoices.delete': 'Delete invoices',
    
    # Purchase Management
    'purchases.view': 'View purchases',
    'purchases.create': 'Create purchases',
    'purchases.finalize': 'Finalize purchases',
    'purchases.delete': 'Delete purchases',
    
    # Finance Management
    'finance.view': 'View finance data',
    'finance.create': 'Create financial transactions',
    'finance.delete': 'Delete financial transactions',
    
    # Inventory Management
    'inventory.view': 'View inventory',
    'inventory.adjust': 'Adjust inventory',
    
    # Job Cards
    'jobcards.view': 'View job cards',
    'jobcards.create': 'Create job cards',
    'jobcards.update': 'Update job cards',
    'jobcards.delete': 'Delete job cards',
    
    # Reports
    'reports.view': 'View reports',
    
    # Audit Logs
    'audit.view': 'View audit logs',
    
    # Returns Management
    'returns.view': 'View returns',
    'returns.create': 'Create returns',
    'returns.finalize': 'Finalize returns',
    'returns.delete': 'Delete returns',
}

# Role-Permission Mappings
ROLE_PERMISSIONS = {
    'admin': [
        # Admin has all permissions
        'users.view', 'users.create', 'users.update', 'users.delete',
        'parties.view', 'parties.create', 'parties.update', 'parties.delete',
        'invoices.view', 'invoices.create', 'invoices.finalize', 'invoices.delete',
        'purchases.view', 'purchases.create', 'purchases.finalize', 'purchases.delete',
        'finance.view', 'finance.create', 'finance.delete',
        'inventory.view', 'inventory.adjust',
        'jobcards.view', 'jobcards.create', 'jobcards.update', 'jobcards.delete',
        'reports.view',
        'audit.view',
        'returns.view', 'returns.create', 'returns.finalize', 'returns.delete',
    ],
    'manager': [
        # Manager has most permissions except user deletion and audit logs
        'users.view', 'users.create', 'users.update',
        'parties.view', 'parties.create', 'parties.update', 'parties.delete',
        'invoices.view', 'invoices.create', 'invoices.finalize',
        'purchases.view', 'purchases.create', 'purchases.finalize',
        'finance.view', 'finance.create',
        'inventory.view', 'inventory.adjust',
        'jobcards.view', 'jobcards.create', 'jobcards.update', 'jobcards.delete',
        'reports.view',
        'returns.view', 'returns.create', 'returns.finalize',
    ],
    'staff': [
        # Staff has limited permissions - mostly view and create
        'parties.view',
        'invoices.view', 'invoices.create',
        'purchases.view', 'purchases.create',
        'finance.view',
        'inventory.view',
        'jobcards.view', 'jobcards.create', 'jobcards.update',
        'reports.view',
        'returns.view', 'returns.create',
    ],
}

# Security Configuration
MAX_LOGIN_ATTEMPTS = 5
LOCKOUT_DURATION_MINUTES = 30
PASSWORD_MIN_LENGTH = 12

def decimal_to_float(obj):
    """Convert Decimal128, datetime, and ObjectId objects to JSON-serializable types"""
    if isinstance(obj, dict):
        return {k: decimal_to_float(v) for k, v in obj.items()}
    elif isinstance(obj, list):
        return [decimal_to_float(item) for item in obj]
    elif isinstance(obj, Decimal128):
        return float(obj.to_decimal())
    elif isinstance(obj, datetime):
        return obj.isoformat()
    elif isinstance(obj, ObjectId):
        return str(obj)
    return obj

def float_to_decimal128(value):
    if value is None:
        return None
    return Decimal128(Decimal(str(value)))

def convert_return_to_decimal(return_data: dict) -> dict:
    """
    Convert float values in return data to Decimal128 for precise storage.
    
    Args:
        return_data: Return document dictionary
    
    Returns:
        Return document with Decimal128 values for weights and amounts
    """
    # Convert top-level fields
    if 'total_weight_grams' in return_data:
        return_data['total_weight_grams'] = Decimal128(
            Decimal(str(return_data['total_weight_grams'])).quantize(Decimal('0.001'))
        )
    if 'total_amount' in return_data:
        return_data['total_amount'] = Decimal128(
            Decimal(str(return_data['total_amount'])).quantize(Decimal('0.01'))
        )
    if 'refund_money_amount' in return_data:
        return_data['refund_money_amount'] = Decimal128(
            Decimal(str(return_data['refund_money_amount'])).quantize(Decimal('0.01'))
        )
    if 'refund_gold_grams' in return_data:
        return_data['refund_gold_grams'] = Decimal128(
            Decimal(str(return_data['refund_gold_grams'])).quantize(Decimal('0.001'))
        )
    
    # Convert items
    if 'items' in return_data:
        for item in return_data['items']:
            if 'weight_grams' in item:
                item['weight_grams'] = Decimal128(
                    Decimal(str(item['weight_grams'])).quantize(Decimal('0.001'))
                )
            if 'amount' in item:
                item['amount'] = Decimal128(
                    Decimal(str(item['amount'])).quantize(Decimal('0.01'))
                )
    
    return return_data

# Status transition validation rules
STATUS_TRANSITIONS = {
    "jobcard": {
        "created": ["in_progress", "cancelled"],
        "in_progress": ["completed", "cancelled"],
        "completed": ["delivered", "cancelled"],
        "invoiced": ["delivered"],  # Can be delivered after invoice is fully paid
        "delivered": [],  # Terminal state
        "cancelled": []  # Terminal state
    },
    "invoice": {
        "draft": ["finalized"],
        "finalized": []  # Terminal state
    },
    "purchase": {
        "draft": ["finalized"],
        "finalized": []  # Terminal state
    }
}

def validate_status_transition(entity_type: str, current_status: str, new_status: str) -> tuple[bool, str]:
    """
    Validate if a status transition is allowed
    
    Args:
        entity_type: Type of entity (jobcard, invoice, purchase)
        current_status: Current status of the entity
        new_status: New status to transition to
    
    Returns:
        Tuple of (is_valid, error_message)
    """
    # If status is not changing, allow it
    if current_status == new_status:
        return True, ""
    
    # Get allowed transitions for this entity type
    transitions = STATUS_TRANSITIONS.get(entity_type, {})
    
    # Normalize status strings (handle variations like "in progress" vs "in_progress")
    current_normalized = current_status.lower().replace(" ", "_")
    new_normalized = new_status.lower().replace(" ", "_")
    
    # Get allowed next statuses
    allowed_next = transitions.get(current_normalized, [])
    
    # Check if transition is allowed
    if new_normalized in allowed_next:
        return True, ""
    
    # If we reach here, transition is not allowed
    allowed_str = ", ".join(allowed_next) if allowed_next else "none (terminal state)"
    error_msg = f"Invalid status transition from '{current_status}' to '{new_status}'. Allowed transitions: {allowed_str}"
    
    return False, error_msg

# ===================================
# TIMESTAMP VALIDATION FUNCTIONS
# ===================================
# Enforce absolute correctness for audit safety

def validate_jobcard_timestamps(status: str, completed_at: Optional[datetime], delivered_at: Optional[datetime]) -> tuple[bool, str]:
    """
    Validate Job Card timestamps based on status
    
    Rules:
    - completed_at must exist when status is 'completed' or 'delivered'
    - delivered_at must exist when status is 'delivered'
    - Status rollback must NOT delete timestamps (audit safety)
    
    Args:
        status: Current status of job card
        completed_at: Completion timestamp
        delivered_at: Delivery timestamp
    
    Returns:
        Tuple of (is_valid, error_message)
    """
    # Normalize status
    status_normalized = status.lower().replace(" ", "_")
    
    # Check completed_at for completed status
    if status_normalized in ['completed', 'delivered']:
        if not completed_at:
            return False, f"Job card with status '{status}' must have completed_at timestamp"
    
    # Check delivered_at for delivered status
    if status_normalized == 'delivered':
        if not delivered_at:
            return False, f"Job card with status '{status}' must have delivered_at timestamp"
    
    return True, ""

def validate_invoice_timestamps(status: str, finalized_at: Optional[datetime], payment_status: str, paid_at: Optional[datetime]) -> tuple[bool, str]:
    """
    Validate Invoice timestamps based on status
    
    Rules:
    - finalized_at must exist when status is 'finalized'
    - paid_at must exist when payment_status is 'paid'
    - Status rollback must NOT delete timestamps (audit safety)
    
    Args:
        status: Current status of invoice
        finalized_at: Finalization timestamp
        payment_status: Payment status of invoice
        paid_at: Full payment timestamp
    
    Returns:
        Tuple of (is_valid, error_message)
    """
    # Check finalized_at for finalized status
    if status == 'finalized':
        if not finalized_at:
            return False, f"Invoice with status 'finalized' must have finalized_at timestamp"
    
    # Check paid_at for paid payment status
    if payment_status == 'paid':
        if not paid_at:
            return False, f"Invoice with payment_status 'paid' must have paid_at timestamp"
    
    return True, ""

def validate_purchase_timestamps(status: str, finalized_at: Optional[datetime]) -> tuple[bool, str]:
    """
    Validate Purchase timestamps based on status
    
    Rules:
    - finalized_at must exist when status is 'finalized'
    - Status rollback must NOT delete timestamps (audit safety)
    
    Args:
        status: Current status of purchase
        finalized_at: Finalization timestamp
    
    Returns:
        Tuple of (is_valid, error_message)
    """
    # Check finalized_at for finalized status
    if status == 'finalized':
        if not finalized_at:
            return False, f"Purchase with status 'finalized' must have finalized_at timestamp"
    
    return True, ""

def calculate_purchase_status(paid_amount: float, total_amount: float) -> str:
    """
    Calculate purchase status based on payment amounts (SERVER-SIDE AUTHORITY).
    
    CRITICAL BUSINESS RULES:
    - A purchase is NEVER "Draft" after creation (auto-finalized)
    - Status reflects payment state for accounting integrity
    
    Args:
        paid_amount: Amount paid to vendor
        total_amount: Total purchase amount
    
    Returns:
        Status string: "Finalized (Unpaid)" | "Partially Paid" | "Paid"
    """
    paid_amount = round(float(paid_amount), 2)
    total_amount = round(float(total_amount), 2)
    
    if paid_amount == 0:
        return "Finalized (Unpaid)"
    elif paid_amount < total_amount:
        return "Partially Paid"
    else:  # paid_amount >= total_amount
        return "Paid"

class PaginationMetadata(BaseModel):
    total_count: int
    page: int
    per_page: int
    total_pages: int
    has_next: bool
    has_prev: bool

class PaginationResponse(BaseModel):
    items: List[Any]
    pagination: PaginationMetadata

def create_pagination_response(items: list, total_count: int, page: int, page_size: int):
    """
    Helper function to create standardized pagination response
    
    Args:
        items: List of items for current page
        total_count: Total number of items across all pages
        page: Current page number (1-indexed)
        page_size: Number of items per page
    
    Returns:
        Dictionary with items and pagination metadata
    """
    total_pages = (total_count + page_size - 1) // page_size  # Ceiling division
    
    return {
        "items": items,
        "pagination": {
            "total_count": total_count,
            "page": page,
            "page_size": page_size,
            "total_pages": total_pages,
            "has_next": page < total_pages,
            "has_prev": page > 1
        }
    }

class UserRole(BaseModel):
    role: str
    permissions: List[str] = []

class User(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    username: str
    email: str
    full_name: str
    role: str
    permissions: List[str] = Field(default_factory=list)
    is_active: bool = True
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    is_deleted: bool = False
    deleted_at: Optional[datetime] = None
    deleted_by: Optional[str] = None
    failed_login_attempts: int = 0
    locked_until: Optional[datetime] = None
    last_login: Optional[datetime] = None

class PasswordResetToken(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    token: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    expires_at: datetime
    used: bool = False
    used_at: Optional[datetime] = None

class AuthAuditLog(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: Optional[str] = None
    username: str
    action: str  # login, logout, login_failed, password_change, password_reset
    ip_address: Optional[str] = None
    user_agent: Optional[str] = None
    success: bool
    failure_reason: Optional[str] = None
    timestamp: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class UserCreate(BaseModel):
    username: str
    email: str
    password: str
    full_name: str
    role: str = "staff"

class UserLogin(BaseModel):
    username: str
    password: str

class TokenResponse(BaseModel):
    access_token: str
    token_type: str = "bearer"
    user: User
    csrf_token: str

class InventoryHeader(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    current_qty: float = 0.0  # Direct stock tracking
    current_weight: float = 0.0  # Direct weight tracking in grams
    is_active: bool = True
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    created_by: str
    is_deleted: bool = False

class StockMovement(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    date: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))  # Alias for created_at
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))  # ISO 8601 UTC timestamp
    movement_type: str
    header_id: str
    header_name: str
    description: str
    qty_delta: float
    weight_delta: float
    purity: int
    reference_type: Optional[str] = None
    reference_id: Optional[str] = None
    created_by: str
    notes: Optional[str] = None
    confirmation_reason: Optional[str] = None  # Required for manual adjustments
    is_deleted: bool = False

class Party(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    phone: Optional[str] = None
    address: Optional[str] = None
    party_type: str
    notes: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    created_by: str
    is_deleted: bool = False

class Worker(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    phone: Optional[str] = None
    role: str
    active: bool = True
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    created_by: str
    is_deleted: bool = False

class JobCardItem(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    category: str
    description: str
    qty: int
    weight_in: float
    weight_out: Optional[float] = None
    purity: int
    work_type: str
    remarks: Optional[str] = None
    making_charge_type: Optional[str] = None  # 'flat' or 'per_gram'
    making_charge_value: Optional[float] = None
    vat_percent: Optional[float] = None
    vat_amount: Optional[float] = None

class JobCard(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    job_card_number: str
    card_type: str = "normal"  # Default to "normal" job card (can be "normal" or "template")
    date_created: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))  # Alias for created_at
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))  # ISO 8601 UTC timestamp
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))  # Updated on every modification
    completed_at: Optional[datetime] = None  # Set when status changes to "completed"
    delivered_at: Optional[datetime] = None  # Set when status changes to "delivered"
    delivery_date: Optional[str] = None  # Date-only field (YYYY-MM-DD) for expected delivery
    status: str = "created"
    customer_type: str = "saved"  # "saved" or "walk_in"
    customer_id: Optional[str] = None  # For saved customers only
    customer_name: Optional[str] = None  # For saved customers only
    walk_in_name: Optional[str] = None  # For walk-in customers only
    walk_in_phone: Optional[str] = None  # For walk-in customers only
    worker_id: Optional[str] = None
    worker_name: Optional[str] = None
    items: List[JobCardItem] = []
    notes: Optional[str] = None
    gold_rate_at_jobcard: Optional[float] = None  # MODULE 8: Gold rate at time of job card creation
    locked: bool = False  # True when linked invoice is finalized
    locked_at: Optional[datetime] = None
    locked_by: Optional[str] = None
    is_invoiced: bool = False  # True when job card has been converted to invoice
    invoice_id: Optional[str] = None  # ID of the invoice created from this job card
    created_by: str
    is_deleted: bool = False
    # Template-specific fields
    template_name: Optional[str] = None  # Required when card_type='template'
    delivery_days_offset: Optional[int] = None  # For templates: days from creation to delivery

class InvoiceItem(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    category: Optional[str] = None  # Inventory category for stock tracking
    description: str
    qty: int
    # Gold-specific weight breakdown (3 decimal precision)
    gross_weight: float = 0.0  # Total weight including stones
    stone_weight: float = 0.0  # Weight of stones/gems
    net_gold_weight: float = 0.0  # Calculated: gross_weight - stone_weight
    weight: float  # Legacy field, maps to net_gold_weight for compatibility
    purity: int
    metal_rate: float  # Gold rate per gram
    gold_value: float  # net_gold_weight * metal_rate
    # Charges breakdown
    making_charge_type: Optional[str] = None  # 'per_gram' or 'flat'
    making_value: float  # Making charges
    stone_charges: float = 0.0  # Stone/gem charges
    wastage_charges: float = 0.0  # Wastage charges
    item_discount: float = 0.0  # Item-level discount
    # Tax
    vat_percent: float
    vat_amount: float
    line_total: float  # gold_value + making_value + stone_charges + wastage_charges + vat_amount - item_discount

class Invoice(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    invoice_number: str
    date: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))  # Alias for created_at
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))  # ISO 8601 UTC timestamp
    due_date: Optional[datetime] = None  # For overdue calculations, defaults to invoice date
    customer_type: str = "saved"  # "saved" or "walk_in"
    customer_id: Optional[str] = None  # For saved customers only
    customer_name: Optional[str] = None  # For saved customers only
    customer_phone: Optional[str] = None  # Customer phone (from party or walk-in)
    customer_address: Optional[str] = None  # Customer address (from party)
    customer_gstin: Optional[str] = None  # Customer GSTIN (from party)
    walk_in_name: Optional[str] = None  # For walk-in customers only
    walk_in_phone: Optional[str] = None  # For walk-in customers only
    worker_id: Optional[str] = None  # Worker assigned to the job (from job card)
    worker_name: Optional[str] = None  # Worker name (from job card)
    invoice_type: str = "sale"
    payment_status: str = "unpaid"
    status: str = "draft"  # "draft" or "finalized" - controls when stock is deducted
    finalized_at: Optional[datetime] = None  # Set when invoice is finalized
    finalized_by: Optional[str] = None
    paid_at: Optional[datetime] = None  # Set when balance becomes zero (first full payment)
    items: List[InvoiceItem] = []
    subtotal: float = 0
    discount_amount: float = 0.0  # Invoice-level discount amount
    # Tax breakdown
    tax_type: str = "cgst_sgst"  # "cgst_sgst" or "igst"
    gst_percent: float = 5.0  # Default 5% GST
    cgst_total: float = 0.0  # CGST amount (for intra-state)
    sgst_total: float = 0.0  # SGST amount (for intra-state)
    igst_total: float = 0.0  # IGST amount (for inter-state)
    vat_total: float = 0  # Legacy field, total tax amount
    grand_total: float = 0
    paid_amount: float = 0
    balance_due: float = 0
    notes: Optional[str] = None
    jobcard_id: Optional[str] = None
    created_by: str
    is_deleted: bool = False

class Account(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    account_type: str  # Must be one of: asset, income, expense, liability, equity
    opening_balance: float = 0
    current_balance: float = 0
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    created_by: str
    is_deleted: bool = False
    
    def model_post_init(self, __context) -> None:
        """Validate account type after initialization"""
        if not validate_account_type(self.account_type):
            raise ValueError(
                f"Invalid account_type '{self.account_type}'. "
                f"Must be one of: {', '.join(VALID_ACCOUNT_TYPES)}"
            )

class Transaction(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    transaction_number: str
    date: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))  # Alias for created_at
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))  # ISO 8601 UTC timestamp
    transaction_type: str
    mode: str
    account_id: str
    account_name: str
    party_id: Optional[str] = None
    party_name: Optional[str] = None
    amount: float
    category: str
    notes: Optional[str] = None
    reference_type: Optional[str] = None  # "invoice", "jobcard", or None for general transactions
    reference_id: Optional[str] = None  # UUID of the related invoice/jobcard
    created_by: str
    is_deleted: bool = False

class DailyClosing(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    date: datetime
    opening_cash: Optional[float] = 0.0
    total_credit: Optional[float] = 0.0
    total_debit: Optional[float] = 0.0
    expected_closing: Optional[float] = 0.0
    actual_closing: float
    difference: Optional[float] = 0.0
    is_locked: bool = False
    closed_by: str
    notes: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class AuditLog(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    timestamp: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    user_id: str
    user_name: str
    module: str
    record_id: str
    action: str
    changes: Optional[Dict[str, Any]] = None

class GoldLedgerEntry(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    party_id: str
    date: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    type: str  # "IN" or "OUT" - IN = shop receives gold from party, OUT = shop gives gold to party
    weight_grams: float  # 3 decimal precision
    purity_entered: int
    purpose: str  # job_work | exchange | advance_gold | adjustment
    reference_type: Optional[str] = None  # invoice | jobcard | purchase | manual
    reference_id: Optional[str] = None
    notes: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    created_by: str
    is_deleted: bool = False
    deleted_at: Optional[datetime] = None
    deleted_by: Optional[str] = None

class Purchase(BaseModel):
    """Purchase model for tracking gold purchases from vendors (Stock IN + Vendor Payable)"""
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    vendor_party_id: str  # Must be a vendor type party
    date: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    description: str
    weight_grams: float  # 3 decimal precision - actual weight
    entered_purity: int  # Purity as entered/claimed by vendor (e.g., 999, 995, 916)
    valuation_purity_fixed: int = 916  # ALWAYS 916 for stock calculations and accounting
    rate_per_gram: float  # Rate per gram for 916 purity - 2 decimal precision
    amount_total: float  # Total amount = weight_grams * rate_per_gram - 2 decimal precision
    # MODULE 4: Payment and Gold Settlement fields
    paid_amount_money: float = 0.0  # Amount paid during purchase (2 decimal precision)
    balance_due_money: float = 0.0  # Auto-calculated: amount_total - paid_amount_money (2 decimal precision)
    payment_mode: Optional[str] = None  # Cash | Bank Transfer | Card | UPI | Online | Cheque
    account_id: Optional[str] = None  # Account from which payment was made
    advance_in_gold_grams: Optional[float] = None  # Gold we gave vendor previously, now used as credit (3 decimals)
    exchange_in_gold_grams: Optional[float] = None  # Gold exchanged from vendor during purchase (3 decimals)
    status: str  # MUST be calculated: "Finalized (Unpaid)" | "Partially Paid" | "Paid" | "Draft" (Draft only for future explicit save-as-draft feature)
    finalized_at: Optional[datetime] = None
    finalized_by: Optional[str] = None
    locked: bool = False  # Finalized purchases are locked
    locked_at: Optional[datetime] = None
    locked_by: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    created_by: str
    is_deleted: bool = False
    deleted_at: Optional[datetime] = None
    deleted_by: Optional[str] = None

class ReturnItem(BaseModel):
    """Item in a return (sales or purchase return)"""
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    description: str
    qty: int
    weight_grams: float  # Input as float, stored as Decimal128 with 3 decimal precision
    purity: int
    amount: float = 0.0  # Input as float, stored as Decimal128 with 2 decimal precision

class Return(BaseModel):
    """
    Return model for tracking sales returns and purchase returns.
    
    Sales Return: Customer returns items from an invoice
    - Stock IN (returned goods back to inventory)
    - Money refund → Transaction (Debit)
    - Gold refund → GoldLedgerEntry (OUT - shop gives gold to customer)
    - Update customer outstanding
    
    Purchase Return: Shop returns items to vendor from a purchase
    - Stock OUT (returned to vendor)
    - Money refund → Transaction (Credit - vendor refunds money to us)
    - Gold refund → GoldLedgerEntry (IN - vendor returns gold to us)
    - Update vendor payable
    """
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    return_number: str  # Auto-generated return number
    return_type: str  # "sale_return" or "purchase_return"
    reference_type: str  # "invoice" or "purchase"
    reference_id: str  # UUID of the related invoice or purchase
    reference_number: Optional[str] = None  # Display number of invoice/purchase
    party_id: str  # Customer (for sale_return) or Vendor (for purchase_return)
    party_name: str  # Party name for display
    party_type: str  # "customer" or "vendor"
    date: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    items: List[ReturnItem] = []
    total_weight_grams: float = 0.0  # Input as float, stored as Decimal128 with 3 decimals
    total_amount: float = 0.0  # Input as float, stored as Decimal128 with 2 decimals
    reason: Optional[str] = None  # Return reason
    
    # Refund details
    refund_mode: str  # "money" | "gold" | "mixed"
    refund_money_amount: float = 0.0  # Input as float, stored as Decimal128 with 2 decimals
    refund_gold_grams: float = 0.0  # Input as float, stored as Decimal128 with 3 decimals
    refund_gold_purity: Optional[int] = None  # Purity of gold being refunded
    
    # Payment details for money refund
    payment_mode: Optional[str] = None  # Cash | Bank Transfer | Card | UPI | Cheque
    account_id: Optional[str] = None  # Account to/from which refund is made
    account_name: Optional[str] = None  # Account name for display
    
    # Status and workflow
    status: str = "draft"  # "draft" or "finalized"
    finalized_at: Optional[datetime] = None
    finalized_by: Optional[str] = None
    
    # Related records created on finalization
    transaction_id: Optional[str] = None  # Transaction record for money refund
    gold_ledger_id: Optional[str] = None  # Gold ledger entry for gold refund
    stock_movement_ids: List[str] = []  # Stock movement records
    
    # Audit fields
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    created_by: str
    notes: Optional[str] = None
    is_deleted: bool = False
    deleted_at: Optional[datetime] = None
    deleted_by: Optional[str] = None

class ShopSettings(BaseModel):
    """Shop/Company settings for invoice printing"""
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    shop_name: str = "Gold Jewellery ERP"
    address: str = "123 Main Street, City, Country"
    phone: str = "+968 1234 5678"
    email: str = "contact@goldjewellery.com"
    gstin: str = "GST1234567890"
    logo_url: Optional[str] = None
    terms_and_conditions: str = "1. Goods once sold cannot be returned.\n2. Gold purity as per invoice.\n3. Making charges are non-refundable."
    authorized_signatory: str = "Authorized Signatory"
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_by: Optional[str] = None

async def create_audit_log(user_id: str, user_name: str, module: str, record_id: str, action: str, changes: Optional[Dict] = None):
    log = AuditLog(
        user_id=user_id,
        user_name=user_name,
        module=module,
        record_id=record_id,
        action=action,
        changes=changes
    )
    await db.audit_logs.insert_one(log.model_dump())

# ============================================================================
# AUTHENTICATION & SECURITY HELPER FUNCTIONS
# ============================================================================

def validate_password_complexity(password: str) -> tuple[bool, str]:
    """
    Validate password meets complexity requirements:
    - Minimum 12 characters
    - At least one uppercase letter
    - At least one lowercase letter
    - At least one number
    - At least one special character
    """
    if len(password) < PASSWORD_MIN_LENGTH:
        return False, f"Password must be at least {PASSWORD_MIN_LENGTH} characters long"
    
    if not any(c.isupper() for c in password):
        return False, "Password must contain at least one uppercase letter"
    
    if not any(c.islower() for c in password):
        return False, "Password must contain at least one lowercase letter"
    
    if not any(c.isdigit() for c in password):
        return False, "Password must contain at least one number"
    
    special_chars = "!@#$%^&*()_+-=[]{}|;:,.<>?"
    if not any(c in special_chars for c in password):
        return False, "Password must contain at least one special character (!@#$%^&*()_+-=[]{}|;:,.<>?)"
    
    return True, ""

async def create_auth_audit_log(username: str, action: str, success: bool, 
                                user_id: Optional[str] = None, 
                                failure_reason: Optional[str] = None,
                                ip_address: Optional[str] = None):
    """Create an authentication audit log entry"""
    log = AuthAuditLog(
        user_id=user_id,
        username=username,
        action=action,
        success=success,
        failure_reason=failure_reason,
        ip_address=ip_address
    )
    await db.auth_audit_logs.insert_one(log.model_dump())

async def check_account_lockout(user_doc: dict) -> tuple[bool, Optional[str]]:
    """
    Check if account is locked due to failed login attempts
    Returns: (is_locked, error_message)
    """
    locked_until = user_doc.get('locked_until')
    if locked_until:
        # Convert string to datetime if needed
        if isinstance(locked_until, str):
            locked_until = datetime.fromisoformat(locked_until)
        
        if datetime.now(timezone.utc) < locked_until:
            remaining = (locked_until - datetime.now(timezone.utc)).total_seconds() / 60
            return True, f"Account is locked due to multiple failed login attempts. Try again in {int(remaining)} minutes."
    
    return False, None

async def handle_failed_login(user_doc: dict, username: str):
    """Handle failed login attempt - increment counter and potentially lock account"""
    failed_attempts = user_doc.get('failed_login_attempts', 0) + 1
    
    update_data = {
        'failed_login_attempts': failed_attempts
    }
    
    # Lock account if max attempts reached
    if failed_attempts >= MAX_LOGIN_ATTEMPTS:
        lockout_time = datetime.now(timezone.utc) + timedelta(minutes=LOCKOUT_DURATION_MINUTES)
        update_data['locked_until'] = lockout_time
    
    await db.users.update_one(
        {"username": username},
        {"$set": update_data}
    )

async def handle_successful_login(user_id: str):
    """Reset failed login attempts and update last login time"""
    await db.users.update_one(
        {"id": user_id},
        {"$set": {
            'failed_login_attempts': 0,
            'locked_until': None,
            'last_login': datetime.now(timezone.utc)
        }}
    )

def get_user_permissions(role: str) -> List[str]:
    """Get permissions for a given role"""
    return ROLE_PERMISSIONS.get(role, [])

def user_has_permission(user: User, required_permission: str) -> bool:
    """Check if user has a specific permission"""
    # Admin always has all permissions
    if user.role == 'admin':
        return True
    
    # Check user's assigned permissions
    user_permissions = user.permissions if user.permissions else get_user_permissions(user.role)
    return required_permission in user_permissions

async def validate_return_against_original(
    db,
    reference_type: str,
    reference_id: str,
    reference_doc: dict,
    return_items: list,
    current_return_id: Optional[str] = None
) -> None:
    """
    Validate that return qty, weight, and amount don't exceed original invoice/purchase totals.
    Uses Decimal for precise arithmetic to avoid floating point errors.
    
    Args:
        db: MongoDB database instance
        reference_type: 'invoice' or 'purchase'
        reference_id: ID of the invoice or purchase
        reference_doc: The invoice or purchase document
        return_items: List of return items with qty, weight_grams, and amount
        current_return_id: ID of current return (for update operations, to exclude it from calculation)
    
    Raises:
        HTTPException: If return exceeds original qty, weight, or amount
    """
    # Calculate current return totals using Decimal for precision
    current_total_qty = sum(item.get('qty', 0) for item in return_items)
    
    # Convert to Decimal for precise arithmetic
    current_total_weight = Decimal('0')
    current_total_amount = Decimal('0')
    
    for item in return_items:
        weight = item.get('weight_grams', 0)
        amount = item.get('amount', 0)
        
        # Handle both Decimal128 (from DB) and float (from API input)
        if isinstance(weight, Decimal128):
            current_total_weight += weight.to_decimal()
        else:
            current_total_weight += Decimal(str(weight))
            
        if isinstance(amount, Decimal128):
            current_total_amount += amount.to_decimal()
        else:
            current_total_amount += Decimal(str(amount))
    
    # Get original totals
    if reference_type == 'invoice':
        # For invoices, calculate from items
        invoice_items = reference_doc.get('items', [])
        original_total_qty = sum(item.get('qty', 0) for item in invoice_items)
        
        # Calculate original weight
        original_total_weight = Decimal('0')
        for item in invoice_items:
            weight = item.get('net_gold_weight', 0) or item.get('weight', 0)
            if isinstance(weight, Decimal128):
                original_total_weight += weight.to_decimal()
            else:
                original_total_weight += Decimal(str(weight))
        
        # Get original amount
        grand_total = reference_doc.get('grand_total', 0)
        if isinstance(grand_total, Decimal128):
            original_total_amount = grand_total.to_decimal()
        else:
            original_total_amount = Decimal(str(grand_total))
        
        entity_name = f"Invoice {reference_doc.get('invoice_number')}"
    else:  # purchase
        # For purchases, single item structure
        original_total_qty = 1  # Purchases are typically single transaction
        
        weight = reference_doc.get('weight_grams', 0)
        if isinstance(weight, Decimal128):
            original_total_weight = weight.to_decimal()
        else:
            original_total_weight = Decimal(str(weight))
        
        amount = reference_doc.get('amount_total', 0)
        if isinstance(amount, Decimal128):
            original_total_amount = amount.to_decimal()
        else:
            original_total_amount = Decimal(str(amount))
        
        entity_name = f"Purchase {reference_id[:8]}..."
    
    # Calculate totals already returned (finalized returns only)
    query = {
        'reference_type': reference_type,
        'reference_id': reference_id,
        'status': 'finalized',
        'is_deleted': False
    }
    
    # Exclude current return if updating
    if current_return_id:
        query['id'] = {'$ne': current_return_id}
    
    existing_returns = await db.returns.find(query).to_list(length=None)
    
    # Sum up already returned quantities, weights, and amounts using Decimal
    already_returned_qty = 0
    already_returned_weight = Decimal('0')
    already_returned_amount = Decimal('0')
    
    for ret in existing_returns:
        ret_items = ret.get('items', [])
        already_returned_qty += sum(item.get('qty', 0) for item in ret_items)
        
        # Sum weights using Decimal
        for item in ret_items:
            weight = item.get('weight_grams', 0)
            if isinstance(weight, Decimal128):
                already_returned_weight += weight.to_decimal()
            else:
                already_returned_weight += Decimal(str(weight))
        
        # Sum amounts using Decimal
        total_amt = ret.get('total_amount', 0)
        if isinstance(total_amt, Decimal128):
            already_returned_amount += total_amt.to_decimal()
        else:
            already_returned_amount += Decimal(str(total_amt))
    
    # Validate quantity
    total_qty_with_new = already_returned_qty + current_total_qty
    if total_qty_with_new > original_total_qty:
        raise HTTPException(
            status_code=400,
            detail=f"Cannot return more quantity than original {entity_name}. "
                   f"Original qty: {original_total_qty}, "
                   f"Already returned: {already_returned_qty}, "
                   f"Current return: {current_total_qty}, "
                   f"Total would be: {total_qty_with_new}"
        )
    
    # Validate weight using Decimal for precision
    total_weight_with_new = already_returned_weight + current_total_weight
    # Allow 0.1% tolerance for rounding
    tolerance = original_total_weight * Decimal('0.001')
    if total_weight_with_new > (original_total_weight + tolerance):
        raise HTTPException(
            status_code=400,
            detail=f"Cannot return more weight than original {entity_name}. "
                   f"Original weight: {float(original_total_weight):.3f}g, "
                   f"Already returned: {float(already_returned_weight):.3f}g, "
                   f"Current return: {float(current_total_weight):.3f}g, "
                   f"Total would be: {float(total_weight_with_new):.3f}g"
        )
    
    # Validate amount using Decimal for precision
    total_amount_with_new = already_returned_amount + current_total_amount
    # Allow 1% tolerance for rounding
    amount_tolerance = original_total_amount * Decimal('0.01')
    if total_amount_with_new > (original_total_amount + amount_tolerance):
        raise HTTPException(
            status_code=400,
            detail=f"Cannot return more amount than original {entity_name}. "
                   f"Original: {float(original_total_amount):.2f} OMR, "
                   f"Already returned: {float(already_returned_amount):.2f} OMR, "
                   f"Current return: {float(current_total_amount):.2f} OMR, "
                   f"Total would be: {float(total_amount_with_new):.2f} OMR"
        )

# ============================================================================
# PERMISSION DECORATOR
# ============================================================================

def require_permission(permission: str):
    """
    Create a FastAPI dependency that checks for a specific permission
    Usage: 
    @api_router.get("/endpoint")
    async def endpoint(current_user: User = Depends(require_permission('permission.name'))):
        ...
    """
    async def permission_checker(current_user: User = Depends(get_current_user)) -> User:
        if not user_has_permission(current_user, permission):
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail=f"You don't have permission to perform this action. Required: {permission}"
            )
        return current_user
    return permission_checker

async def get_current_user(
    request: Request,
    credentials: Optional[HTTPAuthorizationCredentials] = Depends(security)
) -> User:
    """
    Get current user from JWT token - supports both cookie and Authorization header.
    Cookie-based auth is preferred for security (HttpOnly + Secure cookies).
    """
    # Try to get token from cookie first (preferred method)
    token = request.cookies.get("access_token")
    
    # Fallback to Authorization header for backward compatibility
    if not token and credentials:
        token = credentials.credentials
    
    if not token:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Not authenticated"
        )
    
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        user_id = payload.get('user_id')
        if not user_id:
            raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED)
        
        user_doc = await db.users.find_one({"id": user_id, "is_deleted": False}, {"_id": 0})
        if not user_doc:
            raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED)
        
        # Populate permissions based on role if not already set
        if 'permissions' not in user_doc or not user_doc['permissions']:
            user_doc['permissions'] = get_user_permissions(user_doc.get('role', 'staff'))
        
        return User(**user_doc)
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Invalid token")

@api_router.post("/auth/register", response_model=User)
@limiter.limit("5/minute")  # Strict rate limit: 5 registrations per minute per IP
async def register(request: Request, user_data: UserCreate):
    existing = await db.users.find_one({"username": user_data.username, "is_deleted": False})
    if existing:
        raise HTTPException(status_code=400, detail="Username already exists")
    
    # Validate password complexity
    is_valid, error_msg = validate_password_complexity(user_data.password)
    if not is_valid:
        raise HTTPException(status_code=400, detail=error_msg)
    
    hashed_password = pwd_context.hash(user_data.password)
    
    # Assign permissions based on role
    permissions = get_user_permissions(user_data.role)
    
    user = User(
        username=user_data.username,
        email=user_data.email,
        full_name=user_data.full_name,
        role=user_data.role,
        permissions=permissions
    )
    
    user_dict = user.model_dump()
    user_dict['hashed_password'] = hashed_password
    user_dict['created_at'] = user_dict['created_at'].isoformat()
    
    await db.users.insert_one(user_dict)
    
    # Log registration
    await create_auth_audit_log(
        username=user.username,
        action="register",
        success=True,
        user_id=user.id
    )
    
    return user

@api_router.post("/auth/login", response_model=TokenResponse)
@limiter.limit("5/minute")  # Strict rate limit: 5 login attempts per minute per IP
async def login(request: Request, credentials: UserLogin, response: Response):
    user_doc = await db.users.find_one({"username": credentials.username, "is_deleted": False}, {"_id": 0})
    
    # Check if user exists
    if not user_doc:
        await create_auth_audit_log(
            username=credentials.username,
            action="login_failed",
            success=False,
            failure_reason="User not found"
        )
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    # Check account lockout
    is_locked, lock_message = await check_account_lockout(user_doc)
    if is_locked:
        await create_auth_audit_log(
            username=credentials.username,
            action="login_failed",
            success=False,
            user_id=user_doc.get('id'),
            failure_reason="Account locked"
        )
        raise HTTPException(status_code=403, detail=lock_message)
    
    # Verify password
    if not pwd_context.verify(credentials.password, user_doc.get('hashed_password', '')):
        await handle_failed_login(user_doc, credentials.username)
        await create_auth_audit_log(
            username=credentials.username,
            action="login_failed",
            success=False,
            user_id=user_doc.get('id'),
            failure_reason="Invalid password"
        )
        
        # Check if this failure caused a lockout
        failed_attempts = user_doc.get('failed_login_attempts', 0) + 1
        if failed_attempts >= MAX_LOGIN_ATTEMPTS:
            raise HTTPException(
                status_code=403, 
                detail=f"Account locked due to {MAX_LOGIN_ATTEMPTS} failed login attempts. Please try again in {LOCKOUT_DURATION_MINUTES} minutes."
            )
        
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    # Check if user is active
    if not user_doc.get('is_active', False):
        await create_auth_audit_log(
            username=credentials.username,
            action="login_failed",
            success=False,
            user_id=user_doc.get('id'),
            failure_reason="User inactive"
        )
        raise HTTPException(status_code=403, detail="User is inactive")
    
    # Populate permissions
    if 'permissions' not in user_doc or not user_doc['permissions']:
        user_doc['permissions'] = get_user_permissions(user_doc.get('role', 'staff'))
    
    user = User(**user_doc)
    
    # Handle successful login
    await handle_successful_login(user.id)
    
    # Create JWT token
    token = jwt.encode(
        {"user_id": user.id, "exp": datetime.now(timezone.utc) + timedelta(hours=JWT_EXPIRATION_HOURS)},
        JWT_SECRET,
        algorithm=JWT_ALGORITHM
    )
    
    # Generate CSRF token for double-submit cookie pattern
    csrf_token = generate_csrf_token()
    
    # Set HttpOnly + Secure cookie for JWT (XSS protection)
    response.set_cookie(
        key="access_token",
        value=token,
        httponly=True,  # Prevents JavaScript access (XSS protection)
        secure=True,    # Only sent over HTTPS
        samesite="lax", # CSRF protection while allowing navigation
        max_age=JWT_EXPIRATION_HOURS * 3600,  # 24 hours in seconds
        path="/"        # Available to all routes
    )
    
    # Set CSRF token cookie (readable by JavaScript for double-submit pattern)
    response.set_cookie(
        key="csrf_token",
        value=csrf_token,
        httponly=False,  # Must be readable by JavaScript
        secure=True,     # Only sent over HTTPS
        samesite="lax",  # CSRF protection
        max_age=JWT_EXPIRATION_HOURS * 3600,  # Same lifetime as JWT
        path="/"
    )
    
    # Log successful login
    await create_auth_audit_log(
        username=credentials.username,
        action="login",
        success=True,
        user_id=user.id
    )
    
    return TokenResponse(access_token=token, user=user, csrf_token=csrf_token)

@api_router.get("/auth/me", response_model=User)
@limiter.limit("1000/hour")  # General authenticated rate limit: 1000 requests per hour
async def get_me(request: Request, current_user: User = Depends(get_current_user)):
    return current_user

@api_router.post("/auth/logout")
async def logout(response: Response, current_user: User = Depends(get_current_user)):
    """Logout endpoint - clears authentication and CSRF cookies, creates audit log"""
    # Clear the authentication cookie
    response.delete_cookie(
        key="access_token",
        path="/",
        samesite="lax"
    )
    
    # Clear the CSRF token cookie
    response.delete_cookie(
        key="csrf_token",
        path="/",
        samesite="lax"
    )
    
    await create_auth_audit_log(
        username=current_user.username,
        action="logout",
        success=True,
        user_id=current_user.id
    )
    return {"message": "Logged out successfully"}

@api_router.post("/auth/request-password-reset")
@limiter.limit("3/minute")  # Strict rate limit: 3 password reset requests per minute per IP
async def request_password_reset(request: Request, email_data: dict):
    """Request password reset - generates reset token"""
    email = email_data.get('email')
    if not email:
        raise HTTPException(status_code=400, detail="Email is required")
    
    user_doc = await db.users.find_one({"email": email, "is_deleted": False}, {"_id": 0})
    
    # Always return success to prevent email enumeration
    if not user_doc:
        return {"message": "If an account with this email exists, a password reset link will be sent."}
    
    # Generate reset token
    reset_token = str(uuid.uuid4())
    token_obj = PasswordResetToken(
        user_id=user_doc['id'],
        token=reset_token,
        expires_at=datetime.now(timezone.utc) + timedelta(hours=1)  # Token valid for 1 hour
    )
    
    await db.password_reset_tokens.insert_one(token_obj.model_dump())
    
    # Log password reset request
    await create_auth_audit_log(
        username=user_doc['username'],
        action="password_reset_request",
        success=True,
        user_id=user_doc['id']
    )
    
    # In production, send email with reset link containing token
    # For now, return token (in production, this should only be sent via email)
    return {
        "message": "If an account with this email exists, a password reset link will be sent.",
        "reset_token": reset_token  # REMOVE IN PRODUCTION - only for testing
    }

@api_router.post("/auth/reset-password")
@limiter.limit("3/minute")  # Strict rate limit: 3 password reset attempts per minute per IP
async def reset_password(request: Request, reset_data: dict):
    """Reset password using reset token"""
    token = reset_data.get('token')
    new_password = reset_data.get('new_password')
    
    if not token or not new_password:
        raise HTTPException(status_code=400, detail="Token and new password are required")
    
    # Find valid token
    token_doc = await db.password_reset_tokens.find_one({
        "token": token,
        "used": False
    }, {"_id": 0})
    
    if not token_doc:
        raise HTTPException(status_code=400, detail="Invalid or expired reset token")
    
    # Check token expiration
    expires_at = token_doc.get('expires_at')
    if isinstance(expires_at, str):
        expires_at = datetime.fromisoformat(expires_at)
    
    if datetime.now(timezone.utc) > expires_at:
        raise HTTPException(status_code=400, detail="Reset token has expired")
    
    # Validate password complexity
    is_valid, error_msg = validate_password_complexity(new_password)
    if not is_valid:
        raise HTTPException(status_code=400, detail=error_msg)
    
    # Update password
    user_id = token_doc.get('user_id')
    hashed_password = pwd_context.hash(new_password)
    
    await db.users.update_one(
        {"id": user_id},
        {"$set": {
            'hashed_password': hashed_password,
            'failed_login_attempts': 0,  # Reset failed attempts
            'locked_until': None  # Unlock account if locked
        }}
    )
    
    # Mark token as used
    await db.password_reset_tokens.update_one(
        {"token": token},
        {"$set": {
            'used': True,
            'used_at': datetime.now(timezone.utc)
        }}
    )
    
    # Get user for audit log
    user_doc = await db.users.find_one({"id": user_id}, {"_id": 0})
    
    # Log password reset
    await create_auth_audit_log(
        username=user_doc['username'],
        action="password_reset",
        success=True,
        user_id=user_id
    )
    
    return {"message": "Password reset successfully"}

@api_router.get("/users", response_model=List[User])
@limiter.limit("1000/hour")  # General authenticated rate limit: 1000 requests per hour
async def get_users(request: Request, current_user: User = Depends(require_permission('users.view'))):
    users = await db.users.find({"is_deleted": False}, {"_id": 0, "hashed_password": 0}).to_list(1000)
    for user in users:
        if isinstance(user.get('created_at'), str):
            user['created_at'] = datetime.fromisoformat(user['created_at'])
        # Populate permissions if not set
        if 'permissions' not in user or not user['permissions']:
            user['permissions'] = get_user_permissions(user.get('role', 'staff'))
    return users

@api_router.patch("/users/{user_id}")
@limiter.limit("30/minute")  # Sensitive operation: 30 user updates per minute
async def update_user(request: Request, user_id: str, update_data: dict, current_user: User = Depends(require_permission('users.update'))):
    existing = await db.users.find_one({"id": user_id, "is_deleted": False})
    if not existing:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Don't allow password updates through this endpoint
    if 'password' in update_data:
        del update_data['password']
    if 'hashed_password' in update_data:
        del update_data['hashed_password']
    
    # If role is being updated, update permissions accordingly
    if 'role' in update_data:
        update_data['permissions'] = get_user_permissions(update_data['role'])
    
    await db.users.update_one({"id": user_id}, {"$set": update_data})
    await create_audit_log(current_user.id, current_user.full_name, "user", user_id, "update", update_data)
    return {"message": "User updated successfully"}

@api_router.delete("/users/{user_id}")
@limiter.limit("30/minute")  # Sensitive operation: 30 user deletions per minute
async def delete_user(request: Request, user_id: str, current_user: User = Depends(require_permission('users.delete'))):
    if user_id == current_user.id:
        raise HTTPException(status_code=400, detail="Cannot delete your own account")
    
    existing = await db.users.find_one({"id": user_id, "is_deleted": False})
    if not existing:
        raise HTTPException(status_code=404, detail="User not found")
    
    await db.users.update_one(
        {"id": user_id},
        {"$set": {"is_deleted": True, "deleted_at": datetime.now(timezone.utc), "deleted_by": current_user.id}}
    )
    await create_audit_log(current_user.id, current_user.full_name, "user", user_id, "delete")
    return {"message": "User deleted successfully"}

@api_router.get("/auth/audit-logs")
@limiter.limit("50/minute")  # Audit logs rate limit: 50 requests per minute
async def get_auth_audit_logs(
    request: Request,
    limit: int = 100,
    skip: int = 0,
    current_user: User = Depends(require_permission('audit.view'))
):
    """Get authentication audit logs - admin only"""
    logs = await db.auth_audit_logs.find({}, {"_id": 0}).sort("timestamp", -1).skip(skip).limit(limit).to_list(limit)
    total_count = await db.auth_audit_logs.count_documents({})
    
    return {
        "logs": logs,
        "total": total_count,
        "limit": limit,
        "skip": skip
    }

@api_router.get("/permissions")
async def get_available_permissions(current_user: User = Depends(get_current_user)):
    """Get all available permissions in the system - admin only"""
    if current_user.role != 'admin':
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only admins can view available permissions"
        )
    
    return {
        "permissions": PERMISSIONS,
        "role_mappings": ROLE_PERMISSIONS
    }

@api_router.post("/users/{user_id}/change-password")
async def change_password(user_id: str, password_data: dict, current_user: User = Depends(get_current_user)):
    # Users can change their own password, admins can change anyone's
    if user_id != current_user.id and current_user.role != 'admin':
        raise HTTPException(status_code=403, detail="Not authorized")
    
    existing = await db.users.find_one({"id": user_id, "is_deleted": False})
    if not existing:
        raise HTTPException(status_code=404, detail="User not found")
    
    new_password = password_data.get('new_password')
    if not new_password:
        raise HTTPException(status_code=400, detail="New password is required")
    
    # Validate password complexity
    is_valid, error_msg = validate_password_complexity(new_password)
    if not is_valid:
        raise HTTPException(status_code=400, detail=error_msg)
    
    hashed_password = pwd_context.hash(new_password)
    await db.users.update_one({"id": user_id}, {"$set": {"hashed_password": hashed_password}})
    await create_audit_log(current_user.id, current_user.full_name, "user", user_id, "password_change")
    
    # Log password change
    await create_auth_audit_log(
        username=existing['username'],
        action="password_change",
        success=True,
        user_id=user_id
    )
    
    return {"message": "Password changed successfully"}

@api_router.get("/inventory/headers")
async def get_inventory_headers(
    page: int = 1,
    page_size: int = 10,
    current_user: User = Depends(require_permission('inventory.view'))
):
    """Get inventory headers with pagination support"""
    query = {"is_deleted": False}
    
    # Calculate skip value
    skip = (page - 1) * page_size
    
    # Get total count for pagination
    total_count = await db.inventory_headers.count_documents(query)
    
    # Get paginated results
    headers = await db.inventory_headers.find(query, {"_id": 0}).skip(skip).limit(page_size).to_list(page_size)
    
    return create_pagination_response(headers, total_count, page, page_size)

@api_router.post("/inventory/headers", response_model=InventoryHeader)
async def create_inventory_header(header_data: dict, current_user: User = Depends(require_permission('inventory.adjust'))):
    # Validate and sanitize category name
    category_name = header_data['name'].strip()
    
    # Length validation: 3-50 characters
    if len(category_name) < 3:
        raise HTTPException(
            status_code=400,
            detail="Category name must be at least 3 characters long."
        )
    
    if len(category_name) > 50:
        raise HTTPException(
            status_code=400,
            detail="Category name must not exceed 50 characters."
        )
    
    # Character validation: only letters, numbers, and spaces allowed
    import re
    if not re.match(r'^[A-Za-z0-9 ]+$', category_name):
        raise HTTPException(
            status_code=400,
            detail="Category name can only contain letters, numbers, and spaces. Special characters are not allowed."
        )
    
    # Check for duplicate category name (case-insensitive)
    existing_category = await db.inventory_headers.find_one({
        "name": {"$regex": f"^{category_name}$", "$options": "i"},
        "is_deleted": False
    })
    
    if existing_category:
        raise HTTPException(
            status_code=400, 
            detail=f"Category '{category_name}' already exists. Please use a different name."
        )
    
    header = InventoryHeader(name=category_name, created_by=current_user.id)
    await db.inventory_headers.insert_one(header.model_dump())
    await create_audit_log(current_user.id, current_user.full_name, "inventory_header", header.id, "create")
    return header

@api_router.patch("/inventory/headers/{header_id}", response_model=InventoryHeader)
async def update_inventory_header(
    header_id: str, 
    header_data: dict, 
    current_user: User = Depends(require_permission('inventory.adjust'))
):
    """
    Update an existing inventory header (category name)
    Note: current_qty and current_weight are managed through stock movements
    """
    # Find existing header
    existing_header = await db.inventory_headers.find_one({"id": header_id, "is_deleted": False}, {"_id": 0})
    if not existing_header:
        raise HTTPException(status_code=404, detail="Inventory header not found")
    
   # Prepare update data - only allow updating name and is_active
    update_data = {}
    if 'name' in header_data:
        new_name = header_data['name'].strip()
        
        # Length validation: 3-50 characters
        if len(new_name) < 3:
            raise HTTPException(
                status_code=400,
                detail="Category name must be at least 3 characters long."
            )
        
        if len(new_name) > 50:
            raise HTTPException(
                status_code=400,
                detail="Category name must not exceed 50 characters."
            )
        
        # Character validation: only letters, numbers, and spaces allowed
        import re
        if not re.match(r'^[A-Za-z0-9 ]+$', new_name):
            raise HTTPException(
                status_code=400,
                detail="Category name can only contain letters, numbers, and spaces. Special characters are not allowed."
            )
        
        # Check for duplicate category name (case-insensitive), excluding current header
        duplicate_category = await db.inventory_headers.find_one({
            "name": {"$regex": f"^{new_name}$", "$options": "i"},
            "is_deleted": False,
            "id": {"$ne": header_id}  # Exclude current header
        })
        
        if duplicate_category:
            raise HTTPException(
                status_code=400, 
                detail=f"Category '{new_name}' already exists. Please use a different name."
            )
        
        update_data['name'] = new_name
    if 'is_active' in header_data:
        update_data['is_active'] = header_data['is_active']
    
    if not update_data:
        raise HTTPException(status_code=400, detail="No valid fields to update")
    
    # Update the header
    await db.inventory_headers.update_one(
        {"id": header_id},
        {"$set": update_data}
    )
    
    # Create audit log
    await create_audit_log(
        current_user.id, 
        current_user.full_name, 
        "inventory_header", 
        header_id, 
        "update",
        changes={"changes": update_data}
    )
    
    # Fetch and return updated header
    updated_header = await db.inventory_headers.find_one({"id": header_id}, {"_id": 0})
    return InventoryHeader(**updated_header)

@api_router.delete("/inventory/headers/{header_id}")
async def delete_inventory_header(header_id: str, current_user: User = Depends(require_permission('inventory.adjust'))):
    """
    Soft delete an inventory header
    Note: This will not affect existing stock movements (audit trail preserved)
    """
    # Find existing header
    existing_header = await db.inventory_headers.find_one({"id": header_id, "is_deleted": False}, {"_id": 0})
    if not existing_header:
        raise HTTPException(status_code=404, detail="Inventory header not found")
    
    # Check if header has current stock
    if existing_header.get('current_qty', 0) > 0 or existing_header.get('current_weight', 0) > 0:
        raise HTTPException(
            status_code=400, 
            detail=f"Cannot delete inventory header with existing stock. Current: {existing_header.get('current_qty', 0)} qty, {existing_header.get('current_weight', 0)}g"
        )
    
    # Soft delete the header
    await db.inventory_headers.update_one(
        {"id": header_id},
        {"$set": {"is_deleted": True}}
    )
    
    # Create audit log
    await create_audit_log(
        current_user.id, 
        current_user.full_name, 
        "inventory_header", 
        header_id, 
        "delete",
        changes={"header_name": existing_header.get('name')}
    )
    
    return {"message": "Inventory header deleted successfully", "id": header_id}

@api_router.get("/inventory/movements", response_model=List[StockMovement])
async def get_stock_movements(header_id: Optional[str] = None, current_user: User = Depends(require_permission('inventory.view'))):
    if not user_has_permission(current_user, 'inventory.view'):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="You don't have permission to view inventory")
    
    query = {"is_deleted": False}
    if header_id:
        query['header_id'] = header_id
    movements = await db.stock_movements.find(query, {"_id": 0}).sort("date", -1).to_list(1000)
    return movements

@api_router.post("/inventory/movements", response_model=StockMovement)
async def create_stock_movement(movement_data: dict, current_user: User = Depends(require_permission('inventory.adjust'))):
    """
    Create manual stock movement for inventory adjustments.
    
    CRITICAL RESTRICTION - Production ERP Compliance:
    - Stock OUT movements are PROHIBITED via manual entry
    - Stock can only be reduced through Invoice Finalization (POST /api/invoices/{id}/finalize)
    - This ensures proper audit trail, accounting accuracy, and GST compliance
    
    Allowed movement types:
    - "Stock IN": Manual stock additions (returns, found items, corrections)
    - "Adjustment": Inventory reconciliation adjustments (positive only)
    
    Prohibited:
    - "Stock OUT": Must occur ONLY through invoice finalization to maintain:
      * Complete audit trail (tied to invoice)
      * Accurate accounting (revenue recorded)
      * GST compliance (tax collected)
      
    WORKFLOW CONTROL:
    - confirmation_reason is REQUIRED for all manual adjustments
    """
    header = await db.inventory_headers.find_one({"id": movement_data['header_id']}, {"_id": 0})
    if not header:
        raise HTTPException(status_code=404, detail="Header not found")
    
    # CRITICAL VALIDATION: Prevent manual Stock OUT movements
    movement_type = movement_data.get('movement_type', '').strip()
    qty_delta = movement_data.get('qty_delta', 0)
    weight_delta = movement_data.get('weight_delta', 0)
    
    # WORKFLOW CONTROL: Require confirmation_reason for manual adjustments
    confirmation_reason = movement_data.get('confirmation_reason', '').strip()
    if not confirmation_reason:
        raise HTTPException(
            status_code=400,
            detail="confirmation_reason is required for all manual inventory adjustments. Please provide a reason for this stock movement."
        )
    
    # Block Stock OUT movement type entirely
    if movement_type == "Stock OUT":
        raise HTTPException(
            status_code=403,
            detail="Manual 'Stock OUT' movements are prohibited. Stock can only be reduced through Invoice Finalization (POST /api/invoices/{id}/finalize). This restriction ensures audit trail integrity, accounting accuracy, and GST compliance."
        )
    
    # Block negative deltas for Stock IN and Adjustment (attempt to bypass via negative values)
    if movement_type in ["Stock IN", "Adjustment"]:
        if qty_delta < 0 or weight_delta < 0:
            raise HTTPException(
                status_code=400,
                detail=f"Invalid {movement_type} movement: qty_delta and weight_delta must be positive (>= 0). To reduce stock, use Invoice Finalization instead."
            )
    
    # Validate movement_type is one of the allowed types
    allowed_types = ["Stock IN", "Adjustment"]
    if movement_type not in allowed_types:
        raise HTTPException(
            status_code=400,
            detail=f"Invalid movement_type '{movement_type}'. Allowed types: {', '.join(allowed_types)}. Note: 'Stock OUT' is only created automatically through Invoice Finalization."
        )
    
    movement = StockMovement(
        movement_type=movement_type,
        header_id=movement_data['header_id'],
        header_name=header['name'],
        description=movement_data['description'],
        qty_delta=qty_delta,
        weight_delta=weight_delta,
        purity=movement_data['purity'],
        notes=movement_data.get('notes'),
        confirmation_reason=confirmation_reason,
        created_by=current_user.id
    )
    
    # Insert stock movement for audit trail
    await db.stock_movements.insert_one(movement.model_dump())
    
    # Create audit log for manual inventory adjustment
    await create_audit_log(
        current_user.id,
        current_user.full_name,
        "inventory",
        movement.id,
        f"manual_{movement_type.lower().replace(' ', '_')}",
        {
            "movement_type": movement_type,
            "header_name": header['name'],
            "qty_delta": qty_delta,
            "weight_delta": weight_delta,
            "confirmation_reason": confirmation_reason
        }
    )
    
    # DIRECT UPDATE: Update header's current quantity and weight
    new_qty = header.get('current_qty', 0) + qty_delta
    new_weight = header.get('current_weight', 0) + weight_delta
    
    # Validate stock doesn't go negative (safety check, should not happen with positive deltas)
    if new_qty < 0 or new_weight < 0:
        # Delete the movement we just created
        await db.stock_movements.delete_one({"id": movement.id})
        raise HTTPException(
            status_code=400, 
            detail=f"Insufficient stock. Available: {header.get('current_qty', 0)} qty, {header.get('current_weight', 0)}g. Requested: {abs(qty_delta)} qty, {abs(weight_delta)}g"
        )
    
    await db.inventory_headers.update_one(
        {"id": movement_data['header_id']},
        {"$set": {"current_qty": new_qty, "current_weight": new_weight}}
    )
    
    await create_audit_log(current_user.id, current_user.full_name, "stock_movement", movement.id, "create", 
                          changes={"movement_type": movement_type, "qty_delta": qty_delta, "weight_delta": weight_delta})
    return movement

@api_router.delete("/inventory/movements/{movement_id}")
async def delete_stock_movement(movement_id: str, current_user: User = Depends(require_permission('inventory.adjust'))):
    """
    Delete a stock movement and reverse its effect on inventory.
    
    CRITICAL RESTRICTIONS - Production ERP Compliance:
    - CANNOT delete movements linked to invoices or purchases (reference_type set)
    - CANNOT delete Stock OUT movements (these are created only by invoice finalization)
    - CAN ONLY delete manual Stock IN or Adjustment movements
    - This ensures audit trail integrity and prevents financial record corruption
    
    WARNING: This should only be used for corrections of manual entry errors.
    """
    # Find the movement
    movement = await db.stock_movements.find_one({"id": movement_id, "is_deleted": False}, {"_id": 0})
    if not movement:
        raise HTTPException(status_code=404, detail="Stock movement not found")
    
    # CRITICAL VALIDATION 1: Check if linked to invoice or purchase
    if movement.get('reference_type') in ['invoice', 'purchase']:
        raise HTTPException(
            status_code=403,
            detail=f"Cannot delete stock movement linked to {movement.get('reference_type')}. This movement is part of an official transaction and must be preserved for audit trail, accounting accuracy, and GST compliance."
        )
    
    # CRITICAL VALIDATION 2: Prevent deletion of Stock OUT movements
    # Stock OUT should NEVER be created manually, but if somehow exists without reference_type, block deletion
    if movement.get('movement_type') == "Stock OUT":
        raise HTTPException(
            status_code=403,
            detail="Cannot delete 'Stock OUT' movement. Stock OUT movements represent sales/reductions and must be preserved for audit trail. If this movement was created in error, contact system administrator."
        )
    
    # Get the header to reverse the stock changes
    header = await db.inventory_headers.find_one({"id": movement['header_id']}, {"_id": 0})
    if not header:
        raise HTTPException(status_code=404, detail="Inventory header not found")
    
    # Calculate reversed values
    new_qty = header.get('current_qty', 0) - movement['qty_delta']
    new_weight = header.get('current_weight', 0) - movement['weight_delta']
    
    # Validate reversal won't make stock negative
    if new_qty < 0 or new_weight < 0:
        raise HTTPException(
            status_code=400,
            detail=f"Cannot delete movement: would result in negative stock. Current: {header.get('current_qty', 0)} qty, {header.get('current_weight', 0)}g. Movement: {movement['qty_delta']} qty, {movement['weight_delta']}g"
        )
    
    # Soft delete the movement
    await db.stock_movements.update_one(
        {"id": movement_id},
        {"$set": {"is_deleted": True}}
    )
    
    # Reverse the stock change in header
    await db.inventory_headers.update_one(
        {"id": movement['header_id']},
        {"$set": {"current_qty": new_qty, "current_weight": new_weight}}
    )
    
    # Create audit log
    await create_audit_log(
        current_user.id,
        current_user.full_name,
        "stock_movement",
        movement_id,
        "delete",
        changes={
            "movement_type": movement.get('movement_type'),
            "qty_delta": movement.get('qty_delta'),
            "weight_delta": movement.get('weight_delta'),
            "header_id": movement.get('header_id'),
            "header_name": movement.get('header_name')
        }
    )
    
    return {
        "message": "Stock movement deleted successfully and inventory adjusted",
        "id": movement_id,
        "reversed_qty": movement['qty_delta'],
        "reversed_weight": movement['weight_delta']
    }

@api_router.get("/inventory/stock-totals")
async def get_stock_totals(current_user: User = Depends(require_permission('inventory.view'))):
    if not user_has_permission(current_user, 'inventory.view'):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="You don't have permission to view inventory")
    
    # Return current stock directly from inventory headers
    headers = await db.inventory_headers.find({"is_deleted": False}, {"_id": 0}).to_list(1000)
    return [
        {
            "header_id": h['id'], 
            "header_name": h['name'], 
            "total_qty": h.get('current_qty', 0), 
            "total_weight": h.get('current_weight', 0)
        } 
        for h in headers
    ]

# ============================================================================
# NEW ENDPOINTS FOR API COMPLETENESS
# ============================================================================

@api_router.get("/dashboard")
async def get_dashboard(current_user: User = Depends(require_permission('reports.view'))):
    """
    Dashboard endpoint - Returns pre-aggregated statistics
    Combines data from multiple endpoints for convenience
    """
    try:
        # Get inventory headers count and totals
        headers = await db.inventory_headers.find({"is_deleted": False}, {"_id": 0}).to_list(1000)
        total_headers = len(headers)
        total_stock_weight = sum(h.get('current_weight', 0) for h in headers)
        total_stock_qty = sum(h.get('current_qty', 0) for h in headers)
        
        # Get outstanding summary
        invoices = await db.invoices.find(
            {"is_deleted": False, "payment_status": {"$ne": "paid"}}, 
            {"_id": 0}
        ).to_list(10000)
        total_outstanding = sum(inv.get('balance_due', 0) for inv in invoices)
        
        # Get low stock items (qty < 5)
        low_stock_items = len([h for h in headers if h.get('current_qty', 0) < 5])
        
        # Get parties count
        customers_count = await db.parties.count_documents({"is_deleted": False, "party_type": "customer"})
        vendors_count = await db.parties.count_documents({"is_deleted": False, "party_type": "vendor"})
        
        # Get recent invoices (last 5)
        recent_invoices = await db.invoices.find(
            {"is_deleted": False}, 
            {"_id": 0}
        ).sort("created_at", -1).limit(5).to_list(5)
        
        # Get job cards count by status
        total_jobcards = await db.jobcards.count_documents({"is_deleted": False})
        pending_jobcards = await db.jobcards.count_documents({"is_deleted": False, "status": "pending"})
        completed_jobcards = await db.jobcards.count_documents({"is_deleted": False, "status": "completed"})
        
        return {
            "inventory": {
                "total_categories": total_headers,
                "total_stock_weight_grams": round(total_stock_weight, 3),
                "total_stock_qty": round(total_stock_qty, 2),
                "low_stock_items": low_stock_items
            },
            "financial": {
                "total_outstanding_omr": round(total_outstanding, 2),
                "outstanding_invoices_count": len(invoices)
            },
            "parties": {
                "total_customers": customers_count,
                "total_vendors": vendors_count,
                "total": customers_count + vendors_count
            },
            "job_cards": {
                "total": total_jobcards,
                "pending": pending_jobcards,
                "completed": completed_jobcards
            },
            "recent_activity": {
                "recent_invoices": recent_invoices[:5]
            },
            "timestamp": datetime.now(timezone.utc).isoformat()
        }
    except Exception as e:
        logging.error(f"Dashboard error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to load dashboard: {str(e)}")

@api_router.get("/reports")
async def get_reports_list(current_user: User = Depends(require_permission('reports.view'))):
    """
    Reports listing endpoint - Returns available report types with metadata
    """
    reports = [
        {
            "id": "financial-summary",
            "name": "Financial Summary",
            "description": "Overview of sales, purchases, profit, and outstanding balances",
            "category": "financial",
            "endpoints": {
                "view": "/api/reports/financial-summary"
            },
            "supports_filters": False,
            "supports_export": False
        },
        {
            "id": "inventory",
            "name": "Inventory Report",
            "description": "Complete inventory stock levels and movements",
            "category": "inventory",
            "endpoints": {
                "view": "/api/reports/inventory-view",
                "export_excel": "/api/reports/inventory-export",
                "export_pdf": "/api/reports/inventory-pdf"
            },
            "supports_filters": True,
            "supports_export": True,
            "export_formats": ["excel", "pdf"]
        },
        {
            "id": "parties",
            "name": "Parties Report",
            "description": "List of all customers and vendors with contact details",
            "category": "parties",
            "endpoints": {
                "view": "/api/reports/parties-view",
                "export_excel": "/api/reports/parties-export",
                "export_pdf": "/api/reports/parties-pdf"
            },
            "supports_filters": True,
            "supports_export": True,
            "export_formats": ["excel", "pdf"]
        },
        {
            "id": "invoices",
            "name": "Invoices Report",
            "description": "All invoices with payment status and details",
            "category": "sales",
            "endpoints": {
                "view": "/api/reports/invoices-view",
                "export_excel": "/api/reports/invoices-export",
                "export_pdf": "/api/reports/invoices-pdf"
            },
            "supports_filters": True,
            "supports_export": True,
            "export_formats": ["excel", "pdf"]
        },
        {
            "id": "transactions",
            "name": "Transactions Report",
            "description": "All financial transactions and payments",
            "category": "financial",
            "endpoints": {
                "view": "/api/reports/transactions-view",
                "export_excel": "/api/reports/transactions-export",
                "export_pdf": "/api/reports/transactions-pdf"
            },
            "supports_filters": True,
            "supports_export": True,
            "export_formats": ["excel", "pdf"]
        },
        {
            "id": "outstanding",
            "name": "Outstanding Report",
            "description": "Customers with outstanding balances and aging analysis",
            "category": "financial",
            "endpoints": {
                "view": "/api/reports/outstanding",
                "export_excel": "/api/reports/outstanding-export",
                "export_pdf": "/api/reports/outstanding-pdf"
            },
            "supports_filters": True,
            "supports_export": True,
            "export_formats": ["excel", "pdf"]
        },
        {
            "id": "sales-history",
            "name": "Sales History",
            "description": "Historical sales data and trends",
            "category": "sales",
            "endpoints": {
                "view": "/api/reports/sales-history",
                "export_excel": "/api/reports/sales-history-export",
                "export_pdf": "/api/reports/sales-history-pdf"
            },
            "supports_filters": True,
            "supports_export": True,
            "export_formats": ["excel", "pdf"]
        },
        {
            "id": "purchase-history",
            "name": "Purchase History",
            "description": "Historical purchase data from vendors",
            "category": "purchases",
            "endpoints": {
                "view": "/api/reports/purchase-history",
                "export_excel": "/api/reports/purchase-history-export",
                "export_pdf": "/api/reports/purchase-history-pdf"
            },
            "supports_filters": True,
            "supports_export": True,
            "export_formats": ["excel", "pdf"]
        },
        {
            "id": "returns",
            "name": "Returns Report",
            "description": "Sales and purchase returns with financial impact analysis",
            "category": "returns",
            "endpoints": {
                "view": "/api/reports/returns-summary",
                "export_excel": "/api/reports/returns-export",
                "export_pdf": "/api/reports/returns-pdf"
            },
            "supports_filters": True,
            "supports_export": True,
            "export_formats": ["excel", "pdf"]
        }
    ]
    
    return {
        "reports": reports,
        "total_count": len(reports),
        "categories": ["financial", "inventory", "parties", "sales", "purchases", "returns"],
        "timestamp": datetime.now(timezone.utc).isoformat()
    }

@api_router.get("/inventory")
async def get_inventory(
    category: Optional[str] = None,
    min_qty: Optional[float] = None,
    page: int = 1,
    page_size: int = 10,
    current_user: User = Depends(require_permission('inventory.view'))
):
    """
    Basic inventory listing endpoint with pagination support
    Provides a simple interface for inventory data retrieval
    """
    try:
        # Build query
        query = {"is_deleted": False}
        if category:
            query['name'] = {"$regex": category, "$options": "i"}
        
        # Get all headers for filtering (if min_qty is used)
        if min_qty is not None:
            # When filtering by quantity, we need to get all items first
            headers = await db.inventory_headers.find(query, {"_id": 0}).to_list(1000)
            headers = [h for h in headers if h.get('current_qty', 0) >= min_qty]
            
            # Format response with additional computed fields
            inventory_items = []
            for header in headers:
                item = {
                    "id": header.get('id'),
                    "category": header.get('name'),
                    "quantity": round(header.get('current_qty', 0), 2),
                    "weight_grams": round(header.get('current_weight', 0), 3),
                    "is_active": header.get('is_active', True),
                    "created_at": header.get('created_at'),
                    "created_by": header.get('created_by'),
                    "status": "low_stock" if header.get('current_qty', 0) < 5 else "in_stock"
                }
                inventory_items.append(item)
            
            # Sort by weight descending
            inventory_items.sort(key=lambda x: x['weight_grams'], reverse=True)
            
            # Apply pagination to filtered results
            total_count = len(inventory_items)
            skip = (page - 1) * page_size
            inventory_items = inventory_items[skip:skip + page_size]
        else:
            # Efficient pagination without min_qty filter
            # Calculate skip value
            skip = (page - 1) * page_size
            
            # Get total count for pagination
            total_count = await db.inventory_headers.count_documents(query)
            
            # Get paginated results sorted by current_weight descending
            headers = await db.inventory_headers.find(query, {"_id": 0}).sort("current_weight", -1).skip(skip).limit(page_size).to_list(page_size)
            
            # Format response with additional computed fields
            inventory_items = []
            for header in headers:
                item = {
                    "id": header.get('id'),
                    "category": header.get('name'),
                    "quantity": round(header.get('current_qty', 0), 2),
                    "weight_grams": round(header.get('current_weight', 0), 3),
                    "is_active": header.get('is_active', True),
                    "created_at": header.get('created_at'),
                    "created_by": header.get('created_by'),
                    "status": "low_stock" if header.get('current_qty', 0) < 5 else "in_stock"
                }
                inventory_items.append(item)
        
        return create_pagination_response(inventory_items, total_count, page, page_size)
    except Exception as e:
        logging.error(f"Inventory listing error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to load inventory: {str(e)}")

# ============================================================================
# END OF NEW ENDPOINTS
# ============================================================================

@api_router.get("/parties")
@limiter.limit("1000/hour")  # General authenticated rate limit: 1000 requests per hour
async def get_parties(
    request: Request,
    party_type: Optional[str] = None,
    page: int = 1,
    page_size: int = 10,
    current_user: User = Depends(require_permission('parties.view'))
):
    """Get parties with pagination support"""
    query = {"is_deleted": False}
    if party_type:
        query['party_type'] = party_type
    
    # Calculate skip value
    skip = (page - 1) * page_size
    
    # Get total count for pagination
    total_count = await db.parties.count_documents(query)
    
    # Get paginated results
    parties = await db.parties.find(query, {"_id": 0}).skip(skip).limit(page_size).to_list(page_size)
    
    return create_pagination_response(parties, total_count, page, page_size)

@api_router.post("/parties", response_model=Party)
@limiter.limit("1000/hour")  # General authenticated rate limit: 1000 requests per hour
async def create_party(request: Request, party_data: dict, current_user: User = Depends(require_permission('parties.create'))):
    if not user_has_permission(current_user, 'parties.create'):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="You don't have permission to create parties")
    
    # Validate input data using PartyValidator
    try:
        validated_data = PartyValidator(**party_data)
    except Exception as e:
        raise HTTPException(
            status_code=400,
            detail=str(e)
        )
    
    # Validate duplicate phone number
    phone = validated_data.phone
    if phone and phone.strip():  # Only check if phone is provided and not empty
        existing_phone = await db.parties.find_one({
            "phone": phone.strip(),
            "is_deleted": False
        })
        if existing_phone:
            raise HTTPException(
                status_code=400,
                detail=f"Phone number {phone} is already registered with another party: {existing_phone.get('name', 'Unknown')}"
            )
    
    party = Party(**validated_data.dict(), created_by=current_user.id)
    await db.parties.insert_one(party.model_dump())
    await create_audit_log(current_user.id, current_user.full_name, "party", party.id, "create")
    return party

@api_router.get("/parties/outstanding-summary")
async def get_outstanding_summary(current_user: User = Depends(require_permission('parties.view'))):
    invoices = await db.invoices.find({"is_deleted": False, "payment_status": {"$ne": "paid"}}, {"_id": 0}).to_list(10000)
    
    total_customer_due = sum(inv.get('balance_due', 0) for inv in invoices)
    
    party_outstanding = {}
    for inv in invoices:
        cid = inv.get('customer_id')
        if cid:
            if cid not in party_outstanding:
                party_outstanding[cid] = {"customer_id": cid, "customer_name": inv.get('customer_name', ''), "outstanding": 0}
            party_outstanding[cid]['outstanding'] += inv.get('balance_due', 0)
    
    top_10 = sorted(party_outstanding.values(), key=lambda x: x['outstanding'], reverse=True)[:10]
    
    return {"total_customer_due": total_customer_due, "top_10_outstanding": top_10}

@api_router.get("/parties/{party_id}", response_model=Party)
async def get_party(party_id: str, current_user: User = Depends(require_permission('parties.view'))):
    party = await db.parties.find_one({"id": party_id, "is_deleted": False}, {"_id": 0})
    if not party:
        raise HTTPException(status_code=404, detail="Party not found")
    return Party(**party)

@api_router.patch("/parties/{party_id}", response_model=Party)
async def update_party(party_id: str, party_data: dict, current_user: User = Depends(require_permission('parties.update'))):
    if not user_has_permission(current_user, 'parties.update'):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="You don't have permission to update parties")
    
    existing = await db.parties.find_one({"id": party_id, "is_deleted": False})
    if not existing:
        raise HTTPException(status_code=404, detail="Party not found")
    
    # Validate input data using PartyValidator
    # Create a copy with existing values as defaults for partial updates
    update_data = {
        'name': party_data.get('name', existing.get('name')),
        'phone': party_data.get('phone', existing.get('phone')),
        'address': party_data.get('address', existing.get('address')),
        'party_type': party_data.get('party_type', existing.get('party_type')),
        'notes': party_data.get('notes', existing.get('notes'))
    }
    
    try:
        validated_data = PartyValidator(**update_data)
    except Exception as e:
        raise HTTPException(
            status_code=400,
            detail=str(e)
        )
    
    # Validate duplicate phone number if phone is being updated
    phone = validated_data.phone
    if phone and phone.strip():  # Only check if phone is provided and not empty
        existing_phone = await db.parties.find_one({
            "phone": phone.strip(),
            "is_deleted": False,
            "id": {"$ne": party_id}  # Exclude current party
        })
        if existing_phone:
            raise HTTPException(
                status_code=400,
                detail=f"Phone number {phone} is already registered with another party: {existing_phone.get('name', 'Unknown')}"
            )
    
    # Update with validated data
    update_dict = validated_data.dict(exclude_unset=False)
    await db.parties.update_one({"id": party_id}, {"$set": update_dict})
    await create_audit_log(current_user.id, current_user.full_name, "party", party_id, "update", update_dict)
    
    updated = await db.parties.find_one({"id": party_id}, {"_id": 0})
    return Party(**updated)

@api_router.get("/parties/{party_id}/impact")
async def get_party_impact(party_id: str, current_user: User = Depends(require_permission('parties.view'))):
    """
    Get impact summary for party deletion.
    Shows what data is linked to this party and will be affected.
    """
    party = await db.parties.find_one({"id": party_id, "is_deleted": False}, {"_id": 0})
    if not party:
        raise HTTPException(status_code=404, detail="Party not found")
    
    # Count linked records
    linked_invoices_count = await db.invoices.count_documents({"customer_id": party_id, "is_deleted": False})
    linked_purchases_count = await db.purchases.count_documents({"vendor_party_id": party_id, "is_deleted": False})
    linked_jobcards_count = await db.jobcards.count_documents({"customer_id": party_id, "is_deleted": False})
    linked_gold_ledger_count = await db.gold_ledger.count_documents({"party_id": party_id, "is_deleted": False})
    linked_transactions_count = await db.transactions.count_documents({"party_id": party_id, "is_deleted": False})
    
    # Get outstanding balances
    money_outstanding = 0
    invoices = await db.invoices.find({"customer_id": party_id, "is_deleted": False, "status": "finalized"}, {"_id": 0, "balance_due": 1}).to_list(1000)
    for inv in invoices:
        money_outstanding += inv.get('balance_due', 0)
    
    # Get gold balance
    gold_entries = await db.gold_ledger.find({"party_id": party_id, "is_deleted": False}, {"_id": 0, "type": 1, "weight_grams": 1}).to_list(1000)
    gold_balance = 0
    for entry in gold_entries:
        if entry.get('type') == 'IN':
            gold_balance += entry.get('weight_grams', 0)
        else:
            gold_balance -= entry.get('weight_grams', 0)
    
    impact = {
        "party_name": party.get("name"),
        "party_type": party.get("party_type"),
        "phone": party.get("phone"),
        "linked_invoices": linked_invoices_count,
        "linked_purchases": linked_purchases_count,
        "linked_jobcards": linked_jobcards_count,
        "linked_gold_ledger": linked_gold_ledger_count,
        "linked_transactions": linked_transactions_count,
        "total_linked_records": (linked_invoices_count + linked_purchases_count + 
                                linked_jobcards_count + linked_gold_ledger_count + 
                                linked_transactions_count),
        "money_outstanding": round(money_outstanding, 2),
        "gold_balance_grams": round(gold_balance, 3),
        "has_outstanding_balance": money_outstanding != 0 or gold_balance != 0
    }
    
    return impact

@api_router.delete("/parties/{party_id}")
async def delete_party(party_id: str, current_user: User = Depends(require_permission('parties.delete'))):
    if not user_has_permission(current_user, 'parties.delete'):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="You don't have permission to delete parties")
    
    existing = await db.parties.find_one({"id": party_id, "is_deleted": False})
    if not existing:
        raise HTTPException(status_code=404, detail="Party not found")
    
    await db.parties.update_one(
        {"id": party_id},
        {"$set": {"is_deleted": True, "deleted_at": datetime.now(timezone.utc), "deleted_by": current_user.id}}
    )
    await create_audit_log(current_user.id, current_user.full_name, "party", party_id, "delete")
    return {"message": "Party deleted successfully"}

# ============================================================================
# WORKER MANAGEMENT ENDPOINTS
# ============================================================================

# ============================================================================
# WORKER NAME VALIDATION HELPER
# ============================================================================

def validate_worker_name(name: str) -> tuple[bool, str]:
    """
    Validate worker name according to strict data quality rules.
    
    Rules:
    - Length: 3-50 characters
    - Allowed: letters (A-Z, a-z) and spaces only
    - No numbers or special characters
    
    Returns:
        (is_valid, error_message)
    """
    if not name or not name.strip():
        return False, "Worker name is required"
    
    # Trim the name
    name = name.strip()
    
    # Check length (3-50 characters)
    if len(name) < 3:
        return False, "Worker name must be at least 3 characters long"
    
    if len(name) > 50:
        return False, "Worker name cannot exceed 50 characters"
    
    # Check for only letters and spaces (no numbers or special characters)
    if not re.match(r'^[A-Za-z\s]+$', name):
        return False, "Worker name can only contain letters and spaces (no numbers or special characters)"
    
    return True, ""


@api_router.get("/workers")
@limiter.limit("1000/hour")
async def get_workers(
    request: Request,
    active: Optional[bool] = None,
    current_user: User = Depends(get_current_user)
):
    """Get all workers with optional active filter"""
    query = {"is_deleted": False}
    if active is not None:
        query['active'] = active
    
    workers = await db.workers.find(query, {"_id": 0}).sort("name", 1).to_list(None)
    return {"items": workers}

@api_router.post("/workers", response_model=Worker)
@limiter.limit("1000/hour")
async def create_worker(request: Request, worker_data: dict, current_user: User = Depends(get_current_user)):
    """Create a new worker"""
    # Validate worker name format
    name = worker_data.get('name', '')
    is_valid, error_msg = validate_worker_name(name)
    if not is_valid:
        raise HTTPException(status_code=400, detail=error_msg)
    
    # Trim and normalize the name
    name = name.strip()
    worker_data['name'] = name
    
    # Check for duplicate name (case-insensitive)
    existing_name = await db.workers.find_one({
        "name": {"$regex": f"^{re.escape(name)}$", "$options": "i"},
        "is_deleted": False
    })
    if existing_name:
        raise HTTPException(
            status_code=400,
            detail=f"Worker with name '{name}' already exists"
        )
    
    # Validate duplicate phone if provided
    phone = worker_data.get('phone')
    if phone and phone.strip():
        existing_phone = await db.workers.find_one({
            "phone": phone.strip(),
            "is_deleted": False
        })
        if existing_phone:
            raise HTTPException(
                status_code=400,
                detail=f"Phone number {phone} is already registered with another worker: {existing_phone.get('name', 'Unknown')}"
            )
    
    worker = Worker(**worker_data, created_by=current_user.id)
    await db.workers.insert_one(worker.model_dump())
    await create_audit_log(current_user.id, current_user.full_name, "worker", worker.id, "create")
    return worker

@api_router.get("/workers/{worker_id}", response_model=Worker)
async def get_worker(worker_id: str, current_user: User = Depends(get_current_user)):
    """Get a specific worker by ID"""
    worker = await db.workers.find_one({"id": worker_id, "is_deleted": False}, {"_id": 0})
    if not worker:
        raise HTTPException(status_code=404, detail="Worker not found")
    return Worker(**worker)

@api_router.patch("/workers/{worker_id}")
async def update_worker(worker_id: str, update_data: dict, current_user: User = Depends(get_current_user)):
    """Update a worker"""
    existing = await db.workers.find_one({"id": worker_id, "is_deleted": False})
    if not existing:
        raise HTTPException(status_code=404, detail="Worker not found")
    
    # Validate worker name format if name is being changed
    if 'name' in update_data:
        name = update_data['name']
        is_valid, error_msg = validate_worker_name(name)
        if not is_valid:
            raise HTTPException(status_code=400, detail=error_msg)
        
        # Trim and normalize the name
        name = name.strip()
        update_data['name'] = name
        
        # Check for duplicate name (case-insensitive, excluding current worker)
        existing_name = await db.workers.find_one({
            "name": {"$regex": f"^{re.escape(name)}$", "$options": "i"},
            "is_deleted": False,
            "id": {"$ne": worker_id}
        })
        if existing_name:
            raise HTTPException(
                status_code=400,
                detail=f"Worker with name '{name}' already exists"
            )
    
    # Validate duplicate phone if phone is being changed
    if 'phone' in update_data:
        phone = update_data['phone']
        if phone and phone.strip():
            existing_phone = await db.workers.find_one({
                "phone": phone.strip(),
                "is_deleted": False,
                "id": {"$ne": worker_id}
            })
            if existing_phone:
                raise HTTPException(
                    status_code=400,
                    detail=f"Phone number {phone} is already registered with another worker: {existing_phone.get('name', 'Unknown')}"
                )
    
    await db.workers.update_one({"id": worker_id}, {"$set": update_data})
    await create_audit_log(current_user.id, current_user.full_name, "worker", worker_id, "update", update_data)
    return {"message": "Worker updated successfully"}

@api_router.delete("/workers/{worker_id}")
async def delete_worker(worker_id: str, current_user: User = Depends(get_current_user)):
    """Soft delete a worker"""
    existing = await db.workers.find_one({"id": worker_id, "is_deleted": False})
    if not existing:
        raise HTTPException(status_code=404, detail="Worker not found")
    
    # Check if worker is assigned to any active job cards
    active_jobcards = await db.jobcards.count_documents({
        "worker_id": worker_id,
        "is_deleted": False,
        "status": {"$in": ["created", "in_progress"]}
    })
    
    if active_jobcards > 0:
        raise HTTPException(
            status_code=400,
            detail=f"Cannot delete worker. {active_jobcards} active job card(s) are assigned to this worker. Please reassign or complete them first."
        )
    
    await db.workers.update_one(
        {"id": worker_id},
        {"$set": {"is_deleted": True, "deleted_at": datetime.now(timezone.utc), "deleted_by": current_user.id}}
    )
    await create_audit_log(current_user.id, current_user.full_name, "worker", worker_id, "delete")
    return {"message": "Worker deleted successfully"}

@api_router.get("/parties/{party_id}/ledger")
async def get_party_ledger(party_id: str, current_user: User = Depends(require_permission('parties.view'))):
    invoices = await db.invoices.find({"customer_id": party_id, "is_deleted": False}, {"_id": 0}).to_list(1000)
    transactions = await db.transactions.find({"party_id": party_id, "is_deleted": False}, {"_id": 0}).to_list(1000)
    
    outstanding = 0
    for inv in invoices:
        outstanding += inv.get('balance_due', 0)
    
    return {"invoices": invoices, "transactions": transactions, "outstanding": outstanding}

# Gold Ledger Endpoints
@api_router.post("/gold-ledger", response_model=GoldLedgerEntry)
async def create_gold_ledger_entry(entry_data: dict, current_user: User = Depends(require_permission('finance.create'))):
    # Validate required fields
    if 'party_id' not in entry_data:
        raise HTTPException(status_code=400, detail="party_id is required")
    if 'type' not in entry_data or entry_data['type'] not in ['IN', 'OUT']:
        raise HTTPException(status_code=400, detail="type must be 'IN' or 'OUT'")
    if 'weight_grams' not in entry_data:
        raise HTTPException(status_code=400, detail="weight_grams is required")
    if 'purity_entered' not in entry_data:
        raise HTTPException(status_code=400, detail="purity_entered is required")
    if 'purpose' not in entry_data or entry_data['purpose'] not in ['job_work', 'exchange', 'advance_gold', 'adjustment']:
        raise HTTPException(status_code=400, detail="purpose must be one of: job_work, exchange, advance_gold, adjustment")
    
    # Verify party exists
    party = await db.parties.find_one({"id": entry_data['party_id'], "is_deleted": False})
    if not party:
        raise HTTPException(status_code=404, detail="Party not found")
    
    # Round weight to 3 decimal places
    weight_grams = round(float(entry_data['weight_grams']), 3)
    
    # Create entry
    entry = GoldLedgerEntry(
        party_id=entry_data['party_id'],
        date=entry_data.get('date', datetime.now(timezone.utc)),
        type=entry_data['type'],
        weight_grams=weight_grams,
        purity_entered=int(entry_data['purity_entered']),
        purpose=entry_data['purpose'],
        reference_type=entry_data.get('reference_type'),
        reference_id=entry_data.get('reference_id'),
        notes=entry_data.get('notes'),
        created_by=current_user.id
    )
    
    await db.gold_ledger.insert_one(entry.model_dump())
    await create_audit_log(current_user.id, current_user.full_name, "gold_ledger", entry.id, "create")
    return entry

@api_router.get("/gold-ledger")
async def get_gold_ledger_entries(
    party_id: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    page: int = 1,
    per_page: int = 50,
    current_user: User = Depends(require_permission('finance.view'))
):
    """Get gold ledger entries with optional filters and pagination"""
    query = {"is_deleted": False}
    
    # Filter by party_id
    if party_id:
        query['party_id'] = party_id
    
    # Filter by date range
    if date_from or date_to:
        date_query = {}
        if date_from:
            try:
                date_query['$gte'] = datetime.fromisoformat(date_from.replace('Z', '+00:00'))
            except ValueError:
                raise HTTPException(status_code=400, detail="Invalid date_from format. Use ISO format (YYYY-MM-DD or YYYY-MM-DDTHH:MM:SS)")
        if date_to:
            try:
                date_query['$lte'] = datetime.fromisoformat(date_to.replace('Z', '+00:00'))
            except ValueError:
                raise HTTPException(status_code=400, detail="Invalid date_to format. Use ISO format (YYYY-MM-DD or YYYY-MM-DDTHH:MM:SS)")
        query['date'] = date_query
    
    # Calculate skip value
    skip = (page - 1) * per_page
    
    # Get total count for pagination
    total_count = await db.gold_ledger.count_documents(query)
    
    # Get paginated results
    entries = await db.gold_ledger.find(query, {"_id": 0}).sort("date", -1).skip(skip).limit(per_page).to_list(per_page)
    
    return create_pagination_response(entries, total_count, page, per_page)

@api_router.delete("/gold-ledger/{entry_id}")
async def delete_gold_ledger_entry(entry_id: str, current_user: User = Depends(require_permission('finance.delete'))):
    # Check if entry exists
    entry = await db.gold_ledger.find_one({"id": entry_id, "is_deleted": False})
    if not entry:
        raise HTTPException(status_code=404, detail="Gold ledger entry not found")
    
    # Soft delete
    await db.gold_ledger.update_one(
        {"id": entry_id},
        {"$set": {
            "is_deleted": True,
            "deleted_at": datetime.now(timezone.utc),
            "deleted_by": current_user.id
        }}
    )
    
    await create_audit_log(current_user.id, current_user.full_name, "gold_ledger", entry_id, "delete")
    return {"message": "Gold ledger entry deleted successfully"}

# ============================================================================
# MODULE 9/10 - GOLD DEPOSITS (Customer Gold Received) - Specific IN Entry API
# ============================================================================

@api_router.post("/gold-deposits", response_model=GoldLedgerEntry)
async def create_gold_deposit(deposit_data: dict, current_user: User = Depends(require_permission('finance.create'))):
    """
    Create a gold deposit entry - specifically for recording gold RECEIVED from customer.
    This is a convenience endpoint that creates a GoldLedgerEntry with type="IN".
    
    Required fields:
    - party_id: Customer ID who is depositing gold
    - weight_grams: Weight of gold received (3 decimal precision)
    - purity_entered: Purity in karats (e.g., 22, 24, 18)
    - purpose: One of [job_work, exchange, advance_gold, adjustment]
    
    Optional fields:
    - notes: Additional notes about the deposit
    """
    # Validate required fields
    if 'party_id' not in deposit_data:
        raise HTTPException(status_code=400, detail="party_id is required")
    if 'weight_grams' not in deposit_data:
        raise HTTPException(status_code=400, detail="weight_grams is required")
    if 'purity_entered' not in deposit_data:
        raise HTTPException(status_code=400, detail="purity_entered is required")
    if 'purpose' not in deposit_data or deposit_data['purpose'] not in ['job_work', 'exchange', 'advance_gold', 'adjustment']:
        raise HTTPException(status_code=400, detail="purpose must be one of: job_work, exchange, advance_gold, adjustment")
    
    # Verify party exists
    party = await db.parties.find_one({"id": deposit_data['party_id'], "is_deleted": False})
    if not party:
        raise HTTPException(status_code=404, detail="Party not found")
    
    # Round weight to 3 decimal places
    weight_grams = round(float(deposit_data['weight_grams']), 3)
    
    # Validate weight is positive
    if weight_grams <= 0:
        raise HTTPException(status_code=400, detail="weight_grams must be greater than 0")
    
    # Create gold deposit entry with type="IN"
    entry = GoldLedgerEntry(
        party_id=deposit_data['party_id'],
        date=deposit_data.get('date', datetime.now(timezone.utc)),
        type="IN",  # ALWAYS IN for deposits - customer gives gold to shop
        weight_grams=weight_grams,
        purity_entered=int(deposit_data['purity_entered']),
        purpose=deposit_data['purpose'],
        reference_type="manual",  # Manual gold deposit entry
        reference_id=None,
        notes=deposit_data.get('notes'),
        created_by=current_user.id
    )
    
    await db.gold_ledger.insert_one(entry.model_dump())
    await create_audit_log(current_user.id, current_user.full_name, "gold_deposit", entry.id, "create")
    return entry

@api_router.get("/gold-deposits", response_model=List[GoldLedgerEntry])
async def get_gold_deposits(
    party_id: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """
    Get gold deposit entries - filters for type="IN" entries only.
    These are gold deposits received from customers.
    
    Query parameters:
    - party_id: Filter by specific party/customer
    - date_from: Filter entries from this date (ISO format)
    - date_to: Filter entries up to this date (ISO format)
    """
    query = {
        "is_deleted": False,
        "type": "IN"  # Only get IN entries (deposits received)
    }
    
    # Filter by party_id
    if party_id:
        query['party_id'] = party_id
    
    # Filter by date range
    if date_from or date_to:
        date_query = {}
        if date_from:
            try:
                date_query['$gte'] = datetime.fromisoformat(date_from.replace('Z', '+00:00'))
            except ValueError:
                raise HTTPException(status_code=400, detail="Invalid date_from format. Use ISO format (YYYY-MM-DD or YYYY-MM-DDTHH:MM:SS)")
        if date_to:
            try:
                date_query['$lte'] = datetime.fromisoformat(date_to.replace('Z', '+00:00'))
            except ValueError:
                raise HTTPException(status_code=400, detail="Invalid date_to format. Use ISO format (YYYY-MM-DD or YYYY-MM-DDTHH:MM:SS)")
        query['date'] = date_query
    
    entries = await db.gold_ledger.find(query, {"_id": 0}).sort("date", -1).to_list(1000)
    return entries

# Note: Soft delete is handled by existing DELETE /api/gold-ledger/{entry_id} endpoint

# ============================================================================

@api_router.get("/parties/{party_id}/gold-summary")
async def get_party_gold_summary(party_id: str, current_user: User = Depends(require_permission('parties.view'))):
    # Verify party exists
    party = await db.parties.find_one({"id": party_id, "is_deleted": False})
    if not party:
        raise HTTPException(status_code=404, detail="Party not found")
    
    # Get all gold ledger entries for this party
    entries = await db.gold_ledger.find({"party_id": party_id, "is_deleted": False}, {"_id": 0}).to_list(1000)
    
    # Calculate gold balance
    gold_due_from_party = 0.0  # Party owes shop (IN entries - shop received from party)
    gold_due_to_party = 0.0     # Shop owes party (OUT entries - shop gave to party)
    
    for entry in entries:
        weight = round(entry.get('weight_grams', 0), 3)
        if entry.get('type') == 'IN':
            # Shop received gold from party - party owes shop
            gold_due_from_party += weight
        elif entry.get('type') == 'OUT':
            # Shop gave gold to party - shop owes party
            gold_due_to_party += weight
    
    # Round to 3 decimal places
    gold_due_from_party = round(gold_due_from_party, 3)
    gold_due_to_party = round(gold_due_to_party, 3)
    net_gold_balance = round(gold_due_from_party - gold_due_to_party, 3)
    
    return {
        "party_id": party_id,
        "party_name": party.get('name'),
        "gold_due_from_party": gold_due_from_party,  # Party owes shop
        "gold_due_to_party": gold_due_to_party,      # Shop owes party
        "net_gold_balance": net_gold_balance,        # Positive = party owes shop, Negative = shop owes party
        "total_entries": len(entries)
    }

@api_router.get("/parties/{party_id}/summary")
async def get_party_summary(party_id: str, current_user: User = Depends(require_permission('parties.view'))):
    """
    Get comprehensive party summary including both gold and money balances.
    This endpoint combines gold ledger data and financial data (invoices + transactions).
    """
    # Verify party exists
    party = await db.parties.find_one({"id": party_id, "is_deleted": False})
    if not party:
        raise HTTPException(status_code=404, detail="Party not found")
    
    # Get gold ledger entries
    gold_entries = await db.gold_ledger.find({"party_id": party_id, "is_deleted": False}, {"_id": 0}).to_list(1000)
    
    # Calculate gold balances
    gold_due_from_party = 0.0  # Party owes shop (IN entries)
    gold_due_to_party = 0.0     # Shop owes party (OUT entries)
    
    for entry in gold_entries:
        weight = round(entry.get('weight_grams', 0), 3)
        if entry.get('type') == 'IN':
            gold_due_from_party += weight
        elif entry.get('type') == 'OUT':
            gold_due_to_party += weight
    
    gold_due_from_party = round(gold_due_from_party, 3)
    gold_due_to_party = round(gold_due_to_party, 3)
    net_gold_balance = round(gold_due_from_party - gold_due_to_party, 3)
    
    # Get invoices for money calculations - ONLY FINALIZED invoices
    invoices = await db.invoices.find({"customer_id": party_id, "is_deleted": False, "status": "finalized"}, {"_id": 0}).to_list(1000)
    
    # Get transactions
    transactions = await db.transactions.find({"party_id": party_id, "is_deleted": False}, {"_id": 0}).to_list(1000)
    
    # Calculate money balances
    money_due_from_party = 0.0  # Outstanding invoices (party owes shop)
    money_due_to_party = 0.0     # Credits or vendor payables (shop owes party)
    
    # Sum up outstanding invoices
    for inv in invoices:
        balance_due = inv.get('balance_due', 0)
        if balance_due > 0:
            money_due_from_party += balance_due
        elif balance_due < 0:
            # Negative balance means shop owes party (overpayment/credit)
            money_due_to_party += abs(balance_due)
    
    # Check transactions for additional credits (vendor payments, etc.)
    for txn in transactions:
        # If transaction is credit type and to this party, shop owes them
        if txn.get('transaction_type') == 'credit':
            money_due_to_party += txn.get('amount', 0)
    
    money_due_from_party = round(money_due_from_party, 2)
    money_due_to_party = round(money_due_to_party, 2)
    net_money_balance = round(money_due_from_party - money_due_to_party, 2)
    
    # Clean party data for response
    party_clean = {
        "id": party.get('id'),
        "name": party.get('name'),
        "phone": party.get('phone'),
        "address": party.get('address'),
        "party_type": party.get('party_type'),
        "notes": party.get('notes'),
        "created_at": party.get('created_at')
    }
    
    return {
        "party": party_clean,
        "gold": {
            "gold_due_from_party": gold_due_from_party,
            "gold_due_to_party": gold_due_to_party,
            "net_gold_balance": net_gold_balance,
            "total_entries": len(gold_entries)
        },
        "money": {
            "money_due_from_party": money_due_from_party,
            "money_due_to_party": money_due_to_party,
            "net_money_balance": net_money_balance,
            "total_invoices": len(invoices),
            "total_transactions": len(transactions)
        }
    }


# ===========================
# PURCHASES MODULE (Stock IN + Vendor Payable)
# ===========================

@api_router.post("/purchases", response_model=Purchase)
@limiter.limit("1000/hour")  # General authenticated rate limit: 1000 requests per hour
async def create_purchase(request: Request, purchase_data: dict, current_user: User = Depends(require_permission('purchases.create'))):
    """
    Create and auto-finalize a new purchase.
    
    ⚠️ CRITICAL BUSINESS RULE: All purchases are auto-finalized on creation ⚠️
    - Inventory is updated immediately
    - Accounting entries are created immediately
    - No "Draft" state exists for normal purchases
    - Status is calculated based on payment: "Finalized (Unpaid)" | "Partially Paid" | "Paid"
    """
    if not user_has_permission(current_user, 'purchases.create'):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="You don't have permission to create purchases")
    
    # Validate vendor exists and is vendor type
    vendor = await db.parties.find_one({"id": purchase_data.get("vendor_party_id"), "is_deleted": False})
    if not vendor:
        raise HTTPException(status_code=404, detail="Vendor not found")
    if vendor.get("party_type") != "vendor":
        raise HTTPException(status_code=400, detail="Party must be a vendor type")
    
    # ========== CRITICAL VALIDATION: NEVER TRUST FRONTEND ==========
    # Extract and validate weight
    try:
        weight_grams = float(purchase_data.get("weight_grams", 0))
    except (ValueError, TypeError):
        raise HTTPException(status_code=400, detail="Invalid weight value")
    
    if weight_grams <= 0:
        raise HTTPException(status_code=400, detail="Weight must be greater than 0")
    
    # Extract and validate rate
    try:
        rate_per_gram = float(purchase_data.get("rate_per_gram", 0))
    except (ValueError, TypeError):
        raise HTTPException(status_code=400, detail="Invalid rate value")
    
    if rate_per_gram <= 0:
        raise HTTPException(status_code=400, detail="Rate per gram must be greater than 0")
    
    # RECALCULATE total_amount on backend (SINGLE SOURCE OF TRUTH)
    # Never trust client-sent total_amount
    calculated_total = round(weight_grams * rate_per_gram, 2)
    
    # Validate paid amount
    try:
        paid_amount = float(purchase_data.get("paid_amount_money", 0))
    except (ValueError, TypeError):
        raise HTTPException(status_code=400, detail="Invalid paid amount value")
    
    if paid_amount < 0:
        raise HTTPException(status_code=400, detail="Paid amount cannot be negative")
    
    if paid_amount > calculated_total:
        raise HTTPException(status_code=400, detail=f"Paid amount ({paid_amount}) cannot exceed total amount ({calculated_total})")
    
    # Set validated and calculated values
    purchase_data["weight_grams"] = round(weight_grams, 3)
    purchase_data["rate_per_gram"] = round(rate_per_gram, 2)
    purchase_data["amount_total"] = calculated_total  # Backend-calculated, not from client
    purchase_data["paid_amount_money"] = round(paid_amount, 2)
    purchase_data["balance_due_money"] = round(calculated_total - paid_amount, 2)
    
    # Round gold settlement fields to 3 decimals
    if purchase_data.get("advance_in_gold_grams") is not None:
        purchase_data["advance_in_gold_grams"] = round(float(purchase_data["advance_in_gold_grams"]), 3)
    if purchase_data.get("exchange_in_gold_grams") is not None:
        purchase_data["exchange_in_gold_grams"] = round(float(purchase_data["exchange_in_gold_grams"]), 3)
    
    # Validate account exists if payment made
    if purchase_data["paid_amount_money"] > 0:
        if not purchase_data.get("account_id"):
            raise HTTPException(status_code=400, detail="account_id is required when paid_amount_money > 0")
        account = await db.accounts.find_one({"id": purchase_data["account_id"], "is_deleted": False})
        if not account:
            raise HTTPException(status_code=404, detail="Payment account not found")
    
    # Ensure valuation purity is always 916
    purchase_data["valuation_purity_fixed"] = 916
    
    # ========== CRITICAL: CALCULATE STATUS (SERVER-SIDE AUTHORITY) ==========
    # Status is NEVER "draft" - purchases are auto-finalized
    calculated_status = calculate_purchase_status(
        paid_amount=purchase_data["paid_amount_money"],
        total_amount=purchase_data["amount_total"]
    )
    
    # ========== SAFETY ASSERTION: PREVENT "DRAFT" WITH COMMITTED DATA ==========
    # A purchase MUST NEVER be "Draft" if we're creating inventory/accounting entries
    # "Draft" is ONLY for future explicit "save as draft" feature (not yet implemented)
    if calculated_status.lower() == "draft":
        raise HTTPException(
            status_code=500,
            detail="CRITICAL ERROR: Cannot save purchase as 'Draft' when inventory/accounting entries are being created. This violates ERP business rules."
        )

    
    # Set creation and finalization metadata
    finalize_time = datetime.now(timezone.utc)
    purchase_data["created_by"] = current_user.username
    purchase_data["status"] = calculated_status
    purchase_data["finalized_at"] = finalize_time
    purchase_data["finalized_by"] = current_user.username
    purchase_data["locked"] = True
    purchase_data["locked_at"] = finalize_time
    purchase_data["locked_by"] = current_user.username
    
    # Create Purchase model instance
    purchase = Purchase(**purchase_data)
    purchase_id = purchase.id
    
    # Insert purchase
    await db.purchases.insert_one(purchase.model_dump())
    
    # ========== AUTO-FINALIZATION: CREATE ALL ACCOUNTING ENTRIES ==========
    
    # === OPERATION 1: Create Stock IN movement ===
    purity = purchase_data["valuation_purity_fixed"]  # Always 916
    header_name = f"Gold {purity // 41.6:.0f}K"  # 916 = 22K
    
    header = await db.inventory_headers.find_one({"name": header_name, "is_deleted": False})
    if not header:
        # Create new inventory header
        header = InventoryHeader(
            name=header_name,
            purity=purity,
            current_qty=0,
            current_weight=0,
            created_by=current_user.username
        )
        await db.inventory_headers.insert_one(header.model_dump())
    
    # Create Stock IN movement
    movement = StockMovement(
        date=purchase.date,
        movement_type="Stock IN",
        header_id=header.get("id") if isinstance(header, dict) else header.id,
        header_name=header_name,
        description=f"Purchase from {vendor['name']}: {purchase.description}",
        qty_delta=1,
        weight_delta=purchase.weight_grams,
        purity=purity,
        reference_type="purchase",
        reference_id=purchase_id,
        created_by=current_user.username,
        notes=f"Entered purity: {purchase.entered_purity}, Valuation purity: {purity}"
    )
    await db.stock_movements.insert_one(movement.model_dump())
    
    # Update inventory header
    header_id = header.get("id") if isinstance(header, dict) else header.id
    await db.inventory_headers.update_one(
        {"id": header_id},
        {"$inc": {
            "current_qty": 1,
            "current_weight": purchase.weight_grams
        }}
    )
    
    # === OPERATION 2: Create DEBIT transaction if paid_amount_money > 0 ===
    if purchase_data["paid_amount_money"] > 0:
        current_year = datetime.now(timezone.utc).year
        existing_txns = await db.transactions.count_documents({"transaction_number": {"$regex": f"^TXN-{current_year}-"}})
        payment_txn_number = f"TXN-{current_year}-{existing_txns + 1:04d}"
        
        account = await db.accounts.find_one({"id": purchase_data["account_id"], "is_deleted": False})
        
        payment_transaction = Transaction(
            transaction_number=payment_txn_number,
            date=purchase.date,
            transaction_type="debit",
            mode=purchase_data.get("payment_mode", "Cash"),
            account_id=purchase_data["account_id"],
            account_name=account["name"],
            party_id=purchase.vendor_party_id,
            party_name=vendor["name"],
            amount=round(purchase_data["paid_amount_money"], 2),
            category="Purchase Payment",
            reference_type="purchase",
            reference_id=purchase_id,
            notes=f"Payment for purchase: {purchase.description} ({purchase.weight_grams}g)",
            created_by=current_user.username
        )
        await db.transactions.insert_one(payment_transaction.model_dump())
        
        # Update account balance
        delta = -payment_transaction.amount
        await db.accounts.update_one(
            {"id": purchase_data["account_id"]}, 
            {"$inc": {"current_balance": delta}}
        )
    
    # === OPERATION 3: Create GoldLedgerEntry OUT if advance_in_gold_grams > 0 ===
    advance_gold = purchase_data.get("advance_in_gold_grams")
    if advance_gold and advance_gold > 0:
        advance_gold = round(advance_gold, 3)
        
        advance_entry = GoldLedgerEntry(
            party_id=purchase.vendor_party_id,
            date=purchase.date,
            type="OUT",
            weight_grams=advance_gold,
            purity_entered=purchase.entered_purity,
            purpose="advance_gold",
            reference_type="purchase",
            reference_id=purchase_id,
            notes=f"Advance gold settled in purchase: {purchase.description}",
            created_by=current_user.username
        )
        await db.gold_ledger.insert_one(advance_entry.model_dump())
    
    # === OPERATION 4: Create GoldLedgerEntry IN if exchange_in_gold_grams > 0 ===
    exchange_gold = purchase_data.get("exchange_in_gold_grams")
    if exchange_gold and exchange_gold > 0:
        exchange_gold = round(exchange_gold, 3)
        
        exchange_entry = GoldLedgerEntry(
            party_id=purchase.vendor_party_id,
            date=purchase.date,
            type="IN",
            weight_grams=exchange_gold,
            purity_entered=purchase.entered_purity,
            purpose="exchange",
            reference_type="purchase",
            reference_id=purchase_id,
            notes=f"Gold exchanged in purchase: {purchase.description}",
            created_by=current_user.username
        )
        await db.gold_ledger.insert_one(exchange_entry.model_dump())
    
    # === OPERATION 5: Create vendor payable transaction ONLY for balance_due_money ===
    balance_due = purchase_data["balance_due_money"]
    
    if balance_due > 0:
        current_year = datetime.now(timezone.utc).year
        existing_txns = await db.transactions.count_documents({"transaction_number": {"$regex": f"^TXN-{current_year}-"}})
        payable_txn_number = f"TXN-{current_year}-{existing_txns + 1:04d}"
        
        purchases_account = await db.accounts.find_one({"name": "Purchases", "is_deleted": False})
        if not purchases_account:
            purchases_account = Account(
                name="Purchases",
                account_type="expense",
                balance=0,
                created_by=current_user.username
            )
            await db.accounts.insert_one(purchases_account.model_dump())
        
        payable_transaction = Transaction(
            transaction_number=payable_txn_number,
            date=purchase.date,
            transaction_type="credit",
            mode="Vendor Payable",
            account_id=purchases_account.get("id") if isinstance(purchases_account, dict) else purchases_account.id,
            account_name="Purchases",
            party_id=purchase.vendor_party_id,
            party_name=vendor["name"],
            amount=round(balance_due, 2),
            category="Purchase",
            reference_type="purchase",
            reference_id=purchase_id,
            notes=f"Vendor payable for purchase: {purchase.description} ({purchase.weight_grams}g @ {purchase.rate_per_gram}/g)",
            created_by=current_user.username
        )
        await db.transactions.insert_one(payable_transaction.model_dump())
    
    # Create audit log
    await create_audit_log(
        user_id=current_user.id,
        user_name=current_user.username,
        module="purchases",
        record_id=purchase_id,
        action="create",
        changes={
            "vendor_party_id": purchase.vendor_party_id,
            "weight_grams": purchase.weight_grams,
            "entered_purity": purchase.entered_purity,
            "amount_total": purchase.amount_total,
            "paid_amount_money": purchase.paid_amount_money,
            "balance_due_money": purchase.balance_due_money,
            "advance_in_gold_grams": purchase.advance_in_gold_grams,
            "exchange_in_gold_grams": purchase.exchange_in_gold_grams,
            "status": calculated_status,
            "auto_finalized": True
        }
    )
    
    # Return the created purchase with correct status
    return purchase

@api_router.get("/purchases")
@limiter.limit("1000/hour")  # General authenticated rate limit: 1000 requests per hour
async def get_purchases(
    request: Request,
    vendor_party_id: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    status: Optional[str] = None,
    page: int = 1,
    page_size: int = 10,
    current_user: User = Depends(require_permission('purchases.view'))
):
    """Get all purchases with optional filters and pagination"""
    query = {"is_deleted": False}
    
    # Filter by vendor
    if vendor_party_id:
        query["vendor_party_id"] = vendor_party_id
    
    # Filter by date range
    if start_date or end_date:
        query["date"] = {}
        if start_date:
            query["date"]["$gte"] = datetime.fromisoformat(start_date.replace('Z', '+00:00'))
        if end_date:
            query["date"]["$lte"] = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
    
    # Filter by status
    if status:
        query["status"] = status
    
    # Calculate skip value
    skip = (page - 1) * page_size
    
    # Get total count for pagination
    total_count = await db.purchases.count_documents(query)
    
    # Get paginated results
    purchases = await db.purchases.find(query).sort("date", -1).skip(skip).limit(page_size).to_list(page_size)
    
    # CRITICAL FIX: Process purchases through decimal_to_float to handle Decimal serialization
    purchases = [decimal_to_float(p) for p in purchases]
    
    return create_pagination_response(purchases, total_count, page, page_size)

@api_router.patch("/purchases/{purchase_id}")
async def update_purchase(
    purchase_id: str,
    updates: Dict,
    current_user: User = Depends(require_permission('purchases.create'))
):
    """Update a purchase (only draft purchases can be edited)"""
    # Get existing purchase
    existing = await db.purchases.find_one({"id": purchase_id, "is_deleted": False})
    if not existing:
        raise HTTPException(status_code=404, detail="Purchase not found")
    
    # Check if purchase is finalized
    if existing.get("status") == "finalized":
        raise HTTPException(
            status_code=400,
            detail="Cannot edit finalized purchase. Finalized purchases are immutable to maintain financial integrity."
        )
    
    # Validate vendor if being updated
    if "vendor_party_id" in updates:
        vendor = await db.parties.find_one({"id": updates["vendor_party_id"], "is_deleted": False})
        if not vendor:
            raise HTTPException(status_code=404, detail="Vendor not found")
        if vendor.get("party_type") != "vendor":
            raise HTTPException(status_code=400, detail="Party must be a vendor type")
    
    # ========== CRITICAL VALIDATION: NEVER TRUST FRONTEND ==========
    # Get current values or updated values
    weight_grams = updates.get("weight_grams", existing.get("weight_grams", 0))
    rate_per_gram = updates.get("rate_per_gram", existing.get("rate_per_gram", 0))
    
    # Validate weight if provided
    if "weight_grams" in updates:
        try:
            weight_grams = float(updates["weight_grams"])
        except (ValueError, TypeError):
            raise HTTPException(status_code=400, detail="Invalid weight value")
        
        if weight_grams <= 0:
            raise HTTPException(status_code=400, detail="Weight must be greater than 0")
        
        updates["weight_grams"] = round(weight_grams, 3)
    
    # Validate rate if provided
    if "rate_per_gram" in updates:
        try:
            rate_per_gram = float(updates["rate_per_gram"])
        except (ValueError, TypeError):
            raise HTTPException(status_code=400, detail="Invalid rate value")
        
        if rate_per_gram <= 0:
            raise HTTPException(status_code=400, detail="Rate per gram must be greater than 0")
        
        updates["rate_per_gram"] = round(rate_per_gram, 2)
    
    # RECALCULATE total_amount on backend (SINGLE SOURCE OF TRUTH)
    # Never trust client-sent total_amount
    calculated_total = round(float(weight_grams) * float(rate_per_gram), 2)
    updates["amount_total"] = calculated_total
    
    # Validate paid amount
    paid_amount = updates.get("paid_amount_money", existing.get("paid_amount_money", 0))
    if "paid_amount_money" in updates:
        try:
            paid_amount = float(updates["paid_amount_money"])
        except (ValueError, TypeError):
            raise HTTPException(status_code=400, detail="Invalid paid amount value")
        
        if paid_amount < 0:
            raise HTTPException(status_code=400, detail="Paid amount cannot be negative")
        
        if paid_amount > calculated_total:
            raise HTTPException(status_code=400, detail=f"Paid amount ({paid_amount}) cannot exceed total amount ({calculated_total})")
        
        updates["paid_amount_money"] = round(paid_amount, 2)
    
    # Auto-calculate balance_due_money
    updates["balance_due_money"] = round(calculated_total - float(paid_amount), 2)
    
    # Round gold settlement fields to 3 decimals
    if "advance_in_gold_grams" in updates and updates["advance_in_gold_grams"] is not None:
        updates["advance_in_gold_grams"] = round(float(updates["advance_in_gold_grams"]), 3)
    if "exchange_in_gold_grams" in updates and updates["exchange_in_gold_grams"] is not None:
        updates["exchange_in_gold_grams"] = round(float(updates["exchange_in_gold_grams"]), 3)
    
    # Validate account exists if payment made
    if "paid_amount_money" in updates and updates["paid_amount_money"] > 0:
        account_id = updates.get("account_id") or existing.get("account_id")
        if not account_id:
            raise HTTPException(status_code=400, detail="account_id is required when paid_amount_money > 0")
        account = await db.accounts.find_one({"id": account_id, "is_deleted": False})
        if not account:
            raise HTTPException(status_code=404, detail="Payment account not found")
    
    # Update purchase
    await db.purchases.update_one(
        {"id": purchase_id},
        {"$set": updates}
    )
    
    # Create audit log
    await create_audit_log(
        user_id=current_user.id,
        user_name=current_user.username,
        module="purchases",
        record_id=purchase_id,
        action="update",
        changes=updates
    )
    
    # Get updated purchase
    updated = await db.purchases.find_one({"id": purchase_id})
    return updated


@api_router.get("/purchases/{purchase_id}/impact")
async def get_purchase_impact(purchase_id: str, current_user: User = Depends(require_permission('purchases.view'))):
    """
    Get impact summary for purchase actions (finalization or deletion).
    Shows decision-critical data: weight, cost, stock impact, vendor payable, what will be locked.
    """
    purchase = await db.purchases.find_one({"id": purchase_id, "is_deleted": False}, {"_id": 0})
    if not purchase:
        raise HTTPException(status_code=404, detail="Purchase not found")
    
    # Get vendor details
    vendor = await db.parties.find_one(
        {"id": purchase["vendor_party_id"], "is_deleted": False},
        {"_id": 0, "id": 1, "name": 1}
    )
    
    weight_grams = purchase.get("weight_grams", 0)
    entered_purity = purchase.get("entered_purity", 0)
    valuation_purity_fixed = purchase.get("valuation_purity_fixed", 916)
    rate_per_gram = purchase.get("rate_per_gram", 0)
    amount_total = purchase.get("amount_total", 0)
    paid_amount_money = purchase.get("paid_amount_money", 0)
    balance_due_money = purchase.get("balance_due_money", 0)
    advance_in_gold_grams = purchase.get("advance_in_gold_grams", 0)
    exchange_in_gold_grams = purchase.get("exchange_in_gold_grams", 0)
    
    # Build impact summary
    impact = {
        "current_status": purchase.get("status", "draft"),
        "vendor_name": vendor.get("name") if vendor else "Unknown",
        "description": purchase.get("description"),
        "weight_grams": round(weight_grams, 3),
        "entered_purity": entered_purity,
        "valuation_purity": valuation_purity_fixed,
        "rate_per_gram": round(rate_per_gram, 2),
        "amount_total": round(amount_total, 2),
        "paid_amount_money": round(paid_amount_money, 2),
        "balance_due_money": round(balance_due_money, 2),
        "advance_in_gold_grams": round(advance_in_gold_grams, 3) if advance_in_gold_grams else 0,
        "exchange_in_gold_grams": round(exchange_in_gold_grams, 3) if exchange_in_gold_grams else 0,
        "stock_will_increase_by": round(weight_grams, 3),
        "stock_purity_used": valuation_purity_fixed,
        "vendor_payable_will_be": round(balance_due_money, 2),
        "account_debit_amount": round(paid_amount_money, 2) if paid_amount_money > 0 else 0
    }
    
    return impact

@api_router.post("/purchases/{purchase_id}/finalize")
async def finalize_purchase(purchase_id: str, current_user: User = Depends(require_permission('purchases.finalize'))):
    """
    ⚠️ DEPRECATED ENDPOINT ⚠️
    
    This endpoint is no longer needed as purchases are auto-finalized on creation.
    All purchases are immediately committed with proper status:
    - "Finalized (Unpaid)" if paid_amount == 0
    - "Partially Paid" if 0 < paid_amount < total
    - "Paid" if paid_amount == total
    
    Kept for backward compatibility only.
    """
    if not user_has_permission(current_user, 'purchases.finalize'):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="You don't have permission to finalize purchases")
    
    # Get purchase
    purchase = await db.purchases.find_one({"id": purchase_id, "is_deleted": False})
    if not purchase:
        raise HTTPException(status_code=404, detail="Purchase not found")
    
    # Check status - all purchases are now auto-finalized
    purchase_status = purchase.get("status", "")
    if purchase_status in ["Finalized (Unpaid)", "Partially Paid", "Paid"]:
        raise HTTPException(
            status_code=400, 
            detail=f"Purchase is already finalized with status: {purchase_status}. All purchases are auto-finalized on creation."
        )
    
    # Should never reach here in normal operation
    raise HTTPException(
        status_code=400,
        detail="This endpoint is deprecated. All purchases are auto-finalized on creation."
    )

@api_router.get("/jobcards")
async def get_jobcards(
    page: int = 1,
    page_size: int = 10,
    current_user: User = Depends(require_permission('jobcards.view'))
):
    """Get job cards with pagination support"""
    
    # Normal update for unlocked job cards
    await db.jobcards.update_one({"id": jobcard_id}, {"$set": update_data})
    
    # VALIDATION: Verify timestamp consistency after update (audit safety)
    updated_jobcard = await db.jobcards.find_one({"id": jobcard_id})
    is_valid, error_msg = validate_jobcard_timestamps(
        updated_jobcard.get("status", "created"),
        updated_jobcard.get("completed_at"),
        updated_jobcard.get("delivered_at")
    )
    if not is_valid:
        # This should never happen if logic is correct, but safety check
        await create_audit_log(
            current_user.id, 
            current_user.full_name, 
            "jobcard", 
            jobcard_id, 
            "validation_error", 
            {"error": error_msg}
        )
    
    await create_audit_log(current_user.id, current_user.full_name, "jobcard", jobcard_id, "update", update_data)
    return {"message": "Updated successfully"}

@api_router.delete("/jobcards/{jobcard_id}")
async def delete_jobcard(jobcard_id: str, current_user: User = Depends(require_permission('jobcards.delete'))):
    existing = await db.jobcards.find_one({"id": jobcard_id, "is_deleted": False})
    if not existing:
        raise HTTPException(status_code=404, detail="Job card not found")
    
    # Check if job card is locked (linked to finalized invoice)
    if existing.get("locked", False):
        # Admin override: Allow admins to delete locked job cards
        if current_user.role == 'admin':
            # Perform the delete (soft delete)
            await db.jobcards.update_one(
                {"id": jobcard_id},
                {"$set": {"is_deleted": True, "deleted_at": datetime.now(timezone.utc), "deleted_by": current_user.id}}
            )
            
            # Log admin override with special action
            override_details = {
                "action": "admin_override_delete_locked_jobcard",
                "reason": "Admin deleted a locked job card that is linked to a finalized invoice",
                "locked_at": existing.get("locked_at"),
                "locked_by": existing.get("locked_by"),
                "jobcard_number": existing.get("job_card_number"),
                "customer_name": existing.get("customer_name")
            }
            await create_audit_log(
                current_user.id, 
                current_user.full_name, 
                "jobcard", 
                jobcard_id, 
                "admin_override_delete", 
                override_details
            )
            
            return {
                "message": "Job card deleted successfully (admin override)",
                "warning": "This job card was locked and linked to a finalized invoice"
            }
        else:
            # Non-admin users cannot delete locked job cards
            raise HTTPException(
                status_code=403, 
                detail="Cannot delete locked job card. This job card is linked to a finalized invoice. Only admins can override."
            )
    
    # Normal delete for unlocked job cards
    await db.jobcards.update_one(
        {"id": jobcard_id},
        {"$set": {"is_deleted": True, "deleted_at": datetime.now(timezone.utc), "deleted_by": current_user.id}}
    )
    await create_audit_log(current_user.id, current_user.full_name, "jobcard", jobcard_id, "delete")
    return {"message": "Job card deleted successfully"}

@api_router.get("/jobcards/{jobcard_id}/impact")
async def get_jobcard_impact(jobcard_id: str, current_user: User = Depends(require_permission('jobcards.view'))):
    """
    Get impact summary for job card actions (status changes or deletion).
    Shows decision-critical data: items count, total weight, estimated cost, linked data.
    """
    jobcard = await db.jobcards.find_one({"id": jobcard_id, "is_deleted": False}, {"_id": 0})
    if not jobcard:
        raise HTTPException(status_code=404, detail="Job card not found")
    
    # Calculate totals from items
    items = jobcard.get("items", [])
    total_items = len(items)
    total_weight = sum(item.get("weight_in", 0) for item in items)

    # Calculate total making charges based on making_charge_type and making_charge_value
    total_making_charges = 0
    for item in items:
        if item.get('making_charge_value') is not None:
            making_charge_value = float(item.get('making_charge_value', 0))
            if item.get('making_charge_type') == 'per_gram':
                weight = float(item.get('weight_in', 0))
                total_making_charges += making_charge_value * weight
            else:  # 'flat'
                total_making_charges += making_charge_value
    
    # Check if linked to invoice
    linked_invoice = await db.invoices.find_one(
        {"jobcard_id": jobcard_id, "is_deleted": False},
        {"_id": 0, "id": 1, "invoice_number": 1, "status": 1, "grand_total": 1}
    )
    
    # Build impact summary
    impact = {
        "jobcard_number": jobcard.get("job_card_number"),
        "current_status": jobcard.get("status", "created"),
        "is_locked": jobcard.get("locked", False),
        "customer_name": jobcard.get("customer_name") or jobcard.get("walk_in_name"),
        "worker_name": jobcard.get("worker_name"),
        "total_items": total_items,
        "total_weight_grams": round(total_weight, 3),
        "total_making_charges": round(total_making_charges, 2),
        "has_linked_invoice": linked_invoice is not None,
        "linked_invoice": linked_invoice if linked_invoice else None
    }
    
    return impact

@api_router.post("/jobcards/{jobcard_id}/convert-to-invoice")
async def convert_jobcard_to_invoice(jobcard_id: str, invoice_data: dict, current_user: User = Depends(require_permission('jobcards.create'))):
    jobcard = await db.jobcards.find_one({"id": jobcard_id, "is_deleted": False}, {"_id": 0})
    if not jobcard:
        raise HTTPException(status_code=404, detail="Job card not found")
    
    # CRITICAL: Prevent duplicate invoice conversion
    if jobcard.get('is_invoiced', False):
        raise HTTPException(
            status_code=400, 
            detail=f"This job card has already been converted to invoice. Invoice ID: {jobcard.get('invoice_id', 'N/A')}"
        )
    
    # Use customer type from job card (or allow override from invoice_data)
    customer_type = invoice_data.get('customer_type', jobcard.get('customer_type', 'saved'))
    
    # Validate and get customer data
    if customer_type == 'saved':
        # Use job card customer or require in invoice_data
        customer_id = invoice_data.get('customer_id', jobcard.get('customer_id'))
        customer_name = invoice_data.get('customer_name', jobcard.get('customer_name'))
        if not customer_id:
            raise HTTPException(status_code=400, detail="customer_id is required for saved customers")
    elif customer_type == 'walk_in':
        # Use job card walk-in data or require in invoice_data
        walk_in_name = invoice_data.get('walk_in_name', jobcard.get('walk_in_name'))
        walk_in_phone = invoice_data.get('walk_in_phone', jobcard.get('walk_in_phone', ''))
        if not walk_in_name:
            raise HTTPException(status_code=400, detail="walk_in_name is required for walk-in customers")
    
    year = datetime.now(timezone.utc).year
    count = await db.invoices.count_documents({"invoice_number": {"$regex": f"^INV-{year}"}})
    invoice_number = f"INV-{year}-{str(count + 1).zfill(4)}"
    
    vat_percent = 5.0
    invoice_items = []
    subtotal = 0
    
    # MODULE 8: Get metal_rate - Priority: invoice_data override > jobcard gold_rate > default 20.0
    metal_rate = invoice_data.get('metal_rate') or jobcard.get('gold_rate_at_jobcard') or 20.0
    metal_rate = round(float(metal_rate), 2)  # Ensure 2 decimal precision for rate
    
    # First pass: Create invoice items and calculate subtotal
    for item in jobcard.get('items', []):
        weight = item.get('weight_out') or item.get('weight_in') or 0
        weight = float(weight) if weight else 0.0
        gold_value = round(weight * metal_rate, 3)
        
        # Use making charge from job card if provided, otherwise use default
        if item.get('making_charge_value') is not None:
            if item.get('making_charge_type') == 'per_gram':
                making_value = round(float(item.get('making_charge_value', 0)) * weight, 3)
            else:  # flat
                making_value = round(float(item.get('making_charge_value', 0)), 3)
        else:
            making_value = 5.0  # Default
        
        # Use VAT from job card if provided, otherwise use default
        item_vat_percent = item.get('vat_percent') or vat_percent
        
        # Store item temporarily (VAT will be calculated after discount)
        invoice_items.append({
            'category': item.get('category', ''),
            'description': item.get('description', ''),
            'qty': item.get('qty', 1),
            'weight': weight,
            'purity': item.get('purity', 916),
            'metal_rate': metal_rate,
            'gold_value': gold_value,
            'making_value': making_value,
            'vat_percent': item_vat_percent,
        })
        
        subtotal += gold_value + making_value
    
    subtotal = round(subtotal, 3)
    
    # MODULE 7: Get discount amount from invoice_data (default to 0)
    discount_amount = round(float(invoice_data.get('discount_amount', 0)), 3)
    
    # Validate discount
    if discount_amount < 0:
        raise HTTPException(status_code=400, detail="Discount amount cannot be negative")
    if discount_amount > subtotal:
        raise HTTPException(status_code=400, detail=f"Discount amount ({discount_amount:.3f}) cannot exceed subtotal ({subtotal:.3f})")
    
    # MODULE 7: Calculate taxable amount = subtotal - discount
    taxable = round(subtotal - discount_amount, 3)
    
    # MODULE 7: Calculate VAT on taxable amount (after discount)
    vat_total = round(taxable * vat_percent / 100, 3)
    
    # MODULE 7: Calculate grand total = taxable + VAT
    grand_total = round(taxable + vat_total, 3)
    
    # Second pass: Finalize invoice items with proportional VAT distribution
    # VAT is distributed proportionally across items based on their subtotal contribution
    final_invoice_items = []
    for item_data in invoice_items:
        item_subtotal = item_data['gold_value'] + item_data['making_value']
        # Proportional VAT = (item_subtotal / total_subtotal) * total_VAT
        if subtotal > 0:
            item_vat_amount = round((item_subtotal / subtotal) * vat_total, 3)
        else:
            item_vat_amount = 0.0
        # Line total includes proportional share of VAT
        item_line_total = round(item_subtotal + item_vat_amount, 3)
        
        final_invoice_items.append(InvoiceItem(
            category=item_data['category'],
            description=item_data['description'],
            qty=item_data['qty'],
            weight=item_data['weight'],
            purity=item_data['purity'],
            metal_rate=item_data['metal_rate'],
            gold_value=item_data['gold_value'],
            making_value=item_data['making_value'],
            vat_percent=item_data['vat_percent'],
            vat_amount=item_vat_amount,
            line_total=item_line_total
        ))
    
    invoice_items = final_invoice_items
    
    # Create invoice with customer type specific fields
    invoice_dict = {
        "invoice_number": invoice_number,
        "customer_type": customer_type,
        "invoice_type": "service",
        "items": [item.model_dump() for item in invoice_items],
        "subtotal": subtotal,
        "discount_amount": discount_amount,  # MODULE 7: Include discount
        "vat_total": vat_total,
        "grand_total": grand_total,
        "balance_due": grand_total,
        "jobcard_id": jobcard_id,
        "worker_id": jobcard.get("worker_id"),  # Carry forward worker_id from job card
        "worker_name": jobcard.get("worker_name"),  # Carry forward worker_name from job card
        "created_by": current_user.id
    }
    
    # Add customer-specific fields
    if customer_type == 'saved':
        invoice_dict["customer_id"] = customer_id
        invoice_dict["customer_name"] = customer_name
    else:  # walk_in
        invoice_dict["walk_in_name"] = walk_in_name
        invoice_dict["walk_in_phone"] = walk_in_phone
    
    invoice = Invoice(**invoice_dict)
    
    await db.invoices.insert_one(invoice.model_dump())
    await create_audit_log(current_user.id, current_user.full_name, "invoice", invoice.id, "create_from_jobcard")
    
    # CRITICAL: Update job card to mark as invoiced and prevent duplicate conversions
    await db.jobcards.update_one(
        {"id": jobcard_id},
        {
            "$set": {
                "is_invoiced": True,
                "invoice_id": invoice.id
            }
        }
    )
    await create_audit_log(current_user.id, current_user.full_name, "jobcard", jobcard_id, "converted_to_invoice", {"invoice_id": invoice.id})
    
    return invoice

# ============================================================================
# JOB CARD TEMPLATE ENDPOINTS
# ============================================================================

@api_router.get("/jobcard-templates")
async def get_jobcard_templates(current_user: User = Depends(require_permission('jobcards.view'))):
    """Get all job card templates (accessible to all users)"""
    query = {"card_type": "template", "is_deleted": False}
    templates = await db.jobcards.find(query, {"_id": 0}).sort("template_name", 1).to_list(None)
    return {"items": templates}

@api_router.post("/jobcard-templates")
async def create_jobcard_template(template_data: dict, current_user: User = Depends(require_permission('jobcards.create'))):
    """Create a new job card template (admin only)"""
    # Check if user is admin
    if current_user.role != 'admin':
        raise HTTPException(status_code=403, detail="Only administrators can create templates")
    
    # Validate required fields for templates
    if not template_data.get('template_name') or not template_data.get('template_name').strip():
        raise HTTPException(status_code=400, detail="template_name is required for templates")
    
    # Ensure card_type is set to 'template'
    template_data['card_type'] = 'template'
    
    # Templates don't need job card number, use template name as identifier
    template_data['job_card_number'] = f"TEMPLATE-{str(uuid.uuid4())[:8]}"
    
    # Templates should not have customer information
    template_data['customer_type'] = 'saved'  # Default, will be set when creating actual job card
    template_data['customer_id'] = None
    template_data['customer_name'] = None
    template_data['walk_in_name'] = None
    template_data['walk_in_phone'] = None
    
    # Templates always start with 'created' status (will be set when creating actual job card)
    template_data['status'] = 'created'
    
    # Create the template
    template = JobCard(**template_data, created_by=current_user.id)
    await db.jobcards.insert_one(template.model_dump())
    await create_audit_log(current_user.id, current_user.full_name, "jobcard_template", template.id, "create")
    
    return template

@api_router.patch("/jobcard-templates/{template_id}")
async def update_jobcard_template(template_id: str, update_data: dict, current_user: User = Depends(require_permission('jobcards.update'))):
    """Update a job card template (admin only)"""
    # Check if user is admin
    if current_user.role != 'admin':
        raise HTTPException(status_code=403, detail="Only administrators can edit templates")
    
    # Verify the template exists
    existing = await db.jobcards.find_one({"id": template_id, "card_type": "template", "is_deleted": False})
    if not existing:
        raise HTTPException(status_code=404, detail="Template not found")
    
    # Validate template_name if provided
    if 'template_name' in update_data:
        if not update_data['template_name'] or not update_data['template_name'].strip():
            raise HTTPException(status_code=400, detail="template_name cannot be empty")
    
    # Prevent changing card_type from template
    if 'card_type' in update_data and update_data['card_type'] != 'template':
        raise HTTPException(status_code=400, detail="Cannot change card_type of a template")
    
    # Update the template
    await db.jobcards.update_one({"id": template_id}, {"$set": update_data})
    await create_audit_log(current_user.id, current_user.full_name, "jobcard_template", template_id, "update", update_data)
    
    return {"message": "Template updated successfully"}

@api_router.delete("/jobcard-templates/{template_id}")
async def delete_jobcard_template(template_id: str, current_user: User = Depends(require_permission('jobcards.delete'))):
    """Delete a job card template (admin only)"""
    # Check if user is admin
    if current_user.role != 'admin':
        raise HTTPException(status_code=403, detail="Only administrators can delete templates")
    
    # Verify the template exists
    existing = await db.jobcards.find_one({"id": template_id, "card_type": "template", "is_deleted": False})
    if not existing:
        raise HTTPException(status_code=404, detail="Template not found")
    
    # Soft delete the template
    await db.jobcards.update_one(
        {"id": template_id},
        {"$set": {"is_deleted": True, "deleted_at": datetime.now(timezone.utc), "deleted_by": current_user.id}}
    )
    await create_audit_log(current_user.id, current_user.full_name, "jobcard_template", template_id, "delete")
    
    return {"message": "Template deleted successfully"}

@api_router.get("/invoices")
@limiter.limit("1000/hour")  # General authenticated rate limit: 1000 requests per hour
async def get_invoices(
    request: Request,
    page: int = 1,
    page_size: int = 10,
    current_user: User = Depends(require_permission('invoices.view'))
):
    """Get invoices with pagination support"""
    if not user_has_permission(current_user, 'invoices.view'):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="You don't have permission to view invoices")
    
    query = {"is_deleted": False}
    
    # Calculate skip value
    skip = (page - 1) * page_size
    
    # Get total count for pagination
    total_count = await db.invoices.count_documents(query)
    
    # Get paginated results
    invoices = await db.invoices.find(query, {"_id": 0}).sort("date", -1).skip(skip).limit(page_size).to_list(page_size)
    
    return create_pagination_response(invoices, total_count, page, page_size)

@api_router.get("/invoices/returnable")
async def get_returnable_invoices(
    type: str = "sales",
    current_user: User = Depends(require_permission('invoices.view'))
):
    """"
    Get finalized invoices that are available for returns.
    Only returns invoices that:
    - Are finalized (not draft)
    - Are not deleted
    - Returns allowed regardless of balance_due (even if fully paid)
    
    Args:
        type: "sales" or "purchase" to filter invoice type
    """
    if not user_has_permission(current_user, 'invoices.view'):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN, 
            detail="You don't have permission to view invoices"
        )
    
    # Build query filter
    query = {
        "is_deleted": False,
        "status": {"$in": ["finalized", "paid"]}
        # Note: Removed balance_due filter to allow returns on fully paid invoices
        # Accepts both finalized and paid status to support all returnable invoices
    }
    
    # Filter by invoice type
    if type.lower() == "sales":
        # Support both sale and service invoices
        query["invoice_type"] = {"$in": ["sale", "service"]}
    elif type.lower() == "purchase":
        query["invoice_type"] = "purchase"

    # Fetch matching invoices with required fields only
    invoices = await db.invoices.find(
        query,
        {
            "_id": 0,
            "id": 1,
            "invoice_number": 1,
            "date": 1,
            "customer_name": 1,
            "walk_in_name": 1,
            "customer_type": 1,
            "grand_total": 1,
            "balance_due": 1,
            "items": 1
        }
    ).sort("date", -1).limit(100).to_list(100)

    # Format response with party name
    formatted_invoices = []
    for inv in invoices:
        party_name = inv.get("customer_name") or inv.get("walk_in_name") or "Unknown"
        formatted_invoices.append({
            "id": inv["id"],
            "invoice_no": inv["invoice_number"],
            "date": inv["date"].isoformat() if isinstance(inv["date"], datetime) else inv["date"],
            "party_name": party_name,
            "total_amount": float(inv.get("grand_total", 0)),
            "balance_amount": float(inv.get("balance_due", 0)),
            "items": inv.get("items", [])
        })
    
    return formatted_invoices

@api_router.get("/invoices/{invoice_id}", response_model=Invoice)
async def get_invoice(invoice_id: str, current_user: User = Depends(require_permission('invoices.view'))):
    invoice = await db.invoices.find_one({"id": invoice_id, "is_deleted": False}, {"_id": 0})
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    return Invoice(**invoice)

@api_router.patch("/invoices/{invoice_id}")
async def update_invoice(invoice_id: str, update_data: dict, current_user: User = Depends(require_permission('invoices.create'))):
    if not user_has_permission(current_user, 'invoices.create'):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="You don't have permission to update invoices")
    
    existing = await db.invoices.find_one({"id": invoice_id, "is_deleted": False})
    if not existing:
        raise HTTPException(status_code=404, detail="Invoice not found")
    
    # CRITICAL: Only allow editing draft invoices - finalized invoices are immutable
    if existing.get("status") == "finalized":
        raise HTTPException(
            status_code=400, 
            detail="Cannot edit finalized invoice. Finalized invoices are immutable to maintain financial integrity."
        )
    
    # Prevent changing status through this endpoint - use /finalize endpoint instead
    if "status" in update_data:
        del update_data["status"]
    if "finalized_at" in update_data:
        del update_data["finalized_at"]
    if "finalized_by" in update_data:
        del update_data["finalized_by"]
    
    await db.invoices.update_one({"id": invoice_id}, {"$set": update_data})
    await create_audit_log(current_user.id, current_user.full_name, "invoice", invoice_id, "update", update_data)
    return {"message": "Invoice updated successfully"}


@api_router.post("/invoices/{invoice_id}/finalize")
async def finalize_invoice(invoice_id: str, current_user: User = Depends(require_permission('invoices.finalize'))):
    """
    Finalize a draft invoice - this is when all financial operations happen atomically.
    Once finalized, the invoice becomes immutable to maintain financial integrity.
    
    ⚠️ CRITICAL - AUTHORITATIVE STOCK REDUCTION PATH ⚠️
    This endpoint is the ONLY authorized way to reduce inventory stock (Stock OUT).
    Manual Stock OUT movements are prohibited to ensure:
    - Complete audit trail (all stock reductions tied to invoices)
    - Accurate accounting (revenue recorded for all stock leaving)
    - GST compliance (tax collected on all sales)
    - Financial integrity (no unauthorized stock removal)
    
    🔧 INVOICE TYPE HANDLING:
    - SALE invoices: Stock deduction REQUIRED, total weight must be > 0
    - SERVICE invoices: Stock deduction SKIPPED, zero weight allowed (making charges, repair, polish)
    
    Atomic operations performed:
    1. Update invoice status to "finalized"
    2. Create Stock OUT movements (ONLY for SALE invoices)
    3. Directly reduce inventory header (ONLY for SALE invoices)
    4. Lock linked job card (if exists)
    5. Create customer ledger entry
    6. Update customer outstanding balance
    
    All operations succeed together or fail together to maintain data consistency.
    """
    if not user_has_permission(current_user, 'invoices.finalize'):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="You don't have permission to finalize invoices")
    
    # Fetch the invoice
    existing = await db.invoices.find_one({"id": invoice_id, "is_deleted": False})
    if not existing:
        raise HTTPException(status_code=404, detail="Invoice not found")
    
    # Validate invoice is in draft state
    current_status = existing.get("status", "draft")
    if current_status == "finalized":
        raise HTTPException(
            status_code=400, 
            detail="Invoice is already finalized"
        )
    
    # Parse invoice data
    invoice = Invoice(**decimal_to_float(existing))

    # BUSINESS RULE: Stock deduction only applies to SALE invoices
    # SERVICE invoices (making charges, repair, polish) may have zero weight and should skip stock validation
    is_sale_invoice = invoice.invoice_type == "sale"
    
    # VALIDATION: For SALE invoices only, calculate total weight and ensure it's valid
    # This prevents inventory integrity issues from deducting 0g stock
    if is_sale_invoice:
        items = existing.get("items", [])
        total_weight = sum(item.get("weight", 0) * item.get("qty", 1) for item in items)

        if total_weight <= 0:
            raise HTTPException(
                status_code=400,
                detail=f"Cannot finalize sale invoice with total weight {round(total_weight, 3)}g. Sale invoice must have valid item weights for stock deduction."
            )

    # ATOMIC OPERATION: Finalize invoice with all required operations
    finalized_at = datetime.now(timezone.utc)
    
    # Step 1: Update invoice to finalized status
    await db.invoices.update_one(
        {"id": invoice_id},
        {
            "$set": {
                "status": "finalized",
                "finalized_at": finalized_at,
                "finalized_by": current_user.id
            }
        }
    )
    
    # Step 2: DIRECTLY REDUCE from inventory headers and create audit trail
    # ONLY for SALE invoices - SERVICE invoices skip stock deduction entirely
    stock_errors = []
    if is_sale_invoice:
        for item in invoice.items:
            if item.weight > 0 and item.category:
                # Find the inventory header by category name
                header = await db.inventory_headers.find_one(
                    {"name": item.category, "is_deleted": False}, 
                    {"_id": 0}
                )
                
                if header:
                    # Calculate new stock values
                    current_qty = header.get('current_qty', 0)
                    current_weight = header.get('current_weight', 0)
                    new_qty = current_qty - item.qty
                    new_weight = current_weight - item.weight
                    
                    # Check for insufficient stock
                    if new_qty < 0 or new_weight < 0:
                        stock_errors.append(
                            f"{item.category}: Need {item.qty} qty/{item.weight}g, but only {current_qty} qty/{current_weight}g available"
                        )
                        continue
                    
                    # DIRECT UPDATE: Reduce from inventory header
                    await db.inventory_headers.update_one(
                        {"id": header['id']},
                        {"$set": {"current_qty": new_qty, "current_weight": new_weight}}
                    )
                    
                    # Create stock movement for audit trail
                    movement = StockMovement(
                        movement_type="Stock OUT",
                        header_id=header['id'],
                        header_name=header['name'],
                        description=f"Invoice {invoice.invoice_number} - Finalized",
                        qty_delta=-item.qty,
                        weight_delta=-item.weight,
                        purity=item.purity,
                        reference_type="invoice",
                        reference_id=invoice.id,
                        created_by=current_user.id
                    )
                    await db.stock_movements.insert_one(movement.model_dump())
        
        # If there were stock errors, rollback the invoice finalization
        # CRITICAL: Status rollback must NOT delete timestamps (audit safety)
        # Keep finalized_at timestamp for audit trail, only change status
        if stock_errors:
            await db.invoices.update_one(
                {"id": invoice_id},
                {"$set": {"status": "draft", "finalized_by": None}}
            )
            raise HTTPException(
                status_code=400,
                detail=f"Insufficient stock: {'; '.join(stock_errors)}"
            )
    
    # Step 3: Lock the linked job card (make it read-only)
    if invoice.jobcard_id:
        jobcard = await db.jobcards.find_one({"id": invoice.jobcard_id, "is_deleted": False})
        if jobcard:
            await db.jobcards.update_one(
                {"id": invoice.jobcard_id},
                {
                    "$set": {
                        "status": "invoiced",
                        "locked": True,
                        "locked_at": finalized_at,
                        "locked_by": current_user.id
                    }
                }
            )
            await create_audit_log(
                current_user.id,
                current_user.full_name,
                "jobcard",
                invoice.jobcard_id,
                "lock",
                {"locked": True, "reason": f"Invoice {invoice.invoice_number} finalized"}
            )
    
    # Step 4: REMOVED - Invoice finalization does NOT create finance transactions
    # Financial transactions are ONLY created when PAYMENT is received
    # This ensures correct accounting: invoices do not move money, payments do
    
    # Step 5: Outstanding balance is automatically updated
    # The balance_due field in the invoice record already tracks outstanding amount
    # Party ledger calculations aggregate all invoice balance_due values
    
    # Create audit log for finalization
    await create_audit_log(
        current_user.id, 
        current_user.full_name, 
        "invoice", 
        invoice.id, 
        "finalize", 
        {
            "status": "finalized", 
            "finalized_at": finalized_at.isoformat(),
            "jobcard_locked": bool(invoice.jobcard_id),
            "ledger_entry_created": False,  # No ledger entry on finalization
            "customer_type": invoice.customer_type,
            "note": "Finance transactions NOT created - only on payment"
        }
    )
    
    # Fetch and return updated invoice
    updated_invoice = await db.invoices.find_one({"id": invoice_id}, {"_id": 0})
    return decimal_to_float(updated_invoice)

@api_router.post("/invoices/{invoice_id}/add-payment")
async def add_payment_to_invoice(
    invoice_id: str, 
    payment_data: dict, 
    current_user: User = Depends(require_permission('invoices.create'))
):
    """
    Add payment to an invoice and create a transaction record.
    
    MODULE 10: GOLD_EXCHANGE Payment Mode
    When payment_mode = "GOLD_EXCHANGE":
    - Required: gold_weight_grams, rate_per_gram
    - Auto-calculates: gold_money_value = gold_weight_grams × rate_per_gram
    - Creates GoldLedgerEntry (type=OUT - customer uses their gold to pay)
    - Creates Transaction record for financial trace
    - Updates invoice paid_amount and balance_due
    - Only works for saved customers (requires party_id for gold ledger)
    
    For other payment modes (Cash, Bank Transfer, Card, UPI/Online, Cheque):
    - Required fields: amount, payment_mode, account_id
    - Optional: notes
    
    For walk-in customers: Recommend full payment but allow partial
    For saved customers: Allow partial payments (outstanding tracked in ledger)
    """
    try:
        # Fetch invoice first to determine payment mode requirements
        existing = await db.invoices.find_one({"id": invoice_id, "is_deleted": False})
        if not existing:
            raise HTTPException(status_code=404, detail="Invoice not found")
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error fetching invoice {invoice_id}: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Database error: {str(e)}")
    
    invoice = Invoice(**decimal_to_float(existing))
    
    # Check payment mode
    payment_mode = payment_data.get('payment_mode')
    if not payment_mode:
        raise HTTPException(status_code=400, detail="Payment mode is required")
    
    # MODULE 10: Handle GOLD_EXCHANGE payment mode
    if payment_mode == "GOLD_EXCHANGE":
        # Validate GOLD_EXCHANGE is only for saved customers (need party_id for gold ledger)
        if invoice.customer_type != "saved" or not invoice.customer_id:
            raise HTTPException(
                status_code=400, 
                detail="GOLD_EXCHANGE payment mode is only available for saved customers (not walk-in)"
            )
        
        # Validate required fields for GOLD_EXCHANGE
        gold_weight_grams = payment_data.get('gold_weight_grams')
        rate_per_gram = payment_data.get('rate_per_gram')
        
        if not gold_weight_grams or gold_weight_grams <= 0:
            raise HTTPException(status_code=400, detail="gold_weight_grams must be greater than 0 for GOLD_EXCHANGE mode")
        
        if not rate_per_gram or rate_per_gram <= 0:
            raise HTTPException(status_code=400, detail="rate_per_gram must be greater than 0 for GOLD_EXCHANGE mode")
        
        # Round to proper precision
        gold_weight_grams = round(float(gold_weight_grams), 3)  # 3 decimals for gold
        rate_per_gram = round(float(rate_per_gram), 2)  # 2 decimals for money
        
        # Auto-calculate gold money value
        gold_money_value = round(gold_weight_grams * rate_per_gram, 2)
        
        # Calculate new paid amount and balance
        payment_amount = gold_money_value
        new_paid_amount = invoice.paid_amount + payment_amount
        new_balance_due = invoice.grand_total - new_paid_amount
        
        # Validate payment doesn't exceed balance
        if new_balance_due < -0.01:  # Allow small rounding errors
            raise HTTPException(
                status_code=400, 
                detail=f"Gold exchange value ({payment_amount:.2f} OMR) exceeds remaining balance ({invoice.balance_due:.2f} OMR)"
            )
        
        # Check if customer has sufficient gold balance
        gold_in_pipeline = db.gold_ledger.aggregate([
            {
                "$match": {
                    "party_id": invoice.customer_id,
                    "is_deleted": False
                }
            },
            {
                "$group": {
                    "_id": None,
                    "gold_in": {
                        "$sum": {
                            "$cond": [{"$eq": ["$type", "IN"]}, "$weight_grams", 0]
                        }
                    },
                    "gold_out": {
                        "$sum": {
                            "$cond": [{"$eq": ["$type", "OUT"]}, "$weight_grams", 0]
                        }
                    }
                }
            }
        ])
        
        gold_summary_list = await gold_in_pipeline.to_list(length=1)
        if gold_summary_list:
            gold_in = gold_summary_list[0].get("gold_in", 0)
            gold_out = gold_summary_list[0].get("gold_out", 0)
            gold_balance = round(gold_in - gold_out, 3)
        else:
            gold_balance = 0
        
        # Validate customer has sufficient gold
        if gold_balance < gold_weight_grams:
            raise HTTPException(
                status_code=400,
                detail=f"Insufficient gold balance. Customer has {gold_balance:.3f}g available, but {gold_weight_grams:.3f}g requested for payment"
            )
        
        # Get default purity (916 - 22K gold standard)
        purity_entered = payment_data.get('purity_entered', 916)
        
        # Create GoldLedgerEntry (type=OUT - customer uses their gold, shop returns/settles gold)
        gold_ledger_entry = GoldLedgerEntry(
            party_id=invoice.customer_id,
            type="OUT",  # OUT = shop gives gold to party (or party uses their gold with shop)
            weight_grams=gold_weight_grams,
            purity_entered=purity_entered,
            purpose="exchange",  # Exchange purpose for payment settlement
            reference_type="invoice",
            reference_id=invoice_id,
            notes=f"Gold exchange payment for invoice {invoice.invoice_number}. Rate: {rate_per_gram:.2f} OMR/g",
            created_by=current_user.id
        )
        
        # Insert gold ledger entry
        await db.gold_ledger.insert_one(gold_ledger_entry.model_dump())
        
        # Fetch or create default account for gold exchange transactions
        account = await db.accounts.find_one({"name": "Gold Exchange Income", "is_deleted": False}, {"_id": 0})
        if not account:
            # Create default Gold Exchange Income account (INCOME type, not asset)
            account = {
                "id": str(uuid.uuid4()),
                "name": "Gold Exchange Income",
                "account_type": "income",
                "opening_balance": 0,
                "current_balance": 0,
                "created_at": datetime.now(timezone.utc),
                "created_by": current_user.id,
                "is_deleted": False
            }
            await db.accounts.insert_one(account)
        
        account_id = account['id']
        account_name = account['name']
        
        # Determine party details
        party_id = invoice.customer_id
        party_name = invoice.customer_name or "Unknown Customer"
        
        # Generate transaction number
        year = datetime.now(timezone.utc).year
        count = await db.transactions.count_documents({"transaction_number": {"$regex": f"^TXN-{year}"}})
        transaction_number = f"TXN-{year}-{str(count + 1).zfill(4)}"
        
        # Generate transaction numbers for double-entry
        year = datetime.now(timezone.utc).year
        count = await db.transactions.count_documents({"transaction_number": {"$regex": f"^TXN-{year}"}})
        debit_txn_number = f"TXN-{year}-{str(count + 1).zfill(4)}"
        credit_txn_number = f"TXN-{year}-{str(count + 2).zfill(4)}"
        
        # DOUBLE-ENTRY BOOKKEEPING FOR GOLD EXCHANGE:
        # Note: Gold Exchange is tracked in Gold Ledger separately
        # For financial accounting, we only record the money equivalent
        
        # For now, use a simplified approach: Credit Gold Exchange Income account
        # In future, could track: Debit: Gold Asset, Credit: Gold Exchange Income
        
        # Transaction 1: CREDIT Gold Exchange Income (INCOME) - Revenue from gold received
        credit_transaction = Transaction(
            transaction_number=credit_txn_number,
            transaction_type="credit",  # Credit increases income
            mode="GOLD_EXCHANGE",
            account_id=account_id,
            account_name=account_name,
            party_id=party_id,
            party_name=party_name,
            amount=payment_amount,
            category="Invoice Payment - Gold Exchange Income (Credit)",
            notes=f"Gold exchange revenue for {invoice.invoice_number}. {gold_weight_grams:.3f}g @ {rate_per_gram:.2f} OMR/g = {payment_amount:.2f} OMR. {payment_data.get('notes', '')}".strip(),
            reference_type="invoice",
            reference_id=invoice_id,
            created_by=current_user.id
        )
        
        # Insert credit transaction
        await db.transactions.insert_one(credit_transaction.model_dump())
        
        # Update Gold Exchange Income account balance (increase for credit on income)
        await db.accounts.update_one(
            {"id": account_id},
            {"$inc": {"current_balance": payment_amount}}
        )
        
        # Update invoice payment details
        new_payment_status = "paid" if new_balance_due < 0.01 else "partial"
        
        # Prepare update data
        update_data = {
            "paid_amount": new_paid_amount,
            "balance_due": max(0, new_balance_due),
            "payment_status": new_payment_status
        }
        
        # Set paid_at timestamp when invoice becomes fully paid (immutability - set only once)
        if new_payment_status == "paid" and not existing.get("paid_at"):
            update_data["paid_at"] = datetime.now(timezone.utc)
        
        # AUTO-FINALIZE: If invoice is draft and becomes fully paid, auto-finalize it
        # This prevents the "draft + paid" state which breaks delivery rules
        current_status = existing.get("status", "draft")
        if new_payment_status == "paid" and current_status == "draft":
            finalized_at = datetime.now(timezone.utc)
            update_data["status"] = "finalized"
            update_data["finalized_at"] = finalized_at
            update_data["finalized_by"] = current_user.id
            
            # Perform finalization operations atomically
            # 1. Lock the linked job card (if exists)
            if invoice.jobcard_id:
                jobcard = await db.jobcards.find_one({"id": invoice.jobcard_id, "is_deleted": False})
                if jobcard:
                    await db.jobcards.update_one(
                        {"id": invoice.jobcard_id},
                        {
                            "$set": {
                                "status": "invoiced",
                                "locked": True,
                                "locked_at": finalized_at,
                                "locked_by": current_user.id
                            }
                        }
                    )
                    await create_audit_log(
                        current_user.id,
                        current_user.full_name,
                        "jobcard",
                        invoice.jobcard_id,
                        "lock",
                        {"locked": True, "reason": f"Invoice {invoice.invoice_number} auto-finalized due to full gold exchange payment"}
                    )
            
            # 2. Perform stock deduction for SALE invoices
            is_sale_invoice = invoice.invoice_type == "sale"
            gold_stock_errors = []
            
            if is_sale_invoice:
                for item in invoice.items:
                    if item.weight > 0 and item.category:
                        # Find the inventory header by category name
                        header = await db.inventory_headers.find_one(
                            {"name": item.category, "is_deleted": False}, 
                            {"_id": 0}
                        )
                        
                        if header:
                            # Calculate new stock values
                            current_qty = header.get('current_qty', 0)
                            current_weight = header.get('current_weight', 0)
                            new_qty = current_qty - item.qty
                            new_weight = current_weight - item.weight
                            
                            # Check for insufficient stock
                            if new_qty < 0 or new_weight < 0:
                                gold_stock_errors.append(
                                    f"{item.category}: Need {item.qty} qty/{item.weight}g, but only {current_qty} qty/{current_weight}g available"
                                )
                                continue
                            
                            # DIRECT UPDATE: Reduce from inventory header
                            await db.inventory_headers.update_one(
                                {"id": header['id']},
                                {"$set": {"current_qty": new_qty, "current_weight": new_weight}}
                            )
                            
                            # Create stock movement for audit trail
                            movement = StockMovement(
                                movement_type="Stock OUT",
                                header_id=header['id'],
                                header_name=header['name'],
                                description=f"Invoice {invoice.invoice_number} - Auto-finalized (Gold Exchange Payment)",
                                qty_delta=-item.qty,
                                weight_delta=-item.weight,
                                purity=item.purity,
                                reference_type="invoice",
                                reference_id=invoice.id,
                                created_by=current_user.id
                            )
                            await db.stock_movements.insert_one(movement.model_dump())
            
            # If stock errors occurred, rollback finalization but keep payment
            if gold_stock_errors:
                # Rollback finalization status but keep payment information
                update_data["status"] = "draft"
                del update_data["finalized_at"]
                del update_data["finalized_by"]
                
                # Unlock the job card if it was locked
                if invoice.jobcard_id:
                    await db.jobcards.update_one(
                        {"id": invoice.jobcard_id},
                        {
                            "$set": {
                                "locked": False,
                                "locked_at": None,
                                "locked_by": None
                            }
                        }
                    )
                
                # Log the stock error but don't fail the payment
                await create_audit_log(
                    current_user.id,
                    current_user.full_name,
                    "invoice",
                    invoice_id,
                    "auto_finalize_failed",
                    {
                        "reason": "insufficient_stock",
                        "errors": gold_stock_errors,
                        "payment_applied": True,
                        "payment_mode": "GOLD_EXCHANGE"
                    }
                )
            else:
                # 3. NO ledger entry creation on auto-finalization (gold exchange)
                # Payment transaction was ALREADY created with proper double-entry above
                # Auto-finalization only handles stock and job card locking
                
                # Create audit log for auto-finalization
                await create_audit_log(
                    current_user.id,
                    current_user.full_name,
                    "invoice",
                    invoice.id,
                    "auto_finalize",
                    {
                        "status": "finalized",
                        "finalized_at": finalized_at.isoformat(),
                        "jobcard_locked": bool(invoice.jobcard_id),
                        "ledger_entry_created": False,  # No extra ledger entry - payment already created double-entry
                        "customer_type": invoice.customer_type,
                        "trigger": "full_gold_exchange_payment_received"
                    }
                )
        
        await db.invoices.update_one(
            {"id": invoice_id},
            {"$set": update_data}
        )
        
        # Create audit logs
        await create_audit_log(
            current_user.id,
            current_user.full_name,
            "gold_ledger",
            gold_ledger_entry.id,
            "create",
            {
                "invoice_id": invoice_id,
                "gold_weight_grams": gold_weight_grams,
                "rate_per_gram": rate_per_gram,
                "gold_money_value": payment_amount
            }
        )
        
        await create_audit_log(
            current_user.id,
            current_user.full_name,
            "transaction",
            credit_transaction.id,
            "create",
            {
                "invoice_id": invoice_id,
                "payment_mode": "GOLD_EXCHANGE",
                "transaction_type": "credit",
                "gold_weight_grams": gold_weight_grams,
                "rate_per_gram": rate_per_gram,
                "payment_amount": payment_amount
            }
        )
        
        await create_audit_log(
            current_user.id,
            current_user.full_name,
            "invoice",
            invoice_id,
            "add_payment_gold_exchange",
            {
                "gold_weight_grams": gold_weight_grams,
                "rate_per_gram": rate_per_gram,
                "gold_money_value": payment_amount,
                "new_paid_amount": new_paid_amount,
                "new_balance_due": max(0, new_balance_due),
                "gold_balance_before": gold_balance,
                "gold_balance_after": round(gold_balance - gold_weight_grams, 3)
            }
        )
        
        # Return success response with GOLD_EXCHANGE details
        return {
            "message": "Gold exchange payment added successfully (revenue recognized)",
            "transaction_id": credit_transaction.id,
            "transaction_number": credit_txn_number,
            "gold_ledger_entry_id": gold_ledger_entry.id,
            "gold_weight_grams": gold_weight_grams,
            "rate_per_gram": rate_per_gram,
            "gold_money_value": payment_amount,
            "new_paid_amount": new_paid_amount,
            "new_balance_due": max(0, new_balance_due),
            "payment_status": new_payment_status,
            "customer_gold_balance_remaining": round(gold_balance - gold_weight_grams, 3)
        }
    
    # Standard payment modes (Cash, Bank Transfer, Card, UPI/Online, Cheque)
    else:
        # Validate required fields for standard payment modes
        if not payment_data.get('amount') or payment_data['amount'] <= 0:
            raise HTTPException(status_code=400, detail="Payment amount must be greater than 0")
        
        if not payment_data.get('account_id'):
            raise HTTPException(status_code=400, detail="Account ID is required")
        
        # Calculate new paid amount and balance
        payment_amount = float(payment_data['amount'])
        new_paid_amount = invoice.paid_amount + payment_amount
        new_balance_due = invoice.grand_total - new_paid_amount
        
        # Validate payment doesn't exceed balance
        if new_balance_due < -0.01:  # Allow small rounding errors
            raise HTTPException(
                status_code=400, 
                detail=f"Payment amount ({payment_amount}) exceeds remaining balance ({invoice.balance_due})"
            )
        
        # Fetch account
        try:
            account = await db.accounts.find_one({"id": payment_data['account_id'], "is_deleted": False}, {"_id": 0})
            if not account:
                logger.error(f"Account not found: {payment_data['account_id']}")
                raise HTTPException(status_code=404, detail=f"Account with ID {payment_data['account_id']} not found")
        except HTTPException:
            raise
        except Exception as e:
            logger.error(f"Error fetching account: {str(e)}")
            raise HTTPException(status_code=500, detail=f"Database error while fetching account: {str(e)}")
        
        # Determine party details based on customer type
        party_id = None
        party_name = ""
        
        if invoice.customer_type == "saved":
            party_id = invoice.customer_id
            party_name = invoice.customer_name or "Unknown Customer"
        else:  # walk_in
            party_id = None
            party_name = f"{invoice.walk_in_name or 'Walk-in Customer'} (Walk-in)"
        
        # Generate transaction numbers for double-entry
        year = datetime.now(timezone.utc).year
        count = await db.transactions.count_documents({"transaction_number": {"$regex": f"^TXN-{year}"}})
        debit_txn_number = f"TXN-{year}-{str(count + 1).zfill(4)}"
        credit_txn_number = f"TXN-{year}-{str(count + 2).zfill(4)}"
        
        # DOUBLE-ENTRY BOOKKEEPING:
        # Transaction 1: DEBIT Cash/Bank (ASSET) - Money increases in Cash/Bank
        debit_transaction = Transaction(
            transaction_number=debit_txn_number,
            transaction_type="debit",  # Debit increases asset
            mode=payment_data['payment_mode'],
            account_id=payment_data['account_id'],
            account_name=account['name'],
            party_id=party_id,
            party_name=party_name,
            amount=payment_amount,
            category="Invoice Payment - Cash/Bank (Debit)",
            notes=f"Payment for {invoice.invoice_number}. {payment_data.get('notes', '')}".strip(),
            reference_type="invoice",
            reference_id=invoice_id,
            created_by=current_user.id
        )
        
        # Insert debit transaction
        await db.transactions.insert_one(debit_transaction.model_dump())
        
        # Update Cash/Bank account balance (increase for debit on asset)
        await db.accounts.update_one(
            {"id": payment_data['account_id']},
            {"$inc": {"current_balance": payment_amount}}
        )
        
        # Transaction 2: CREDIT Sales Income (INCOME) - Revenue recognized
        # Get or create Sales Income account
        sales_account = await db.accounts.find_one({"name": "Sales Income", "is_deleted": False})
        if not sales_account:
            sales_account = {
                "id": str(uuid.uuid4()),
                "name": "Sales Income",
                "account_type": "income",
                "opening_balance": 0,
                "current_balance": 0,
                "created_at": datetime.now(timezone.utc),
                "created_by": current_user.id,
                "is_deleted": False
            }
            await db.accounts.insert_one(sales_account)
        
        credit_transaction = Transaction(
            transaction_number=credit_txn_number,
            transaction_type="credit",  # Credit increases income
            mode=payment_data['payment_mode'],
            account_id=sales_account['id'],
            account_name=sales_account['name'],
            party_id=party_id,
            party_name=party_name,
            amount=payment_amount,
            category="Invoice Payment - Sales Income (Credit)",
            notes=f"Revenue for {invoice.invoice_number}. {payment_data.get('notes', '')}".strip(),
            reference_type="invoice",
            reference_id=invoice_id,
            created_by=current_user.id
        )
        
        # Insert credit transaction
        await db.transactions.insert_one(credit_transaction.model_dump())
        
        # Update Sales Income account balance (increase for credit on income)
        await db.accounts.update_one(
            {"id": sales_account['id']},
            {"$inc": {"current_balance": payment_amount}}
        )
        
        # Update invoice payment details
        new_payment_status = "paid" if new_balance_due < 0.01 else "partial"
        
        # Prepare update data
        update_data = {
            "paid_amount": new_paid_amount,
            "balance_due": max(0, new_balance_due),  # Ensure no negative balance
            "payment_status": new_payment_status
        }
        
        # Set paid_at timestamp when invoice becomes fully paid (immutability - set only once)
        if new_payment_status == "paid" and not existing.get("paid_at"):
            update_data["paid_at"] = datetime.now(timezone.utc)
        
        # AUTO-FINALIZE: If invoice is draft and becomes fully paid, auto-finalize it
        # This prevents the "draft + paid" state which breaks delivery rules
        current_status = existing.get("status", "draft")
        if new_payment_status == "paid" and current_status == "draft":
            finalized_at = datetime.now(timezone.utc)
            update_data["status"] = "finalized"
            update_data["finalized_at"] = finalized_at
            update_data["finalized_by"] = current_user.id
            
            # Perform finalization operations atomically
            # 1. Lock the linked job card (if exists)
            if invoice.jobcard_id:
                jobcard = await db.jobcards.find_one({"id": invoice.jobcard_id, "is_deleted": False})
                if jobcard:
                    await db.jobcards.update_one(
                        {"id": invoice.jobcard_id},
                        {
                            "$set": {
                                "status": "invoiced",
                                "locked": True,
                                "locked_at": finalized_at,
                                "locked_by": current_user.id
                            }
                        }
                    )
                    await create_audit_log(
                        current_user.id,
                        current_user.full_name,
                        "jobcard",
                        invoice.jobcard_id,
                        "lock",
                        {"locked": True, "reason": f"Invoice {invoice.invoice_number} auto-finalized due to full payment"}
                    )
            
            # 2. Perform stock deduction for SALE invoices
            is_sale_invoice = invoice.invoice_type == "sale"
            stock_errors = []
            
            if is_sale_invoice:
                for item in invoice.items:
                    if item.weight > 0 and item.category:
                        # Find the inventory header by category name
                        header = await db.inventory_headers.find_one(
                            {"name": item.category, "is_deleted": False}, 
                            {"_id": 0}
                        )
                        
                        if header:
                            # Calculate new stock values
                            current_qty = header.get('current_qty', 0)
                            current_weight = header.get('current_weight', 0)
                            new_qty = current_qty - item.qty
                            new_weight = current_weight - item.weight
                            
                            # Check for insufficient stock
                            if new_qty < 0 or new_weight < 0:
                                stock_errors.append(
                                    f"{item.category}: Need {item.qty} qty/{item.weight}g, but only {current_qty} qty/{current_weight}g available"
                                )
                                continue
                            
                            # DIRECT UPDATE: Reduce from inventory header
                            await db.inventory_headers.update_one(
                                {"id": header['id']},
                                {"$set": {"current_qty": new_qty, "current_weight": new_weight}}
                            )
                            
                            # Create stock movement for audit trail
                            movement = StockMovement(
                                movement_type="Stock OUT",
                                header_id=header['id'],
                                header_name=header['name'],
                                description=f"Invoice {invoice.invoice_number} - Auto-finalized (Payment)",
                                qty_delta=-item.qty,
                                weight_delta=-item.weight,
                                purity=item.purity,
                                reference_type="invoice",
                                reference_id=invoice.id,
                                created_by=current_user.id
                            )
                            await db.stock_movements.insert_one(movement.model_dump())
            
            # If stock errors occurred, rollback finalization but keep payment
            if stock_errors:
                # Rollback finalization status but keep payment information
                update_data["status"] = "draft"
                del update_data["finalized_at"]
                del update_data["finalized_by"]
                
                # Unlock the job card if it was locked
                if invoice.jobcard_id:
                    await db.jobcards.update_one(
                        {"id": invoice.jobcard_id},
                        {
                            "$set": {
                                "locked": False,
                                "locked_at": None,
                                "locked_by": None
                            }
                        }
                    )
                
                # Log the stock error but don't fail the payment
                await create_audit_log(
                    current_user.id,
                    current_user.full_name,
                    "invoice",
                    invoice_id,
                    "auto_finalize_failed",
                    {
                        "reason": "insufficient_stock",
                        "errors": stock_errors,
                        "payment_applied": True
                    }
                )
            else:
                # 3. NO ledger entry creation on auto-finalization
                # Payment transaction was ALREADY created with proper double-entry above
                # Auto-finalization only handles stock and job card locking
                
                # Create audit log for auto-finalization
                await create_audit_log(
                    current_user.id,
                    current_user.full_name,
                    "invoice",
                    invoice.id,
                    "auto_finalize",
                    {
                        "status": "finalized",
                        "finalized_at": finalized_at.isoformat(),
                        "jobcard_locked": bool(invoice.jobcard_id),
                        "ledger_entry_created": False,  # No extra ledger entry - payment already created double-entry
                        "customer_type": invoice.customer_type,
                        "trigger": "full_payment_received"
                    }
                )
        
        await db.invoices.update_one(
            {"id": invoice_id},
            {"$set": update_data}
        )
        
        # Create audit logs for both transactions (double-entry)
        await create_audit_log(
            current_user.id,
            current_user.full_name,
            "transaction",
            debit_transaction.id,
            "create",
            {
                "invoice_id": invoice_id, 
                "payment_amount": payment_amount,
                "transaction_type": "debit",
                "account": account['name']
            }
        )
        
        await create_audit_log(
            current_user.id,
            current_user.full_name,
            "transaction",
            credit_transaction.id,
            "create",
            {
                "invoice_id": invoice_id, 
                "payment_amount": payment_amount,
                "transaction_type": "credit",
                "account": "Sales Income"
            }
        )
        
        await create_audit_log(
            current_user.id,
            current_user.full_name,
            "invoice",
            invoice_id,
            "add_payment",
            {
                "amount": payment_amount,
                "new_paid_amount": new_paid_amount,
                "new_balance_due": max(0, new_balance_due),
                "payment_mode": payment_data['payment_mode'],
                "double_entry": True
            }
        )
        
        # Return success response with updated invoice details
        return {
            "message": "Payment added successfully (double-entry recorded)",
            "debit_transaction_id": debit_transaction.id,
            "credit_transaction_id": credit_transaction.id,
            "debit_transaction_number": debit_txn_number,
            "credit_transaction_number": credit_txn_number,
            "new_paid_amount": new_paid_amount,
            "new_balance_due": max(0, new_balance_due),
            "payment_status": new_payment_status,
            "is_walk_in_partial_payment": invoice.customer_type == "walk_in" and new_balance_due > 0.01
        }


@api_router.get("/invoices/{invoice_id}/impact")
async def get_invoice_impact(invoice_id: str, current_user: User = Depends(require_permission('invoices.view'))):
    """
    Get impact summary for invoice actions (finalization or deletion).
    Shows decision-critical data: items, totals, linked data, what will be locked.
    """
    invoice = await db.invoices.find_one({"id": invoice_id, "is_deleted": False}, {"_id": 0})
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    
    # Get items breakdown
    items = invoice.get("items", [])
    total_items = len(items)
    total_weight = sum(item.get("weight", 0) for item in items)
    subtotal = invoice.get("subtotal", 0)
    vat_total = invoice.get("vat_total", 0)
    discount_amount = invoice.get("discount_amount", 0)
    grand_total = invoice.get("grand_total", 0)
    paid_amount = invoice.get("paid_amount", 0)
    balance_due = invoice.get("balance_due", 0)
    
    # Check linked job card
    linked_jobcard = None
    if invoice.get("jobcard_id"):
        jobcard = await db.jobcards.find_one(
            {"id": invoice["jobcard_id"], "is_deleted": False},
            {"_id": 0, "id": 1, "job_card_number": 1, "status": 1}
        )
        if jobcard:
            linked_jobcard = jobcard
    
    # Build impact summary
    impact = {
        "invoice_number": invoice.get("invoice_number"),
        "current_status": invoice.get("status", "draft"),
        "customer_name": invoice.get("customer_name"),
        "total_items": total_items,
        "total_weight_grams": round(total_weight, 3),
        "subtotal": round(subtotal, 2),
        "vat_total": round(vat_total, 2),
        "discount_amount": round(discount_amount, 2),
        "grand_total": round(grand_total, 2),
        "paid_amount": round(paid_amount, 2),
        "balance_due": round(balance_due, 2),
        "payment_status": invoice.get("payment_status", "unpaid"),
        "has_linked_jobcard": linked_jobcard is not None,
        "linked_jobcard": linked_jobcard if linked_jobcard else None,
        "will_lock_jobcard": linked_jobcard is not None and invoice.get("status") == "draft",
        "finalized_at": invoice.get("finalized_at")
    }
    
    return impact

@api_router.delete("/invoices/{invoice_id}")
async def delete_invoice(invoice_id: str, current_user: User = Depends(require_permission('invoices.delete'))):
    if not user_has_permission(current_user, 'invoices.delete'):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="You don't have permission to delete invoices")
    
    existing = await db.invoices.find_one({"id": invoice_id, "is_deleted": False})
    if not existing:
        raise HTTPException(status_code=404, detail="Invoice not found")
    
    # CRITICAL: Only allow deleting draft invoices
    # Finalized invoices should not be deleted to maintain financial integrity
    if existing.get("status") == "finalized":
        raise HTTPException(
            status_code=400, 
            detail="Cannot delete finalized invoice. Finalized invoices are immutable to maintain financial integrity."
        )
    
    await db.invoices.update_one(
        {"id": invoice_id},
        {"$set": {"is_deleted": True}}
    )
    await create_audit_log(current_user.id, current_user.full_name, "invoice", invoice_id, "delete")
    return {"message": "Invoice deleted successfully"}

@api_router.get("/invoices/{invoice_id}/pdf")
async def generate_invoice_pdf(invoice_id: str, current_user: User = Depends(require_permission('invoices.view'))):
    from fastapi.responses import StreamingResponse
    from io import BytesIO
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.units import inch
    from reportlab.pdfgen import canvas
    from reportlab.lib import colors
    from reportlab.platypus import Table, TableStyle
    
    invoice = await db.invoices.find_one({"id": invoice_id, "is_deleted": False}, {"_id": 0})
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    
    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    
    # Header
    p.setFont("Helvetica-Bold", 20)
    p.drawString(50, height - 50, "Gold Shop ERP")
    p.setFont("Helvetica", 10)
    p.drawString(50, height - 70, "The Artisan Ledger")
    
    # Invoice details
    p.setFont("Helvetica-Bold", 16)
    p.drawString(50, height - 120, f"Invoice #{invoice.get('invoice_number', '')}")
    p.setFont("Helvetica", 10)
    invoice_date = invoice.get('date', '')
    if isinstance(invoice_date, str):
        date_str = invoice_date[:10]
    else:
        date_str = str(invoice_date)[:10]
    p.drawString(50, height - 140, f"Date: {date_str}")
    p.drawString(50, height - 155, f"Customer: {invoice.get('customer_name', 'N/A')}")
    p.drawString(50, height - 170, f"Type: {invoice.get('invoice_type', 'sale').upper()}")
    p.drawString(50, height - 185, f"Status: {invoice.get('payment_status', 'unpaid').upper()}")
    
    # Items table
    y_position = height - 230
    p.setFont("Helvetica-Bold", 10)
    p.drawString(50, y_position, "Item")
    p.drawString(250, y_position, "Qty")
    p.drawString(300, y_position, "Weight")
    p.drawString(370, y_position, "Rate")
    p.drawString(450, y_position, "Total")
    
    p.setFont("Helvetica", 9)
    y_position -= 20
    
    for item in invoice.get('items', []):
        p.drawString(50, y_position, item.get('description', '')[:30])
        p.drawString(250, y_position, str(item.get('qty', 0)))
        p.drawString(300, y_position, f"{item.get('weight', 0)}g")
        p.drawString(370, y_position, f"{item.get('metal_rate', 0):.2f}")
        p.drawString(450, y_position, f"{item.get('line_total', 0):.2f}")
        y_position -= 15
        
        if y_position < 100:
            p.showPage()
            y_position = height - 50
    
    # Totals
    y_position -= 20
    p.setFont("Helvetica-Bold", 10)
    p.drawString(370, y_position, "Subtotal:")
    p.drawString(450, y_position, f"{invoice.get('subtotal', 0):.2f} OMR")
    
    # MODULE 7: Add discount line if discount exists
    discount_amount = invoice.get('discount_amount', 0)
    if discount_amount > 0:
        y_position -= 15
        p.setFont("Helvetica", 10)
        p.drawString(370, y_position, "Discount:")
        p.drawString(450, y_position, f"-{discount_amount:.2f} OMR")
    
    y_position -= 15
    p.setFont("Helvetica-Bold", 10)
    p.drawString(370, y_position, "VAT:")
    p.drawString(450, y_position, f"{invoice.get('vat_total', 0):.2f} OMR")
    y_position -= 15
    p.setFont("Helvetica-Bold", 12)
    p.drawString(370, y_position, "Grand Total:")
    p.drawString(450, y_position, f"{invoice.get('grand_total', 0):.2f} OMR")
    y_position -= 15
    p.setFont("Helvetica", 10)
    p.drawString(370, y_position, "Balance Due:")
    p.drawString(450, y_position, f"{invoice.get('balance_due', 0):.2f} OMR")
    
    # Footer
    p.setFont("Helvetica-Oblique", 8)
    p.drawString(50, 50, "Thank you for your business!")
    
    p.save()
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=invoice_{invoice.get('invoice_number', 'unknown')}.pdf"}
    )

@api_router.get("/invoices/{invoice_id}/full-details")
async def get_invoice_full_details(invoice_id: str, current_user: User = Depends(require_permission('invoices.view'))):
    """
    Get invoice with full details including payment transactions for professional invoice printing
    """
    invoice = await db.invoices.find_one({"id": invoice_id, "is_deleted": False}, {"_id": 0})
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    
    # Get all payment transactions for this invoice
    payments = await db.transactions.find({
        "reference_type": "invoice",
        "reference_id": invoice_id,
        "is_deleted": False
    }, {"_id": 0}).to_list(length=100)
    
    # Get customer details if saved customer
    customer_details = None
    if invoice.get('customer_type') == 'saved' and invoice.get('customer_id'):
        party = await db.parties.find_one({"id": invoice['customer_id'], "is_deleted": False}, {"_id": 0})
        if party:
            customer_details = {
                "name": party.get('name'),
                "phone": party.get('phone'),
                "address": party.get('address'),
                "gstin": party.get('gstin')  # If you add GSTIN field to Party model
            }
    
    return {
        "invoice": Invoice(**invoice),
        "payments": payments,
        "customer_details": customer_details
    }

@api_router.get("/settings/shop")
async def get_shop_settings(current_user: User = Depends(get_current_user)):
    """
    Get shop settings for invoice printing. Returns placeholder data if not configured.
    """
    settings = await db.shop_settings.find_one({}, {"_id": 0})
    if not settings:
        # Return default placeholder settings
        default_settings = ShopSettings()
        return default_settings
    return ShopSettings(**settings)

@api_router.put("/settings/shop")
async def update_shop_settings(settings_data: dict, current_user: User = Depends(require_permission('users.update'))):
    """
    Update shop settings for invoice printing (admin only)
    """
    settings_data['updated_at'] = datetime.now(timezone.utc)
    settings_data['updated_by'] = current_user.id
    
    existing = await db.shop_settings.find_one({})
    if existing:
        await db.shop_settings.update_one({"id": existing['id']}, {"$set": settings_data})
    else:
        # Create new settings
        new_settings = ShopSettings(**settings_data)
        await db.shop_settings.insert_one(new_settings.model_dump())
    
    await create_audit_log(current_user.id, current_user.full_name, "settings", "shop_settings", "update", settings_data)
    return {"message": "Shop settings updated successfully"}

@api_router.post("/invoices", response_model=Invoice)
async def create_invoice(invoice_data: dict, current_user: User = Depends(require_permission('invoices.create'))):
    if not user_has_permission(current_user, 'invoices.create'):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="You don't have permission to create invoices")
    
    year = datetime.now(timezone.utc).year
    count = await db.invoices.count_documents({"invoice_number": {"$regex": f"^INV-{year}"}})
    invoice_number = f"INV-{year}-{str(count + 1).zfill(4)}"
    
    # Remove conflicting keys and add required fields
    invoice_data_clean = {k: v for k, v in invoice_data.items() if k not in ['invoice_number', 'created_by']}
    invoice = Invoice(**invoice_data_clean, invoice_number=invoice_number, created_by=current_user.id)
    await db.invoices.insert_one(invoice.model_dump())
    
    # Stock movements will ONLY happen when invoice is finalized via /invoices/{id}/finalize endpoint
    
    await create_audit_log(current_user.id, current_user.full_name, "invoice", invoice.id, "create")
    return invoice

@api_router.get("/accounts", response_model=List[Account])
async def get_accounts(current_user: User = Depends(require_permission('finance.view'))):
    if not user_has_permission(current_user, 'finance.view'):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="You don't have permission to view finance data")
    
    accounts = await db.accounts.find({"is_deleted": False}, {"_id": 0}).to_list(1000)
    return accounts

@api_router.get("/accounts/{account_id}", response_model=Account)
async def get_account(account_id: str, current_user: User = Depends(require_permission('finance.view'))):
    """Get a single account by ID"""
    account = await db.accounts.find_one({"id": account_id, "is_deleted": False}, {"_id": 0})
    if not account:
        raise HTTPException(status_code=404, detail="Account not found")
    return account

@api_router.post("/accounts", status_code=201)
async def create_account(account_data: dict, current_user: User = Depends(require_permission('finance.create'))):
    """
    Create a new account with strict type validation.
    
    Valid account types: asset, income, expense, liability, equity
    """
    # Validate account type
    account_type = account_data.get('account_type', '').lower()
    if not validate_account_type(account_type):
        raise HTTPException(
            status_code=400,
            detail=f"Invalid account_type '{account_type}'. Must be one of: {', '.join(VALID_ACCOUNT_TYPES)}"
        )
    
    # Ensure account_type is lowercase
    account_data['account_type'] = account_type
    
    try:
        account = Account(**account_data, created_by=current_user.id)
        await db.accounts.insert_one(account.model_dump())
        await create_audit_log(current_user.id, current_user.full_name, "account", account.id, "create")
        return account
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

@api_router.patch("/accounts/{account_id}")
async def update_account(account_id: str, update_data: dict, current_user: User = Depends(require_permission('finance.create'))):
    """
    Update an account. If account_type is being changed, validates it's a valid type.
    """
    existing = await db.accounts.find_one({"id": account_id, "is_deleted": False})
    if not existing:
        raise HTTPException(status_code=404, detail="Account not found")
    
    # Validate account_type if it's being updated
    if 'account_type' in update_data:
        account_type = update_data['account_type'].lower()
        if not validate_account_type(account_type):
            raise HTTPException(
                status_code=400,
                detail=f"Invalid account_type '{account_type}'. Must be one of: {', '.join(VALID_ACCOUNT_TYPES)}"
            )
        update_data['account_type'] = account_type
    
    await db.accounts.update_one({"id": account_id}, {"$set": update_data})
    await create_audit_log(current_user.id, current_user.full_name, "account", account_id, "update", update_data)
    return {"message": "Account updated successfully"}

@api_router.delete("/accounts/{account_id}")
@limiter.limit("30/minute")  # Sensitive finance operation: 30 deletions per minute
async def delete_account(request: Request, account_id: str, current_user: User = Depends(require_permission('finance.delete'))):
    """
    Delete an account. Protected accounts (Cash, Bank, Sales Income, etc.) cannot be deleted.
    """
    existing = await db.accounts.find_one({"id": account_id, "is_deleted": False})
    if not existing:
        raise HTTPException(status_code=404, detail="Account not found")
    
    # Check if this is a protected account
    account_name = existing.get('name', '')
    if account_name in PROTECTED_ACCOUNTS:
        raise HTTPException(
            status_code=400,
            detail=f"Cannot delete protected account '{account_name}'. This account is required for system operations."
        )
    
    # Check if account has transactions
    transactions = await db.transactions.find_one({"account_id": account_id, "is_deleted": False})
    if transactions:
        raise HTTPException(status_code=400, detail="Cannot delete account with existing transactions")
    
    await db.accounts.update_one(
        {"id": account_id},
        {"$set": {"is_deleted": True}}
    )
    await create_audit_log(current_user.id, current_user.full_name, "account", account_id, "delete")
    return {"message": "Account deleted successfully"}

@api_router.get("/transactions")
async def get_transactions(
    page: int = 1,
    page_size: int = 10,
    account_id: Optional[str] = None,
    account_type: Optional[str] = None,  # "cash" or "bank"
    transaction_type: Optional[str] = None,  # "credit" or "debit"
    reference_type: Optional[str] = None,  # "invoice", "purchase", "manual"
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: User = Depends(require_permission('finance.view'))
):
    """
    Get transactions with pagination and filtering support.
    Includes running balance calculation for each transaction.
    """
    query = {"is_deleted": False}
    
    # Apply filters
    if account_id:
        query["account_id"] = account_id
    
    if transaction_type:
        query["transaction_type"] = transaction_type
    
    if reference_type:
        if reference_type == "manual":
            query["reference_type"] = None
        else:
            query["reference_type"] = reference_type
    
    # Date range filter
    if start_date or end_date:
        date_query = {}
        if start_date:
            date_query["$gte"] = datetime.fromisoformat(start_date.replace('Z', '+00:00'))
        if end_date:
            date_query["$lte"] = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
        if date_query:
            query["date"] = date_query
    
    # Account type filter (cash vs bank)
    if account_type:
        accounts = await db.accounts.find({"account_type": account_type, "is_deleted": False}, {"_id": 0}).to_list(1000)
        account_ids = [acc['id'] for acc in accounts]
        if account_ids:
            query["account_id"] = {"$in": account_ids}
        else:
            # No accounts of this type exist, return empty
            return create_pagination_response([], 0, page, page_size)
    
    # Calculate skip value
    skip = (page - 1) * page_size
    
    # Get total count for pagination
    total_count = await db.transactions.count_documents(query)
    
    # Get paginated results sorted by date (newest first)
    transactions = await db.transactions.find(query, {"_id": 0}).sort("date", -1).skip(skip).limit(page_size).to_list(page_size)
    
    # Enhance each transaction with account type and running balance
    account_cache = {}
    for txn in transactions:
        # Get account details
        if txn['account_id'] not in account_cache:
            account = await db.accounts.find_one({"id": txn['account_id']}, {"_id": 0})
            if account:
                account_cache[txn['account_id']] = account
        
        if txn['account_id'] in account_cache:
            account = account_cache[txn['account_id']]
            txn['account_type'] = account['account_type']
            txn['account_current_balance'] = account['current_balance']
        else:
            txn['account_type'] = 'unknown'
            txn['account_current_balance'] = 0.0
        
        # Set transaction source
        if txn.get('reference_type') == 'invoice':
            txn['transaction_source'] = 'Invoice Payment'
        elif txn.get('reference_type') == 'purchase':
            txn['transaction_source'] = 'Purchase Payment'
        elif txn.get('reference_type') == 'jobcard':
            txn['transaction_source'] = 'Job Card'
        else:
            txn['transaction_source'] = 'Manual Entry'
    
    # Calculate running balance for display
    # For simplicity, we'll calculate the balance at the time of transaction
    # by getting all transactions for that account up to that date
    for txn in transactions:
        # Get all transactions for this account up to and including this transaction date
        prior_txns = await db.transactions.find({
            "account_id": txn['account_id'],
            "is_deleted": False,
            "date": {"$lte": txn['date']}
        }, {"_id": 0}).sort("date", 1).to_list(10000)
        
        # Get the account's opening balance
        account = account_cache.get(txn['account_id'])
        if account:
            running_balance = account['opening_balance']
        else:
            running_balance = 0.0
        
        # Calculate running balance up to current transaction
        balance_before = running_balance
        for pt in prior_txns:
            if pt['transaction_type'] == 'credit':
                running_balance += pt['amount']
            else:
                running_balance -= pt['amount']
            
            # If this is the current transaction, capture before and after
            if pt['id'] == txn['id']:
                if txn['transaction_type'] == 'credit':
                    balance_before = running_balance - txn['amount']
                else:
                    balance_before = running_balance + txn['amount']
                break
        
        txn['balance_before'] = round(balance_before, 3)
        txn['balance_after'] = round(running_balance, 3)
    
    return create_pagination_response(transactions, total_count, page, page_size)

@api_router.post("/transactions", response_model=Transaction)
async def create_transaction(transaction_data: dict, current_user: User = Depends(require_permission('finance.create'))):
    """
    Create a manual finance transaction.
    
    WARNING: Manual transactions should be created in pairs for double-entry bookkeeping.
    Consider creating both debit and credit sides to maintain balance.
    
    Balance update rules:
    - ASSET/EXPENSE accounts: Debit increases, Credit decreases
    - INCOME/LIABILITY/EQUITY accounts: Credit increases, Debit decreases
    """
    year = datetime.now(timezone.utc).year
    count = await db.transactions.count_documents({"transaction_number": {"$regex": f"^TXN-{year}"}})
    transaction_number = f"TXN-{year}-{str(count + 1).zfill(4)}"
    
    account = await db.accounts.find_one({"id": transaction_data['account_id']}, {"_id": 0})
    if not account:
        raise HTTPException(status_code=404, detail="Account not found")
    
    # Validate transaction type
    transaction_type = transaction_data.get('transaction_type', '').lower()
    if transaction_type not in ['debit', 'credit']:
        raise HTTPException(
            status_code=400,
            detail="transaction_type must be either 'debit' or 'credit'"
        )
    
    # Remove conflicting keys and add required fields
    transaction_data_clean = {k: v for k, v in transaction_data.items() if k not in ['transaction_number', 'account_name', 'created_by']}
    transaction = Transaction(
        **transaction_data_clean,
        transaction_number=transaction_number,
        account_name=account['name'],
        created_by=current_user.id
    )
    
    await db.transactions.insert_one(transaction.model_dump())
    
    # Calculate balance delta using account-type-aware logic
    account_type = account.get('account_type', 'asset')
    amount = transaction.amount
    delta = calculate_balance_delta(account_type, transaction_type, amount)
    
    await db.accounts.update_one(
        {"id": transaction.account_id},
        {"$inc": {"current_balance": delta}}
    )
    
    await create_audit_log(
        current_user.id,
        current_user.full_name,
        "transaction",
        transaction.id,
        "create",
        {
            "account_type": account_type,
            "transaction_type": transaction_type,
            "amount": amount,
            "balance_delta": delta
        }
    )
    
    return transaction


@api_router.get("/transactions/summary")
async def get_transactions_summary(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    account_id: Optional[str] = None,
    current_user: User = Depends(require_permission('finance.view'))
):
    """
    Get transaction summary including:
    - Total credit (money IN)
    - Total debit (money OUT)
    - Net flow
    - Account-wise breakdown
    - Cash vs Bank breakdown
    """
    try:
        query = {"is_deleted": False}
        
        # Date range filter
        if start_date or end_date:
            date_query = {}
            if start_date:
                try:
                    date_query["$gte"] = datetime.fromisoformat(start_date.replace('Z', '+00:00'))
                except ValueError:
                    raise HTTPException(status_code=400, detail="Invalid start_date format")
            if end_date:
                try:
                    date_query["$lte"] = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
                except ValueError:
                    raise HTTPException(status_code=400, detail="Invalid end_date format")
            if date_query:
                query["date"] = date_query
        
        # Account filter
        if account_id:
            query["account_id"] = account_id
        
        # Get all transactions matching the query
        transactions = await db.transactions.find(query, {"_id": 0}).to_list(10000)
        
        # Calculate totals
        total_credit = 0.0
        total_debit = 0.0
        account_breakdown = {}
        
        for txn in transactions:
            try:
                txn_type = txn.get('transaction_type', 'debit')
                amount = float(txn.get('amount', 0))
                
                if txn_type == 'credit':
                    total_credit += amount
                else:
                    total_debit += amount
                
                # Account-wise breakdown
                acc_id = txn.get('account_id')
                if acc_id:
                    if acc_id not in account_breakdown:
                        account_breakdown[acc_id] = {
                            'account_id': acc_id,
                            'account_name': txn.get('account_name', 'Unknown'),
                            'credit': 0.0,
                            'debit': 0.0
                        }
                    
                    if txn_type == 'credit':
                        account_breakdown[acc_id]['credit'] += amount
                    else:
                        account_breakdown[acc_id]['debit'] += amount
            except (KeyError, TypeError, ValueError) as e:
                # Skip malformed transaction
                continue
        
        # Get accounts to determine cash vs bank
        try:
            accounts = await db.accounts.find({"is_deleted": False}, {"_id": 0}).to_list(1000)
            account_types = {acc['id']: acc.get('account_type', 'unknown') for acc in accounts if 'id' in acc}
        except Exception:
            account_types = {}
        
        # Add account type to breakdown and calculate net
        for acc_id, breakdown in account_breakdown.items():
            breakdown['account_type'] = account_types.get(acc_id, 'unknown')
            breakdown['net'] = round(breakdown['credit'] - breakdown['debit'], 3)
            breakdown['credit'] = round(breakdown['credit'], 3)
            breakdown['debit'] = round(breakdown['debit'], 3)
        
        # Cash vs Bank breakdown
        cash_credit = 0.0
        cash_debit = 0.0
        bank_credit = 0.0
        bank_debit = 0.0
        
        for breakdown in account_breakdown.values():
            acc_type = breakdown.get('account_type', 'unknown')
            if acc_type in ['cash', 'petty']:
                cash_credit += breakdown['credit']
                cash_debit += breakdown['debit']
            elif acc_type == 'bank':
                bank_credit += breakdown['credit']
                bank_debit += breakdown['debit']
        
        # FIX: Net Flow should represent actual cash flow (Cash + Bank movements only)
        # This gives meaningful business insight: "How much money moved in/out?"
        cash_net = cash_credit - cash_debit
        bank_net = bank_credit - bank_debit
        net_flow = cash_net + bank_net
        
        return {
            "total_credit": round(total_credit, 3),
            "total_debit": round(total_debit, 3),
            "net_flow": round(net_flow, 3),
            "transaction_count": len(transactions),
            "cash_summary": {
                "credit": round(cash_credit, 3),
                "debit": round(cash_debit, 3),
                "net": round(cash_credit - cash_debit, 3)
            },
            "bank_summary": {
                "credit": round(bank_credit, 3),
                "debit": round(bank_debit, 3),
                "net": round(bank_credit - bank_debit, 3)
            },
            "account_breakdown": list(account_breakdown.values())
        }
    except HTTPException:
        raise
    except Exception as e:
        # Log the error and return a valid empty response
        print(f"Error in get_transactions_summary: {str(e)}")
        return {
            "total_credit": 0.0,
            "total_debit": 0.0,
            "net_flow": 0.0,
            "transaction_count": 0,
            "cash_summary": {
                "credit": 0.0,
                "debit": 0.0,
                "net": 0.0
            },
            "bank_summary": {
                "credit": 0.0,
                "debit": 0.0,
                "net": 0.0
            },
            "account_breakdown": []
        }


@api_router.get("/daily-closings", response_model=List[DailyClosing])
async def get_daily_closings(current_user: User = Depends(require_permission('finance.view'))):
    try:
        closings = await db.daily_closings.find({}, {"_id": 0}).sort("date", -1).to_list(1000)
        return closings if closings else []
    except Exception as e:
        print(f"Error in get_daily_closings: {str(e)}")
        # Return empty list instead of crashing
        return []

@api_router.post("/daily-closings", response_model=DailyClosing)
async def create_daily_closing(closing_data: dict, current_user: User = Depends(require_permission('finance.create'))):
    """
    Create a new daily closing record.
    If opening_cash, total_credit, total_debit, expected_closing, or difference are not provided,
    they will be auto-calculated based on the date and transactions.
    """
    try:
        # Parse the date from closing_data
        closing_date = closing_data.get('date')
        if isinstance(closing_date, str):
            target_date = datetime.fromisoformat(closing_date.replace('Z', '+00:00'))
        elif isinstance(closing_date, datetime):
            target_date = closing_date
        else:
            raise ValueError("Invalid date format")
        
        # Check if we need to auto-calculate fields
        needs_calculation = (
            'opening_cash' not in closing_data or
            'total_credit' not in closing_data or
            'total_debit' not in closing_data or
            'expected_closing' not in closing_data
        )
        
        if needs_calculation:
            # Auto-calculate missing fields
            start_of_day = target_date.replace(hour=0, minute=0, second=0, microsecond=0, tzinfo=timezone.utc)
            end_of_day = target_date.replace(hour=23, minute=59, second=59, microsecond=999999, tzinfo=timezone.utc)
            
            # Get previous day's closing for opening_cash
            previous_date = target_date - timedelta(days=1)
            previous_closing = await db.daily_closings.find_one(
                {"date": {"$gte": previous_date.replace(hour=0, minute=0, second=0, tzinfo=timezone.utc),
                          "$lte": previous_date.replace(hour=23, minute=59, second=59, tzinfo=timezone.utc)}},
                {"_id": 0}
            )
            opening_cash = round(previous_closing['actual_closing'], 3) if previous_closing else 0.0
            
            # Get all transactions for the target date
            transactions = await db.transactions.find(
                {
                    "date": {"$gte": start_of_day, "$lte": end_of_day},
                    "is_deleted": False
                },
                {"_id": 0}
            ).to_list(None)
            
            # Calculate total credit and debit
            total_credit = round(sum(t['amount'] for t in transactions if t['transaction_type'] == 'credit'), 3)
            total_debit = round(sum(t['amount'] for t in transactions if t['transaction_type'] == 'debit'), 3)
            expected_closing = round(opening_cash + total_credit - total_debit, 3)
            
            # Update closing_data with calculated values
            closing_data['opening_cash'] = opening_cash
            closing_data['total_credit'] = total_credit
            closing_data['total_debit'] = total_debit
            closing_data['expected_closing'] = expected_closing
        
        # Calculate difference: actual_closing - expected_closing
        actual_closing = closing_data.get('actual_closing', 0.0)
        expected_closing = closing_data.get('expected_closing', 0.0)
        closing_data['difference'] = round(actual_closing - expected_closing, 3)
        
        # Create the closing record
        closing = DailyClosing(**closing_data, closed_by=current_user.id)
        await db.daily_closings.insert_one(closing.model_dump())
        await create_audit_log(current_user.id, current_user.full_name, "daily_closing", closing.id, "create")
        return closing
        
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to create daily closing: {str(e)}")

@api_router.get("/daily-closings/calculate/{date}")
async def calculate_daily_closing(date: str, current_user: User = Depends(require_permission('finance.view'))):
    """
    Auto-calculate daily closing values for a specific date.
    Returns:
    - opening_cash: actual_closing from previous day's closing (or 0 if first day)
    - total_credit: sum of all credit transactions for the date
    - total_debit: sum of all debit transactions for the date
    - expected_closing: opening_cash + total_credit - total_debit
    """
    try:
        # Parse the date string (format: YYYY-MM-DD)
        target_date = datetime.strptime(date, "%Y-%m-%d")
        start_of_day = target_date.replace(hour=0, minute=0, second=0, microsecond=0, tzinfo=timezone.utc)
        end_of_day = target_date.replace(hour=23, minute=59, second=59, microsecond=999999, tzinfo=timezone.utc)
        
        # Get previous day's closing for opening_cash
        previous_date = target_date - timedelta(days=1)
        previous_closing = await db.daily_closings.find_one(
            {"date": {"$gte": previous_date.replace(hour=0, minute=0, second=0, tzinfo=timezone.utc),
                      "$lte": previous_date.replace(hour=23, minute=59, second=59, tzinfo=timezone.utc)}},
            {"_id": 0}
        )
        opening_cash = previous_closing['actual_closing'] if previous_closing else 0.0
        
        # Get all transactions for the target date
        transactions = await db.transactions.find(
            {
                "date": {"$gte": start_of_day, "$lte": end_of_day},
                "is_deleted": False
            },
            {"_id": 0}
        ).to_list(None)
        
        # Calculate total credit and debit
        total_credit = sum(t['amount'] for t in transactions if t['transaction_type'] == 'credit')
        total_debit = sum(t['amount'] for t in transactions if t['transaction_type'] == 'debit')
        
        # Round to 3 decimal places (OMR standard)
        opening_cash = round(opening_cash, 3)
        total_credit = round(total_credit, 3)
        total_debit = round(total_debit, 3)
        expected_closing = round(opening_cash + total_credit - total_debit, 3)
        
        return {
            "date": date,
            "opening_cash": opening_cash,
            "total_credit": total_credit,
            "total_debit": total_debit,
            "expected_closing": expected_closing,
            "transaction_count": len(transactions),
            "credit_count": sum(1 for t in transactions if t['transaction_type'] == 'credit'),
            "debit_count": sum(1 for t in transactions if t['transaction_type'] == 'debit'),
            "has_previous_closing": previous_closing is not None
        }
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to calculate daily closing: {str(e)}")


@api_router.patch("/daily-closings/{closing_id}", response_model=DailyClosing)
async def update_daily_closing(closing_id: str, update_data: dict, current_user: User = Depends(require_permission('finance.create'))):
    """
    Update an existing daily closing record.
    Can update actual_closing, notes, or lock status.
    Automatically recalculates difference when actual_closing is updated.
    """
    try:
        # Find the existing closing record
        existing_closing = await db.daily_closings.find_one({"id": closing_id}, {"_id": 0})
        if not existing_closing:
            raise HTTPException(status_code=404, detail="Daily closing record not found")
        
        # Check if record is locked
        if existing_closing.get('is_locked', False) and not update_data.get('is_locked') == False:
            raise HTTPException(status_code=403, detail="Cannot update a locked daily closing record")
        
        # If actual_closing is being updated, recalculate difference
        if 'actual_closing' in update_data:
            expected_closing = existing_closing.get('expected_closing', 0.0)
            actual_closing = update_data['actual_closing']
            update_data['difference'] = round(actual_closing - expected_closing, 3)
        
        # Update the record
        result = await db.daily_closings.update_one(
            {"id": closing_id},
            {"$set": update_data}
        )
        
        if result.modified_count == 0 and result.matched_count == 0:
            raise HTTPException(status_code=404, detail="Daily closing record not found")
        
        # Get the updated record
        updated_closing = await db.daily_closings.find_one({"id": closing_id}, {"_id": 0})
        
        # Create audit log
        await create_audit_log(current_user.id, current_user.full_name, "daily_closing", closing_id, "update")
        
        return DailyClosing(**updated_closing)
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to update daily closing: {str(e)}")

@api_router.get("/audit-logs")
async def get_audit_logs(
    module: Optional[str] = None,
    action: Optional[str] = None,
    user_id: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    page: int = 1,
    page_size: int = 10,
    current_user: User = Depends(require_permission('audit.view'))
):
    """
    Get audit logs with comprehensive filtering options and pagination.
    
    Filters:
    - module: Filter by module name (e.g., 'invoice', 'jobcard', 'party')
    - action: Filter by action type (e.g., 'create', 'update', 'delete')
    - user_id: Filter by user ID who performed the action
    - date_from: Filter logs from this date (ISO format: YYYY-MM-DD)
    - date_to: Filter logs up to this date (ISO format: YYYY-MM-DD)
    - page: Page number (default: 1)
    - page_size: Items per page (default: 10)
    """
    query = {}
    
    # Module filter
    if module:
        query['module'] = module
    
    # Action filter
    if action:
        query['action'] = action
    
    # User filter
    if user_id:
        query['user_id'] = user_id
    
    # Date range filter
    if date_from or date_to:
        date_query = {}
        if date_from:
            try:
                from_date = datetime.fromisoformat(date_from.replace('Z', '+00:00'))
                date_query['$gte'] = from_date
            except ValueError:
                pass  # Skip invalid date format
        if date_to:
            try:
                to_date = datetime.fromisoformat(date_to.replace('Z', '+00:00'))
                # Add time to include the entire end date
                to_date = to_date.replace(hour=23, minute=59, second=59)
                date_query['$lte'] = to_date
            except ValueError:
                pass  # Skip invalid date format
        if date_query:
            query['timestamp'] = date_query
    
    # Calculate skip value
    skip = (page - 1) * page_size
    
    # Get total count for pagination
    total_count = await db.audit_logs.count_documents(query)
    
    # Get paginated results
    logs = await db.audit_logs.find(query, {"_id": 0}).sort("timestamp", -1).skip(skip).limit(page_size).to_list(page_size)
    
    return create_pagination_response(logs, total_count, page, page_size)

@api_router.get("/reports/inventory-export")
async def export_inventory(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    movement_type: Optional[str] = None,
    category: Optional[str] = None,
    current_user: User = Depends(require_permission('reports.view'))
):
    from fastapi.responses import StreamingResponse
    from io import BytesIO
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    
    # Build query with filters
    query = {"is_deleted": False}
    if start_date:
        query['date'] = {"$gte": datetime.fromisoformat(start_date)}
    if end_date:
        end_dt = datetime.fromisoformat(end_date)
        if 'date' in query:
            query['date']['$lte'] = end_dt
        else:
            query['date'] = {"$lte": end_dt}
    if movement_type:
        query['movement_type'] = movement_type
    if category:
        query['header_name'] = category
    
    # Get filtered inventory data
    movements = await db.stock_movements.find(query, {"_id": 0}).sort("date", -1).to_list(10000)
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventory Movements"
    
    # Headers
    headers = ["Date", "Type", "Category", "Description", "Quantity", "Weight (g)", "Purity", "Notes"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    
    # Data
    for row_idx, movement in enumerate(movements, 2):
        ws.cell(row=row_idx, column=1, value=str(movement.get('date', ''))[:10])
        ws.cell(row=row_idx, column=2, value=movement.get('movement_type', ''))
        ws.cell(row=row_idx, column=3, value=movement.get('header_name', ''))
        ws.cell(row=row_idx, column=4, value=movement.get('description', ''))
        ws.cell(row=row_idx, column=5, value=movement.get('qty_delta', 0))
        ws.cell(row=row_idx, column=6, value=movement.get('weight_delta', 0))
        ws.cell(row=row_idx, column=7, value=movement.get('purity', 0))
        ws.cell(row=row_idx, column=8, value=movement.get('notes', ''))
    
    # Adjust column widths
    for col in range(1, 9):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=inventory_export.xlsx"}
    )

@api_router.get("/reports/parties-export")
async def export_parties(
    party_type: Optional[str] = None,
    current_user: User = Depends(require_permission('reports.view'))
):
    from fastapi.responses import StreamingResponse
    from io import BytesIO
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    
    # Build query with filters
    query = {"is_deleted": False}
    if party_type:
        query['party_type'] = party_type
    
    parties = await db.parties.find(query, {"_id": 0}).to_list(10000)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Parties"
    
    headers = ["Name", "Phone", "Type", "Address", "Notes", "Created At"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    for row_idx, party in enumerate(parties, 2):
        ws.cell(row=row_idx, column=1, value=party.get('name', ''))
        ws.cell(row=row_idx, column=2, value=party.get('phone', ''))
        ws.cell(row=row_idx, column=3, value=party.get('party_type', ''))
        ws.cell(row=row_idx, column=4, value=party.get('address', ''))
        ws.cell(row=row_idx, column=5, value=party.get('notes', ''))
        ws.cell(row=row_idx, column=6, value=str(party.get('created_at', ''))[:10])
    
    for col in range(1, 7):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=parties_export.xlsx"}
    )

@api_router.get("/reports/invoices-export")
async def export_invoices(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    invoice_type: Optional[str] = None,
    payment_status: Optional[str] = None,
    current_user: User = Depends(require_permission('reports.view'))
):
    """
    Enhanced Excel Export with Multiple Sheets:
    - Sheet 1: Invoice Summary
    - Sheet 2: Line Items (detailed breakdown)
    - Sheet 3: Totals & Statistics
    
    All numeric columns are proper numbers (not strings) for Excel calculations
    """
    from fastapi.responses import StreamingResponse
    from io import BytesIO
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    # Build query with filters
    query = {"is_deleted": False}
    if start_date:
        query['date'] = {"$gte": datetime.fromisoformat(start_date)}
    if end_date:
        end_dt = datetime.fromisoformat(end_date)
        if 'date' in query:
            query['date']['$lte'] = end_dt
        else:
            query['date'] = {"$lte": end_dt}
    if invoice_type:
        query['invoice_type'] = invoice_type
    if payment_status:
        query['payment_status'] = payment_status
    
    invoices = await db.invoices.find(query, {"_id": 0}).sort("date", -1).to_list(10000)
    
    wb = openpyxl.Workbook()
    
    # ===========================================================================
    # SHEET 1: Invoice Summary
    # ===========================================================================
    ws1 = wb.active
    ws1.title = "Invoice Summary"
    
    # Headers with styling
    summary_headers = [
        "Invoice #", "Date", "Customer", "Customer Type", "Type", 
        "Status", "Grand Total", "Paid Amount", "Balance Due", "Payment Status"
    ]
    
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col, header in enumerate(summary_headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Data rows with proper data types
    for row_idx, inv in enumerate(invoices, 2):
        ws1.cell(row=row_idx, column=1, value=inv.get('invoice_number', ''))
        ws1.cell(row=row_idx, column=2, value=str(inv.get('date', ''))[:10])
        
        # Customer name (handle walk-in)
        customer_name = inv.get('walk_in_name') or inv.get('customer_name', 'N/A')
        ws1.cell(row=row_idx, column=3, value=customer_name)
        ws1.cell(row=row_idx, column=4, value=inv.get('customer_type', 'walk_in'))
        ws1.cell(row=row_idx, column=5, value=inv.get('invoice_type', ''))
        ws1.cell(row=row_idx, column=6, value=inv.get('status', 'draft'))
        
        # Numeric columns (proper numbers, not strings)
        ws1.cell(row=row_idx, column=7, value=float(inv.get('grand_total', 0)))
        ws1.cell(row=row_idx, column=8, value=float(inv.get('paid_amount', 0)))
        ws1.cell(row=row_idx, column=9, value=float(inv.get('balance_due', 0)))
        ws1.cell(row=row_idx, column=10, value=inv.get('payment_status', ''))
    
    # Set column widths
    ws1.column_dimensions['A'].width = 15
    ws1.column_dimensions['B'].width = 12
    ws1.column_dimensions['C'].width = 25
    ws1.column_dimensions['D'].width = 15
    ws1.column_dimensions['E'].width = 10
    ws1.column_dimensions['F'].width = 12
    ws1.column_dimensions['G'].width = 15
    ws1.column_dimensions['H'].width = 15
    ws1.column_dimensions['I'].width = 15
    ws1.column_dimensions['J'].width = 15
    
    # ===========================================================================
    # SHEET 2: Invoice Line Items (Detailed)
    # ===========================================================================
    ws2 = wb.create_sheet(title="Invoice Line Items")
    
    line_item_headers = [
        "Invoice #", "Date", "Customer", "Item Category", "Description", 
        "Qty", "Purity", "Weight (g)", "Gold Rate", "Gold Value", 
        "Making Charge", "VAT %", "VAT Amount", "Line Total"
    ]
    
    for col, header in enumerate(line_item_headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Populate line items from all invoices
    current_row = 2
    for inv in invoices:
        invoice_number = inv.get('invoice_number', '')
        invoice_date = str(inv.get('date', ''))[:10]
        customer_name = inv.get('walk_in_name') or inv.get('customer_name', 'N/A')
        
        for item in inv.get('items', []):
            ws2.cell(row=current_row, column=1, value=invoice_number)
            ws2.cell(row=current_row, column=2, value=invoice_date)
            ws2.cell(row=current_row, column=3, value=customer_name)
            ws2.cell(row=current_row, column=4, value=item.get('category', ''))
            ws2.cell(row=current_row, column=5, value=item.get('description', ''))
            
            # Numeric values
            ws2.cell(row=current_row, column=6, value=int(item.get('qty', 1)))
            ws2.cell(row=current_row, column=7, value=int(item.get('purity', 916)))
            ws2.cell(row=current_row, column=8, value=float(item.get('weight', 0)))
            ws2.cell(row=current_row, column=9, value=float(item.get('metal_rate', 0)))
            ws2.cell(row=current_row, column=10, value=float(item.get('gold_value', 0)))
            ws2.cell(row=current_row, column=11, value=float(item.get('making_value', 0)))
            ws2.cell(row=current_row, column=12, value=float(item.get('vat_percent', 5)))
            ws2.cell(row=current_row, column=13, value=float(item.get('vat_amount', 0)))
            ws2.cell(row=current_row, column=14, value=float(item.get('line_total', 0)))
            
            current_row += 1
    
    # Set column widths for line items
    for col in ['A', 'B', 'C', 'D', 'E']:
        ws2.column_dimensions[col].width = 15
    for col in ['F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']:
        ws2.column_dimensions[col].width = 12
    
    # ===========================================================================
    # SHEET 3: Totals & Statistics
    # ===========================================================================
    ws3 = wb.create_sheet(title="Totals")
    
    # Calculate totals
    total_invoices = len(invoices)
    metal_total = sum(
        item.get('gold_value', 0) 
        for inv in invoices 
        for item in inv.get('items', [])
    )
    making_total = sum(
        item.get('making_value', 0) 
        for inv in invoices 
        for item in inv.get('items', [])
    )
    vat_total = sum(inv.get('vat_total', 0) for inv in invoices)
    grand_total = sum(inv.get('grand_total', 0) for inv in invoices)
    paid_total = sum(inv.get('paid_amount', 0) for inv in invoices)
    outstanding_total = sum(inv.get('balance_due', 0) for inv in invoices)
    
    # Style for totals sheet
    title_font = Font(bold=True, size=14)
    label_font = Font(bold=True)
    
    ws3['A1'] = 'INVOICE TOTALS SUMMARY'
    ws3['A1'].font = title_font
    
    row = 3
    totals_data = [
        ('Total Invoices', total_invoices),
        ('', ''),
        ('Metal Total (OMR)', metal_total),
        ('Making Charges Total (OMR)', making_total),
        ('VAT Total (OMR)', vat_total),
        ('', ''),
        ('Grand Total (OMR)', grand_total),
        ('Total Paid (OMR)', paid_total),
        ('Total Outstanding (OMR)', outstanding_total),
    ]
    
    for label, value in totals_data:
        ws3.cell(row=row, column=1, value=label).font = label_font
        if value != '':
            # Make sure numeric values are stored as numbers
            if isinstance(value, (int, float)):
                ws3.cell(row=row, column=2, value=float(value))
        row += 1
    
    ws3.column_dimensions['A'].width = 30
    ws3.column_dimensions['B'].width = 20
    
    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=invoices_export.xlsx"}
    )

@api_router.get("/reports/transactions-export")
async def export_transactions(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    transaction_type: Optional[str] = None,
    party_id: Optional[str] = None,
    current_user: User = Depends(require_permission('reports.view'))
):
    """Export transactions report as Excel"""
    from fastapi.responses import StreamingResponse
    from io import BytesIO
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    
    # Get filtered transaction data
    data = await view_transactions_report(
        start_date=start_date,
        end_date=end_date,
        transaction_type=transaction_type,
        party_id=party_id,
        sort_by='date_desc',
        current_user=current_user
    )
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transactions"
    
    # Headers
    headers = ["Date", "Transaction #", "Type", "Mode", "Party Name", "Account", "Amount (OMR)", "Category", "Notes"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    
    # Data
    for row_idx, txn in enumerate(data['transactions'], 2):
        txn_date = txn.get('date', '')
        if isinstance(txn_date, str):
            txn_date = txn_date[:10]
        elif hasattr(txn_date, 'strftime'):
            txn_date = txn_date.strftime('%Y-%m-%d')
        
        ws.cell(row=row_idx, column=1, value=txn_date)
        ws.cell(row=row_idx, column=2, value=txn.get('transaction_number', ''))
        ws.cell(row=row_idx, column=3, value=txn.get('transaction_type', ''))
        ws.cell(row=row_idx, column=4, value=txn.get('mode', ''))
        ws.cell(row=row_idx, column=5, value=txn.get('party_name', ''))
        ws.cell(row=row_idx, column=6, value=txn.get('account_name', ''))
        ws.cell(row=row_idx, column=7, value=txn.get('amount', 0))
        ws.cell(row=row_idx, column=8, value=txn.get('category', ''))
        ws.cell(row=row_idx, column=9, value=txn.get('notes', ''))
    
    # Add summary at the bottom
    last_row = len(data['transactions']) + 3
    ws.cell(row=last_row, column=1, value="Summary:").font = Font(bold=True)
    ws.cell(row=last_row + 1, column=1, value="Total Credit:")
    ws.cell(row=last_row + 1, column=2, value=data['summary']['total_credit'])
    ws.cell(row=last_row + 2, column=1, value="Total Debit:")
    ws.cell(row=last_row + 2, column=2, value=data['summary']['total_debit'])
    ws.cell(row=last_row + 3, column=1, value="Net Balance:")
    ws.cell(row=last_row + 3, column=2, value=data['summary']['net_balance'])
    
    # Adjust column widths
    for col in range(1, 10):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=transactions_export_{datetime.now().strftime('%Y%m%d')}.xlsx"}
    )

@api_router.get("/reports/outstanding-export")
async def export_outstanding(
    party_id: Optional[str] = None,
    party_type: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: User = Depends(require_permission('reports.view'))
):
    """Export outstanding report as Excel"""
    from fastapi.responses import StreamingResponse
    from io import BytesIO
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    
    # Get filtered outstanding data
    data = await get_outstanding_report(
        party_id=party_id,
        party_type=party_type,
        start_date=start_date,
        end_date=end_date,
        include_paid=False,
        current_user=current_user
    )
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Outstanding"
    
    # Headers
    headers = [
        "Party Name", "Type", "Total Invoiced", "Total Paid", "Outstanding", 
        "Overdue 0-7d", "Overdue 8-30d", "Overdue 31+d", "Last Invoice Date", "Last Payment Date"
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    
    # Data
    for row_idx, party in enumerate(data['parties'], 2):
        last_invoice = party.get('last_invoice_date')
        if last_invoice and hasattr(last_invoice, 'strftime'):
            last_invoice = last_invoice.strftime('%Y-%m-%d')
        elif isinstance(last_invoice, str):
            last_invoice = last_invoice[:10]
        else:
            last_invoice = ''
        
        last_payment = party.get('last_payment_date')
        if last_payment and hasattr(last_payment, 'strftime'):
            last_payment = last_payment.strftime('%Y-%m-%d')
        elif isinstance(last_payment, str):
            last_payment = last_payment[:10]
        else:
            last_payment = ''
        
        ws.cell(row=row_idx, column=1, value=party.get('party_name', ''))
        ws.cell(row=row_idx, column=2, value=party.get('party_type', ''))
        ws.cell(row=row_idx, column=3, value=party.get('total_invoiced', 0))
        ws.cell(row=row_idx, column=4, value=party.get('total_paid', 0))
        ws.cell(row=row_idx, column=5, value=party.get('total_outstanding', 0))
        ws.cell(row=row_idx, column=6, value=party.get('overdue_0_7', 0))
        ws.cell(row=row_idx, column=7, value=party.get('overdue_8_30', 0))
        ws.cell(row=row_idx, column=8, value=party.get('overdue_31_plus', 0))
        ws.cell(row=row_idx, column=9, value=last_invoice)
        ws.cell(row=row_idx, column=10, value=last_payment)
    
    # Add summary at the bottom
    last_row = len(data['parties']) + 3
    ws.cell(row=last_row, column=1, value="Summary:").font = Font(bold=True)
    ws.cell(row=last_row + 1, column=1, value="Customer Due (Receivable):")
    ws.cell(row=last_row + 1, column=2, value=data['summary']['customer_due'])
    ws.cell(row=last_row + 2, column=1, value="Vendor Payable:")
    ws.cell(row=last_row + 2, column=2, value=data['summary']['vendor_payable'])
    ws.cell(row=last_row + 3, column=1, value="Total Outstanding:")
    ws.cell(row=last_row + 3, column=2, value=data['summary']['total_outstanding'])
    ws.cell(row=last_row + 4, column=1, value="Overdue 0-7 Days:")
    ws.cell(row=last_row + 4, column=2, value=data['summary']['total_overdue_0_7'])
    ws.cell(row=last_row + 5, column=1, value="Overdue 8-30 Days:")
    ws.cell(row=last_row + 5, column=2, value=data['summary']['total_overdue_8_30'])
    ws.cell(row=last_row + 6, column=1, value="Overdue 31+ Days:")
    ws.cell(row=last_row + 6, column=2, value=data['summary']['total_overdue_31_plus'])
    
    # Adjust column widths
    for col in range(1, 11):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=outstanding_export_{datetime.now().strftime('%Y%m%d')}.xlsx"}
    )

# New VIEW endpoints for displaying reports in UI
@api_router.get("/reports/inventory-view")
async def view_inventory_report(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    movement_type: Optional[str] = None,
    category: Optional[str] = None,
    sort_by: Optional[str] = None,  # NEW: "date_asc", "date_desc"
    current_user: User = Depends(require_permission('reports.view'))
):
    """View inventory movements with filters - returns JSON for UI"""
    query = {"is_deleted": False}
    if start_date:
        query['date'] = {"$gte": datetime.fromisoformat(start_date)}
    if end_date:
        end_dt = datetime.fromisoformat(end_date)
        if 'date' in query:
            query['date']['$lte'] = end_dt
        else:
            query['date'] = {"$lte": end_dt}
    if movement_type:
        query['movement_type'] = movement_type
    if category:
        query['header_name'] = category
    
    # Apply sorting
    sort_field = "date"
    sort_direction = -1  # Default: newest first
    
    if sort_by == "date_asc":
        sort_direction = 1
    elif sort_by == "date_desc":
        sort_direction = -1
    
    movements = await db.stock_movements.find(query, {"_id": 0}).sort(sort_field, sort_direction).to_list(10000)
    
    # Calculate totals
    total_in = sum(m.get('qty_delta', 0) for m in movements if m.get('qty_delta', 0) > 0)
    total_out = sum(abs(m.get('qty_delta', 0)) for m in movements if m.get('qty_delta', 0) < 0)
    total_weight_in = sum(m.get('weight_delta', 0) for m in movements if m.get('weight_delta', 0) > 0)
    total_weight_out = sum(abs(m.get('weight_delta', 0)) for m in movements if m.get('weight_delta', 0) < 0)
    
    return {
        "movements": movements,
        "summary": {
            "total_in": total_in,
            "total_out": total_out,
            "total_weight_in": total_weight_in,
            "total_weight_out": total_weight_out,
            "net_quantity": total_in - total_out,
            "net_weight": total_weight_in - total_weight_out
        },
        "count": len(movements)
    }

@api_router.get("/reports/parties-view")
async def view_parties_report(
    party_type: Optional[str] = None,
    sort_by: Optional[str] = None,  # NEW: "outstanding_desc", "name_asc"
    current_user: User = Depends(require_permission('reports.view'))
):
    """View parties with filters - returns JSON for UI"""
    query = {"is_deleted": False}
    if party_type:
        query['party_type'] = party_type
    
    parties = await db.parties.find(query, {"_id": 0}).to_list(10000)
    
    # Get outstanding amounts for each party
    for party in parties:
        invoices = await db.invoices.find(
            {"customer_id": party['id'], "is_deleted": False},
            {"_id": 0, "balance_due": 1}
        ).to_list(1000)
        party['outstanding'] = sum(inv.get('balance_due', 0) for inv in invoices)
    
    # Apply sorting
    if sort_by == "outstanding_desc":
        parties = sorted(parties, key=lambda x: x.get('outstanding', 0), reverse=True)
    elif sort_by == "name_asc":
        parties = sorted(parties, key=lambda x: x.get('name', '').lower())
    
    return {
        "parties": parties,
        "count": len(parties)
    }

@api_router.get("/reports/invoices-view")
async def view_invoices_report(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    invoice_type: Optional[str] = None,
    payment_status: Optional[str] = None,
    party_id: Optional[str] = None,  # NEW: Filter by specific party
    sort_by: Optional[str] = None,  # NEW: "date_asc", "date_desc", "amount_desc", "outstanding_desc"
    current_user: User = Depends(require_permission('reports.view'))
):
    """View invoices with filters - returns JSON for UI"""
    query = {"is_deleted": False}
    if start_date:
        query['date'] = {"$gte": datetime.fromisoformat(start_date)}
    if end_date:
        end_dt = datetime.fromisoformat(end_date)
        if 'date' in query:
            query['date']['$lte'] = end_dt
        else:
            query['date'] = {"$lte": end_dt}
    if invoice_type:
        query['invoice_type'] = invoice_type
    if payment_status:
        query['payment_status'] = payment_status
    if party_id:
        query['customer_id'] = party_id
    
    # Apply sorting
    sort_field = "date"
    sort_direction = -1  # Default: newest first
    
    if sort_by == "date_asc":
        sort_field = "date"
        sort_direction = 1
    elif sort_by == "date_desc":
        sort_field = "date"
        sort_direction = -1
    elif sort_by == "amount_desc":
        sort_field = "grand_total"
        sort_direction = -1
    elif sort_by == "outstanding_desc":
        sort_field = "balance_due"
        sort_direction = -1
    
    invoices = await db.invoices.find(query, {"_id": 0}).sort(sort_field, sort_direction).to_list(10000)
    
    # Calculate totals
    total_amount = sum(inv.get('grand_total', 0) for inv in invoices)
    total_paid = sum(inv.get('paid_amount', 0) for inv in invoices)
    total_balance = sum(inv.get('balance_due', 0) for inv in invoices)
    
    return {
        "invoices": invoices,
        "summary": {
            "total_amount": total_amount,
            "total_paid": total_paid,
            "total_balance": total_balance
        },
        "count": len(invoices)
    }

@api_router.get("/reports/transactions-view")
async def view_transactions_report(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    transaction_type: Optional[str] = None,
    account_id: Optional[str] = None,
    party_id: Optional[str] = None,  # NEW: Filter by specific party
    sort_by: Optional[str] = None,  # NEW: "date_asc", "date_desc", "amount_desc"
    current_user: User = Depends(require_permission('reports.view'))
):
    """View financial transactions with filters - returns JSON for UI"""
    query = {"is_deleted": False}
    if start_date:
        query['date'] = {"$gte": datetime.fromisoformat(start_date)}
    if end_date:
        end_dt = datetime.fromisoformat(end_date)
        if 'date' in query:
            query['date']['$lte'] = end_dt
        else:
            query['date'] = {"$lte": end_dt}
    if transaction_type:
        query['transaction_type'] = transaction_type
    if account_id:
        query['account_id'] = account_id
    if party_id:
        query['party_id'] = party_id
    
    # Apply sorting
    sort_field = "date"
    sort_direction = -1  # Default: newest first
    
    if sort_by == "date_asc":
        sort_field = "date"
        sort_direction = 1
    elif sort_by == "date_desc":
        sort_field = "date"
        sort_direction = -1
    elif sort_by == "amount_desc":
        sort_field = "amount"
        sort_direction = -1
    
    transactions = await db.transactions.find(query, {"_id": 0}).sort(sort_field, sort_direction).to_list(10000)
    
    # Calculate totals
    total_credit = sum(txn.get('amount', 0) for txn in transactions if txn.get('transaction_type') == 'credit')
    total_debit = sum(txn.get('amount', 0) for txn in transactions if txn.get('transaction_type') == 'debit')
    
    return {
        "transactions": transactions,
        "summary": {
            "total_credit": total_credit,
            "total_debit": total_debit,
            "net_balance": total_credit - total_debit
        },
        "count": len(transactions)
    }

@api_router.get("/reports/invoice/{invoice_id}")
async def get_invoice_report(invoice_id: str, current_user: User = Depends(require_permission('reports.view'))):
    """Get detailed report for a single invoice"""
    invoice = await db.invoices.find_one({"id": invoice_id, "is_deleted": False}, {"_id": 0})
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    
    # Get related payments/transactions
    transactions = await db.transactions.find(
        {"party_id": invoice.get('customer_id'), "is_deleted": False},
        {"_id": 0}
    ).to_list(1000)
    
    return {
        "invoice": invoice,
        "transactions": transactions
    }

@api_router.get("/reports/party/{party_id}/ledger-report")
async def get_party_ledger_report(
    party_id: str,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: User = Depends(require_permission('reports.view'))
):
    """Get detailed ledger report for a party with date filtering"""
    party = await db.parties.find_one({"id": party_id, "is_deleted": False}, {"_id": 0})
    if not party:
        raise HTTPException(status_code=404, detail="Party not found")
    
    # Build query for invoices
    invoice_query = {"customer_id": party_id, "is_deleted": False}
    if start_date:
        invoice_query['date'] = {"$gte": datetime.fromisoformat(start_date)}
    if end_date:
        end_dt = datetime.fromisoformat(end_date)
        if 'date' in invoice_query:
            invoice_query['date']['$lte'] = end_dt
        else:
            invoice_query['date'] = {"$lte": end_dt}
    
    invoices = await db.invoices.find(invoice_query, {"_id": 0}).sort("date", -1).to_list(1000)
    
    # Build query for transactions
    txn_query = {"party_id": party_id, "is_deleted": False}
    if start_date:
        txn_query['date'] = {"$gte": datetime.fromisoformat(start_date)}
    if end_date:
        end_dt = datetime.fromisoformat(end_date)
        if 'date' in txn_query:
            txn_query['date']['$lte'] = end_dt
        else:
            txn_query['date'] = {"$lte": end_dt}
    
    transactions = await db.transactions.find(txn_query, {"_id": 0}).sort("date", -1).to_list(1000)
    
    # Calculate totals
    total_invoiced = sum(inv.get('grand_total', 0) for inv in invoices)
    total_paid = sum(txn.get('amount', 0) for txn in transactions if txn.get('transaction_type') == 'debit')
    total_outstanding = sum(inv.get('balance_due', 0) for inv in invoices)
    
    return {
        "party": party,
        "invoices": invoices,
        "transactions": transactions,
        "summary": {
            "total_invoiced": total_invoiced,
            "total_paid": total_paid,
            "total_outstanding": total_outstanding
        }
    }

@api_router.get("/reports/inventory/{header_id}/stock-report")
async def get_inventory_stock_report(
    header_id: str,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: User = Depends(require_permission('reports.view'))
):
    """Get detailed stock report for a specific inventory category"""
    header = await db.inventory_headers.find_one({"id": header_id, "is_deleted": False}, {"_id": 0})
    if not header:
        raise HTTPException(status_code=404, detail="Inventory category not found")
    
    # Build query for movements
    query = {"header_id": header_id, "is_deleted": False}
    if start_date:
        query['date'] = {"$gte": datetime.fromisoformat(start_date)}
    if end_date:
        end_dt = datetime.fromisoformat(end_date)
        if 'date' in query:
            query['date']['$lte'] = end_dt
        else:
            query['date'] = {"$lte": end_dt}
    
    movements = await db.stock_movements.find(query, {"_id": 0}).sort("date", -1).to_list(10000)
    
    # Calculate stock totals
    total_in = sum(m.get('qty_delta', 0) for m in movements if m.get('qty_delta', 0) > 0)
    total_out = sum(abs(m.get('qty_delta', 0)) for m in movements if m.get('qty_delta', 0) < 0)
    current_stock = total_in - total_out
    
    total_weight_in = sum(m.get('weight_delta', 0) for m in movements if m.get('weight_delta', 0) > 0)
    total_weight_out = sum(abs(m.get('weight_delta', 0)) for m in movements if m.get('weight_delta', 0) < 0)
    current_weight = total_weight_in - total_weight_out
    
    return {
        "header": header,
        "movements": movements,
        "summary": {
            "total_in": total_in,
            "total_out": total_out,
            "current_stock": current_stock,
            "total_weight_in": total_weight_in,
            "total_weight_out": total_weight_out,
            "current_weight": current_weight
        },
        "count": len(movements)
    }

@api_router.get("/reports/financial-summary")
async def get_financial_summary(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: User = Depends(require_permission('reports.view'))
):
    """
    Get financial summary with LEDGER-ACCURATE calculations
    
    CRITICAL: All calculations derived from Accounts + Transactions ONLY
    Invoices are informational - NOT authoritative for balances
    
    Accounting Rules:
    - Total Sales = SUM of credits to Income accounts
    - Cash Balance = Current balance of Cash Asset accounts
    - Bank Balance = Current balance of Bank Asset accounts
    - Total Credit = SUM of all Credit transactions
    - Total Debit = SUM of all Debit transactions
    - Net Flow = Total Credit - Total Debit
    - Net Profit = Total Income - Total Expenses
    """
    # Build query for transactions
    txn_query = {"is_deleted": False}
    if start_date:
        txn_query['date'] = {"$gte": datetime.fromisoformat(start_date)}
    if end_date:
        end_dt = datetime.fromisoformat(end_date)
        if 'date' in txn_query:
            txn_query['date']['$lte'] = end_dt
        else:
            txn_query['date'] = {"$lte": end_dt}
    
    # Build query for invoices (only for outstanding calculation)
    invoice_query = {"is_deleted": False, "status": "finalized"}
    if start_date:
        invoice_query['date'] = {"$gte": datetime.fromisoformat(start_date)}
    if end_date:
        end_dt = datetime.fromisoformat(end_date)
        if 'date' in invoice_query:
            invoice_query['date']['$lte'] = end_dt
        else:
            invoice_query['date'] = {"$lte": end_dt}
    
    # Get data from database
    transactions = await db.transactions.find(txn_query, {"_id": 0}).to_list(10000)
    accounts = await db.accounts.find({"is_deleted": False}, {"_id": 0}).to_list(1000)
    invoices = await db.invoices.find(invoice_query, {"_id": 0}).to_list(10000)
    
    # ============================================================================
    # LEDGER-BASED CALCULATIONS (Authoritative Source)
    # ============================================================================
    
    # Build account maps for efficient lookups
    account_type_map = {acc.get('id'): acc.get('account_type', '').lower() for acc in accounts if 'id' in acc}
    account_name_map = {acc.get('id'): acc.get('name', '') for acc in accounts if 'id' in acc}
    
    # Calculate balances from ACCOUNTS table (current state)
    cash_balance = sum(
        acc.get('current_balance', 0) for acc in accounts 
        if acc.get('account_type', '').lower() == 'asset' and 
        'cash' in acc.get('name', '').lower()
    )
    
    bank_balance = sum(
        acc.get('current_balance', 0) for acc in accounts 
        if acc.get('account_type', '').lower() == 'asset' and 
        'bank' in acc.get('name', '').lower()
    )
    
    # Calculate Total Sales from INCOME ACCOUNTS (credits increase income)
    # Sum of all credits to Income-type accounts
    # Subtract sales returns (debits to income accounts)
    total_sales_credits = sum(
        txn.get('amount', 0) for txn in transactions
        if txn.get('transaction_type') == 'credit' and
        account_type_map.get(txn.get('account_id'), '') == 'income'
    )
    
    # Sales returns reduce total sales (debits or category="sales_return")
    total_sales_returns = sum(
        txn.get('amount', 0) for txn in transactions
        if (txn.get('category') == 'sales_return' or 
            (txn.get('transaction_type') == 'debit' and
             account_type_map.get(txn.get('account_id'), '') == 'income'))
    )
    
    # Net Sales = Gross Sales - Returns
    total_sales = total_sales_credits - total_sales_returns
    
    # Calculate Total Income and Expenses for Net Profit
    total_income = sum(
        acc.get('current_balance', 0) for acc in accounts
        if acc.get('account_type', '').lower() == 'income'
    )
    
    total_expenses = sum(
        acc.get('current_balance', 0) for acc in accounts
        if acc.get('account_type', '').lower() == 'expense'
    )
    
    # Net Profit = Income - Expenses (ledger balances)
    # Note: Expense accounts have positive balances representing costs incurred
    net_profit = total_income - total_expenses
    
    # Calculate Total Credit and Debit from TRANSACTIONS
    total_credit = sum(
        txn.get('amount', 0) for txn in transactions 
        if txn.get('transaction_type') == 'credit'
    )
    
    total_debit = sum(
        txn.get('amount', 0) for txn in transactions 
        if txn.get('transaction_type') == 'debit'
    )
    
    # Net Flow = Total Credits - Total Debits
    net_flow = total_credit - total_debit
    
    # ============================================================================
    # INVOICE-BASED CALCULATIONS (Informational Only)
    # ============================================================================
    
    # Outstanding is still calculated from invoices (customer/vendor balances)
    total_outstanding = sum(inv.get('balance_due', 0) for inv in invoices)
    
    # Total account balance (sum of all account balances)
    total_account_balance = sum(acc.get('current_balance', 0) for acc in accounts)
    
    # Daily closing difference (for reconciliation)
    daily_closing_difference = 0
    closing_query = {"is_deleted": False}
    if start_date and end_date:
        closing_query['date'] = {
            "$gte": datetime.fromisoformat(start_date),
            "$lte": datetime.fromisoformat(end_date)
        }
    else:
        today = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0)
        closing_query['date'] = {"$gte": today}
    
    closings = await db.daily_closings.find(closing_query, {"_id": 0}).to_list(100)
    for closing in closings:
        expected = closing.get('expected_closing', 0)
        actual = closing.get('actual_closing', 0)
        daily_closing_difference += (actual - expected)
    
    # Get returns metrics for the period
    returns_query = {"is_deleted": False, "status": "finalized"}
    if start_date:
        returns_query['date'] = {"$gte": datetime.fromisoformat(start_date)}
    if end_date:
        end_dt = datetime.fromisoformat(end_date)
        if 'date' in returns_query:
            returns_query['date']['$lte'] = end_dt
        else:
            returns_query['date'] = {"$lte": end_dt}
    
    returns_docs = await db.returns.find(returns_query, {"_id": 0}).to_list(1000)
    
    total_sales_returns_count = len([r for r in returns_docs if r.get('return_type') == 'sale_return'])
    total_purchase_returns_count = len([r for r in returns_docs if r.get('return_type') == 'purchase_return'])
    
    return {
        "total_sales": total_sales,  # LEDGER-BASED: Net Sales (Credits - Returns)
        "total_sales_returns": total_sales_returns,  # Total sales returns amount
        "total_credit": total_credit,  # LEDGER-BASED: All credits
        "total_debit": total_debit,  # LEDGER-BASED: All debits
        "net_flow": net_flow,  # LEDGER-BASED: Credit - Debit
        "cash_balance": cash_balance,  # LEDGER-BASED: Cash account balance
        "bank_balance": bank_balance,  # LEDGER-BASED: Bank account balance
        "net_profit": net_profit,  # LEDGER-BASED: Income - Expenses (includes returns impact)
        "total_account_balance": total_account_balance,  # Sum of all accounts
        "total_outstanding": total_outstanding,  # Invoice-based (informational)
        "daily_closing_difference": daily_closing_difference,  # Reconciliation
        "returns_summary": {
            "sales_returns_count": total_sales_returns_count,
            "purchase_returns_count": total_purchase_returns_count,
            "total_returns_count": len(returns_docs)
        }
    }


@api_router.get("/reports/outstanding")
async def get_outstanding_report(
    party_id: Optional[str] = None,
    party_type: Optional[str] = None,  # "customer", "vendor", or None for both
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    include_paid: bool = False,  # Include fully paid invoices
    current_user: User = Depends(require_permission('reports.view'))
):
    """
    Get outstanding report with overdue buckets
    Shows total invoiced, paid, outstanding per party
    Includes overdue buckets: 0-7, 8-30, 31+ days
    """
    # Build invoice query
    invoice_query = {"is_deleted": False, "status": "finalized"}
    
    if not include_paid:
        # Only include invoices with outstanding balance
        invoice_query['balance_due'] = {"$gt": 0}
    
    if start_date:
        invoice_query['date'] = {"$gte": datetime.fromisoformat(start_date)}
    if end_date:
        end_dt = datetime.fromisoformat(end_date)
        if 'date' in invoice_query:
            invoice_query['date']['$lte'] = end_dt
        else:
            invoice_query['date'] = {"$lte": end_dt}
    
    # Get all invoices
    invoices = await db.invoices.find(invoice_query, {"_id": 0}).to_list(10000)
    
    # Get all transactions for payment tracking
    # CRITICAL FIX: Include "Purchase" category for vendor payables from purchase finalization
    transactions = await db.transactions.find(
        {"is_deleted": False, "category": {"$in": ["Sales Invoice", "Purchase Invoice", "Purchase"]}},
        {"_id": 0}
    ).to_list(10000)
    
    # Calculate today for overdue calculations
    today = datetime.now(timezone.utc)
    
    # Group by party
    party_data = {}
    
    for inv in invoices:
        # Determine party info
        if inv.get('customer_type') == 'walk_in':
            party_key = f"walk_in_{inv.get('walk_in_name', 'Unknown')}"
            party_name = f"{inv.get('walk_in_name', 'Unknown')} (Walk-in)"
            ptype = "customer"
        elif inv.get('customer_id'):
            party_key = inv['customer_id']
            party_name = inv.get('customer_name', 'Unknown')
            ptype = "customer" if inv.get('invoice_type') == 'sale' else "vendor"
        else:
            continue
        
        # Filter by party_id if specified
        if party_id and party_key != party_id:
            continue
        
        # Filter by party_type if specified
        if party_type and ptype != party_type:
            continue
        
        # Initialize party data
        if party_key not in party_data:
            party_data[party_key] = {
                "party_id": party_key,
                "party_name": party_name,
                "party_type": ptype,
                "total_invoiced": 0,
                "total_paid": 0,
                "total_outstanding": 0,
                "overdue_0_7": 0,
                "overdue_8_30": 0,
                "overdue_31_plus": 0,
                "last_invoice_date": None,
                "last_payment_date": None,
                "invoice_count": 0
            }
        
        # Update totals
        party_data[party_key]['total_invoiced'] += inv.get('grand_total', 0)
        party_data[party_key]['total_paid'] += inv.get('paid_amount', 0)
        party_data[party_key]['total_outstanding'] += inv.get('balance_due', 0)
        party_data[party_key]['invoice_count'] += 1
        
        # Update last invoice date
        inv_date = inv.get('date')
        if inv_date:
            if isinstance(inv_date, str):
                inv_date = datetime.fromisoformat(inv_date)
            if not party_data[party_key]['last_invoice_date'] or inv_date > party_data[party_key]['last_invoice_date']:
                party_data[party_key]['last_invoice_date'] = inv_date
        
        # Calculate overdue only if balance_due > 0
        if inv.get('balance_due', 0) > 0:
            # Use due_date if available, else fallback to invoice date
            due_date = inv.get('due_date') or inv.get('date')
            if due_date:
                if isinstance(due_date, str):
                    due_date = datetime.fromisoformat(due_date)
                
                # CRITICAL FIX: Ensure due_date is timezone-aware to match today
                if due_date.tzinfo is None:
                    due_date = due_date.replace(tzinfo=timezone.utc)
                
                overdue_days = (today - due_date).days
                
                if overdue_days >= 0:  # Only count if actually overdue
                    outstanding_amount = inv.get('balance_due', 0)
                    if 0 <= overdue_days <= 7:
                        party_data[party_key]['overdue_0_7'] += outstanding_amount
                    elif 8 <= overdue_days <= 30:
                        party_data[party_key]['overdue_8_30'] += outstanding_amount
                    elif overdue_days > 30:
                        party_data[party_key]['overdue_31_plus'] += outstanding_amount
    
    # CRITICAL FIX: Process Purchase transactions to add vendor payables from purchase finalization
    # These are credit transactions (we owe vendor) with category "Purchase" and transaction_type "credit"
    for txn in transactions:
        if txn.get('category') == 'Purchase' and txn.get('transaction_type') == 'credit':
            party_key = txn.get('party_id')
            if not party_key:
                continue
            
            # Filter by party_id if specified
            if party_id and party_key != party_id:
                continue
            
            # Filter by party_type (vendor payables only)
            if party_type and party_type != 'vendor':
                continue
            
            # Get party name
            party_name = txn.get('party_name', 'Unknown Vendor')
            
            # Initialize party data if not exists
            if party_key not in party_data:
                party_data[party_key] = {
                    "party_id": party_key,
                    "party_name": party_name,
                    "party_type": "vendor",
                    "total_invoiced": 0,
                    "total_paid": 0,
                    "total_outstanding": 0,
                    "overdue_0_7": 0,
                    "overdue_8_30": 0,
                    "overdue_31_plus": 0,
                    "last_invoice_date": None,
                    "last_payment_date": None,
                    "invoice_count": 0
                }
            
            # Add vendor payable amount to outstanding
            txn_amount = txn.get('amount', 0)
            party_data[party_key]['total_outstanding'] += txn_amount
            
            # Calculate overdue for vendor payables
            txn_date = txn.get('date')
            if txn_date and txn_amount > 0:
                if isinstance(txn_date, str):
                    txn_date = datetime.fromisoformat(txn_date)
                
                # CRITICAL FIX: Ensure txn_date is timezone-aware to match today
                if txn_date.tzinfo is None:
                    txn_date = txn_date.replace(tzinfo=timezone.utc)
                
                overdue_days = (today - txn_date).days
                
                if overdue_days >= 0:  # Only count if actually overdue
                    if 0 <= overdue_days <= 7:
                        party_data[party_key]['overdue_0_7'] += txn_amount
                    elif 8 <= overdue_days <= 30:
                        party_data[party_key]['overdue_8_30'] += txn_amount
                    elif overdue_days > 30:
                        party_data[party_key]['overdue_31_plus'] += txn_amount
    
    # Get last payment dates from transactions
    for txn in transactions:
        party_key = txn.get('party_id')
        if party_key and party_key in party_data:
            txn_date = txn.get('date')
            if txn_date:
                if isinstance(txn_date, str):
                    txn_date = datetime.fromisoformat(txn_date)
                if not party_data[party_key]['last_payment_date'] or txn_date > party_data[party_key]['last_payment_date']:
                    party_data[party_key]['last_payment_date'] = txn_date
    
    # Convert dates to ISO strings for JSON serialization
    for party in party_data.values():
        if party['last_invoice_date']:
            party['last_invoice_date'] = party['last_invoice_date'].isoformat()
        if party['last_payment_date']:
            party['last_payment_date'] = party['last_payment_date'].isoformat()
    
    # Calculate summary totals
    customer_due = sum(p['total_outstanding'] for p in party_data.values() if p['party_type'] == 'customer')
    vendor_payable = sum(p['total_outstanding'] for p in party_data.values() if p['party_type'] == 'vendor')
    total_outstanding = customer_due + vendor_payable
    
    total_overdue_0_7 = sum(p['overdue_0_7'] for p in party_data.values())
    total_overdue_8_30 = sum(p['overdue_8_30'] for p in party_data.values())
    total_overdue_31_plus = sum(p['overdue_31_plus'] for p in party_data.values())
    
    return {
        "summary": {
            "customer_due": customer_due,
            "vendor_payable": vendor_payable,
            "total_outstanding": total_outstanding,
            "total_overdue_0_7": total_overdue_0_7,
            "total_overdue_8_30": total_overdue_8_30,
            "total_overdue_31_plus": total_overdue_31_plus
        },
        "parties": list(party_data.values())
    }


# ==================== PDF EXPORT ENDPOINTS ====================

@api_router.get("/reports/outstanding-pdf")
async def export_outstanding_pdf(
    party_id: Optional[str] = None,
    party_type: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: User = Depends(require_permission('reports.view'))
):
    """Export outstanding report as PDF"""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import inch
    from reportlab.pdfgen import canvas
    from reportlab.lib import colors
    from reportlab.platypus import Table, TableStyle
    from io import BytesIO
    from fastapi.responses import StreamingResponse
    
    # Get data from outstanding report
    params = {}
    if party_id:
        params['party_id'] = party_id
    if party_type:
        params['party_type'] = party_type
    if start_date:
        params['start_date'] = start_date
    if end_date:
        params['end_date'] = end_date
    
    # Call outstanding report endpoint logic
    from urllib.parse import urlencode
    import httpx
    
    # Get outstanding data by calling the endpoint function directly
    data = await get_outstanding_report(
        party_id=party_id,
        party_type=party_type,
        start_date=start_date,
        end_date=end_date,
        include_paid=False,
        current_user=current_user
    )
    
    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    
    # Header
    c.setFont("Helvetica-Bold", 16)
    c.drawString(inch, height - inch, "Outstanding Report")
    
    # Date range
    c.setFont("Helvetica", 10)
    date_str = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    if start_date or end_date:
        date_str += f" | Period: {start_date or 'Start'} to {end_date or 'End'}"
    c.drawString(inch, height - inch - 0.3*inch, date_str)
    
    # Summary section
    y_position = height - inch - 0.8*inch
    c.setFont("Helvetica-Bold", 12)
    c.drawString(inch, y_position, "Summary")
    y_position -= 0.3*inch
    
    c.setFont("Helvetica", 10)
    summary = data['summary']
    c.drawString(inch, y_position, f"Customer Due: {summary['customer_due']:.3f}")
    c.drawString(inch + 2.5*inch, y_position, f"Vendor Payable: {summary['vendor_payable']:.3f}")
    y_position -= 0.2*inch
    c.drawString(inch, y_position, f"Total Outstanding: {summary['total_outstanding']:.3f}")
    y_position -= 0.3*inch
    
    c.setFont("Helvetica-Bold", 11)
    c.drawString(inch, y_position, "Overdue Buckets:")
    y_position -= 0.2*inch
    c.setFont("Helvetica", 10)
    c.drawString(inch, y_position, f"0-7 days: {summary['total_overdue_0_7']:.3f}")
    c.drawString(inch + 2*inch, y_position, f"8-30 days: {summary['total_overdue_8_30']:.3f}")
    c.drawString(inch + 4*inch, y_position, f"31+ days: {summary['total_overdue_31_plus']:.3f}")
    y_position -= 0.5*inch
    
    # Parties table
    c.setFont("Helvetica-Bold", 12)
    c.drawString(inch, y_position, "Party-wise Outstanding")
    y_position -= 0.3*inch
    
    # Create table data
    table_data = [['Party Name', 'Type', 'Invoiced', 'Paid', 'Outstanding', '0-7d', '8-30d', '31+d']]
    for party in data['parties'][:20]:  # Limit to 20 parties per page
        table_data.append([
            party['party_name'][:25],
            party['party_type'],
            f"{party['total_invoiced']:.2f}",
            f"{party['total_paid']:.2f}",
            f"{party['total_outstanding']:.2f}",
            f"{party['overdue_0_7']:.2f}",
            f"{party['overdue_8_30']:.2f}",
            f"{party['overdue_31_plus']:.2f}"
        ])
    
    # Create table
    table = Table(table_data, colWidths=[2*inch, 0.7*inch, 0.8*inch, 0.8*inch, 0.9*inch, 0.6*inch, 0.7*inch, 0.7*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    # Draw table
    table.wrapOn(c, width, height)
    table.drawOn(c, inch, y_position - len(table_data) * 0.25*inch)
    
    c.save()
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=outstanding_report_{datetime.now().strftime('%Y%m%d')}.pdf"}
    )


@api_router.get("/reports/invoices-pdf")
async def export_invoices_pdf(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    invoice_type: Optional[str] = None,
    payment_status: Optional[str] = None,
    party_id: Optional[str] = None,
    current_user: User = Depends(require_permission('reports.view'))
):
    """Export invoices report as PDF"""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import inch
    from reportlab.pdfgen import canvas
    from reportlab.lib import colors
    from reportlab.platypus import Table, TableStyle
    from io import BytesIO
    from fastapi.responses import StreamingResponse
    
    # Get data
    data = await view_invoices_report(
        start_date=start_date,
        end_date=end_date,
        invoice_type=invoice_type,
        payment_status=payment_status,
        party_id=party_id,
        sort_by=None,
        current_user=current_user
    )
    
    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    
    # Header
    c.setFont("Helvetica-Bold", 16)
    c.drawString(inch, height - inch, "Invoices Report")
    
    c.setFont("Helvetica", 10)
    date_str = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    if start_date or end_date:
        date_str += f" | Period: {start_date or 'Start'} to {end_date or 'End'}"
    c.drawString(inch, height - inch - 0.3*inch, date_str)
    
    # Summary
    y_position = height - inch - 0.8*inch
    c.setFont("Helvetica-Bold", 12)
    c.drawString(inch, y_position, "Summary")
    y_position -= 0.3*inch
    
    c.setFont("Helvetica", 10)
    summary = data['summary']
    c.drawString(inch, y_position, f"Total Amount: {summary['total_amount']:.3f}")
    c.drawString(inch + 2.5*inch, y_position, f"Total Paid: {summary['total_paid']:.3f}")
    y_position -= 0.2*inch
    c.drawString(inch, y_position, f"Total Balance: {summary['total_balance']:.3f}")
    c.drawString(inch + 2.5*inch, y_position, f"Count: {data['count']}")
    y_position -= 0.5*inch
    
    # Table
    c.setFont("Helvetica-Bold", 12)
    c.drawString(inch, y_position, "Invoices")
    y_position -= 0.3*inch
    
    table_data = [['Invoice #', 'Date', 'Customer', 'Type', 'Amount', 'Paid', 'Balance']]
    for inv in data['invoices'][:25]:
        inv_date = inv.get('date', '')
        if isinstance(inv_date, str):
            inv_date = inv_date[:10]
        elif hasattr(inv_date, 'strftime'):
            inv_date = inv_date.strftime('%Y-%m-%d')
        
        customer = inv.get('customer_name') or inv.get('walk_in_name') or 'N/A'
        table_data.append([
            inv.get('invoice_number', '')[:15],
            inv_date,
            customer[:20],
            inv.get('invoice_type', '')[:4],
            f"{inv.get('grand_total', 0):.2f}",
            f"{inv.get('paid_amount', 0):.2f}",
            f"{inv.get('balance_due', 0):.2f}"
        ])
    
    table = Table(table_data, colWidths=[1.2*inch, 0.9*inch, 1.5*inch, 0.6*inch, 0.8*inch, 0.8*inch, 0.8*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    table.wrapOn(c, width, height)
    table.drawOn(c, inch, y_position - len(table_data) * 0.25*inch)
    
    c.save()
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=invoices_report_{datetime.now().strftime('%Y%m%d')}.pdf"}
    )


@api_router.get("/reports/parties-pdf")
async def export_parties_pdf(
    party_type: Optional[str] = None,
    current_user: User = Depends(require_permission('reports.view'))
):
    """Export parties report as PDF"""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import inch
    from reportlab.pdfgen import canvas
    from reportlab.lib import colors
    from reportlab.platypus import Table, TableStyle
    from io import BytesIO
    from fastapi.responses import StreamingResponse
    
    # Get data
    data = await view_parties_report(
        party_type=party_type,
        sort_by="outstanding_desc",
        current_user=current_user
    )
    
    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    
    # Header
    c.setFont("Helvetica-Bold", 16)
    c.drawString(inch, height - inch, "Parties Report")
    
    c.setFont("Helvetica", 10)
    c.drawString(inch, height - inch - 0.3*inch, f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    
    # Table
    y_position = height - inch - 0.8*inch
    c.setFont("Helvetica-Bold", 12)
    c.drawString(inch, y_position, f"Total Parties: {data['count']}")
    y_position -= 0.4*inch
    
    table_data = [['Party Name', 'Type', 'Phone', 'Email', 'Outstanding']]
    for party in data['parties'][:30]:
        table_data.append([
            party.get('name', '')[:25],
            party.get('party_type', '')[:8],
            party.get('phone', '')[:15],
            party.get('email', '')[:20],
            f"{party.get('outstanding', 0):.2f}"
        ])
    
    table = Table(table_data, colWidths=[2*inch, 0.8*inch, 1.2*inch, 1.5*inch, 1*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    table.wrapOn(c, width, height)
    table.drawOn(c, inch, y_position - len(table_data) * 0.25*inch)
    
    c.save()
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=parties_report_{datetime.now().strftime('%Y%m%d')}.pdf"}
    )


@api_router.get("/reports/transactions-pdf")
async def export_transactions_pdf(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    transaction_type: Optional[str] = None,
    party_id: Optional[str] = None,
    current_user: User = Depends(require_permission('reports.view'))
):
    """Export transactions report as PDF"""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import inch
    from reportlab.pdfgen import canvas
    from reportlab.lib import colors
    from reportlab.platypus import Table, TableStyle
    from io import BytesIO
    from fastapi.responses import StreamingResponse
    
    # Get data
    data = await view_transactions_report(
        start_date=start_date,
        end_date=end_date,
        transaction_type=transaction_type,
        account_id=None,
        party_id=party_id,
        sort_by=None,
        current_user=current_user
    )
    
    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    
    # Header
    c.setFont("Helvetica-Bold", 16)
    c.drawString(inch, height - inch, "Transactions Report")
    
    c.setFont("Helvetica", 10)
    date_str = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    if start_date or end_date:
        date_str += f" | Period: {start_date or 'Start'} to {end_date or 'End'}"
    c.drawString(inch, height - inch - 0.3*inch, date_str)
    
    # Summary
    y_position = height - inch - 0.8*inch
    c.setFont("Helvetica-Bold", 12)
    c.drawString(inch, y_position, "Summary")
    y_position -= 0.3*inch
    
    c.setFont("Helvetica", 10)
    summary = data['summary']
    c.drawString(inch, y_position, f"Total Credit: {summary['total_credit']:.3f}")
    c.drawString(inch + 2.5*inch, y_position, f"Total Debit: {summary['total_debit']:.3f}")
    y_position -= 0.2*inch
    c.drawString(inch, y_position, f"Net Balance: {summary['net_balance']:.3f}")
    c.drawString(inch + 2.5*inch, y_position, f"Count: {data['count']}")
    y_position -= 0.5*inch
    
    # Table
    c.setFont("Helvetica-Bold", 12)
    c.drawString(inch, y_position, "Transactions")
    y_position -= 0.3*inch
    
    table_data = [['TXN #', 'Date', 'Type', 'Account', 'Party', 'Amount']]
    for txn in data['transactions'][:30]:
        txn_date = txn.get('date', '')
        if isinstance(txn_date, str):
            txn_date = txn_date[:10]
        elif hasattr(txn_date, 'strftime'):
            txn_date = txn_date.strftime('%Y-%m-%d')
        
        table_data.append([
            txn.get('transaction_number', '')[:15],
            txn_date,
            txn.get('transaction_type', '')[:6],
            txn.get('account_name', '')[:20],
            txn.get('party_name', 'N/A')[:15],
            f"{txn.get('amount', 0):.2f}"
        ])
    
    table = Table(table_data, colWidths=[1.2*inch, 0.9*inch, 0.7*inch, 1.5*inch, 1.2*inch, 0.8*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    table.wrapOn(c, width, height)
    table.drawOn(c, inch, y_position - len(table_data) * 0.25*inch)
    
    c.save()
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=transactions_report_{datetime.now().strftime('%Y%m%d')}.pdf"}
    )


@api_router.get("/reports/inventory-pdf")
async def export_inventory_pdf(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    movement_type: Optional[str] = None,
    category: Optional[str] = None,
    current_user: User = Depends(require_permission('reports.view'))
):
    """Export inventory report as PDF"""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import inch
    from reportlab.pdfgen import canvas
    from reportlab.lib import colors
    from reportlab.platypus import Table, TableStyle
    from io import BytesIO
    from fastapi.responses import StreamingResponse
    
    # Get data
    data = await view_inventory_report(
        start_date=start_date,
        end_date=end_date,
        movement_type=movement_type,
        category=category,
        sort_by=None,
        current_user=current_user
    )
    
    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    
    # Header
    c.setFont("Helvetica-Bold", 16)
    c.drawString(inch, height - inch, "Inventory Report")
    
    c.setFont("Helvetica", 10)
    date_str = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    if start_date or end_date:
        date_str += f" | Period: {start_date or 'Start'} to {end_date or 'End'}"
    c.drawString(inch, height - inch - 0.3*inch, date_str)
    
    # Summary
    y_position = height - inch - 0.8*inch
    c.setFont("Helvetica-Bold", 12)
    c.drawString(inch, y_position, "Summary")
    y_position -= 0.3*inch
    
    c.setFont("Helvetica", 10)
    summary = data['summary']
    c.drawString(inch, y_position, f"Total In: {summary['total_in']:.2f} pcs")
    c.drawString(inch + 2.5*inch, y_position, f"Total Out: {summary['total_out']:.2f} pcs")
    y_position -= 0.2*inch
    c.drawString(inch, y_position, f"Weight In: {summary['total_weight_in']:.3f} g")
    c.drawString(inch + 2.5*inch, y_position, f"Weight Out: {summary['total_weight_out']:.3f} g")
    y_position -= 0.5*inch
    
    # Table
    c.setFont("Helvetica-Bold", 12)
    c.drawString(inch, y_position, "Stock Movements")
    y_position -= 0.3*inch
    
    table_data = [['Date', 'Category', 'Type', 'Qty', 'Weight', 'Reference']]
    for mov in data['movements'][:30]:
        mov_date = mov.get('date', '')
        if isinstance(mov_date, str):
            mov_date = mov_date[:10]
        elif hasattr(mov_date, 'strftime'):
            mov_date = mov_date.strftime('%Y-%m-%d')
        
        table_data.append([
            mov_date,
            (mov.get('header_name') or '')[:15],
            (mov.get('movement_type') or '')[:10],
            f"{mov.get('qty_delta', 0):.1f}",
            f"{mov.get('weight_delta', 0):.2f}",
            (mov.get('reference_type') or '')[:12]
        ])
    
    table = Table(table_data, colWidths=[0.9*inch, 1.3*inch, 1*inch, 0.7*inch, 0.9*inch, 1.2*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    table.wrapOn(c, width, height)
    table.drawOn(c, inch, y_position - len(table_data) * 0.25*inch)
    
    c.save()
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=inventory_report_{datetime.now().strftime('%Y%m%d')}.pdf"}
    )


# ============================================================================
# MODULE 5/10: SALES HISTORY REPORT (Finalized Invoices Only)
# ============================================================================

@api_router.get("/reports/sales-history")
async def get_sales_history_report(
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    party_id: Optional[str] = None,
    search: Optional[str] = None,
    current_user: User = Depends(require_permission('reports.view'))
):
    """
    Get sales history report showing ONLY finalized invoices.
    
    Filters:
    - date_from/date_to: Date range filter
    - party_id: Filter by specific party (or "all" for all parties)
    - search: Search in customer name, phone, or invoice_id
    
    Returns table with:
    - invoice_id
    - customer name + phone (handles both saved and walk-in)
    - date
    - total_weight_grams (sum of all item weights)
    - purity summary ("Mixed" if multiple purities, otherwise single purity)
    - grand_total
    """
    # Query for FINALIZED invoices only
    query = {
        "is_deleted": False,
        "status": "finalized"  # CRITICAL: Only finalized invoices
    }
    
    # Date filters
    if date_from:
        query['date'] = {"$gte": datetime.fromisoformat(date_from)}
    if date_to:
        end_dt = datetime.fromisoformat(date_to)
        if 'date' in query:
            query['date']['$lte'] = end_dt
        else:
            query['date'] = {"$lte": end_dt}
    
    # Party filter
    if party_id and party_id != 'all':
        query['customer_id'] = party_id
    
    # Get invoices
    invoices = await db.invoices.find(query, {"_id": 0}).sort("date", -1).to_list(10000)
    
    # Process each invoice to calculate totals and purity summary
    sales_records = []
    total_sales = 0.0
    total_weight = 0.0
    
    for inv in invoices:
        # Get customer info (handle both saved and walk-in)
        if inv.get('customer_type') == 'walk_in':
            customer_name = inv.get('walk_in_name', 'Walk-in Customer')
            customer_phone = inv.get('walk_in_phone', '')
        else:
            customer_name = inv.get('customer_name', 'Unknown Customer')
            # Fetch phone from party if saved customer
            customer_phone = ''
            if inv.get('customer_id'):
                party = await db.parties.find_one({"id": inv['customer_id']}, {"_id": 0, "phone": 1})
                if party:
                    customer_phone = party.get('phone', '')
        
        # Calculate total weight and purity summary
        items = inv.get('items', [])
        invoice_weight = sum(item.get('weight', 0) for item in items)
        
        # Check for purity diversity
        purities = list(set(item.get('purity') for item in items if item.get('purity')))
        if len(purities) == 0:
            purity_summary = "N/A"
        elif len(purities) == 1:
            purity_summary = f"{purities[0]}K"
        else:
            purity_summary = "Mixed"
        
        # Apply search filter (if provided)
        if search:
            search_lower = search.lower()
            if not (
                search_lower in customer_name.lower() or
                search_lower in customer_phone.lower() or
                search_lower in inv.get('invoice_number', '').lower()
            ):
                continue  # Skip this invoice if search doesn't match
        
        # Format date
        invoice_date = inv.get('date', '')
        if isinstance(invoice_date, str):
            invoice_date = invoice_date[:10]
        elif hasattr(invoice_date, 'strftime'):
            invoice_date = invoice_date.strftime('%Y-%m-%d')
        
        sales_records.append({
            "invoice_id": inv.get('invoice_number', ''),
            "customer_name": customer_name,
            "customer_phone": customer_phone,
            "date": invoice_date,
            "total_weight_grams": round(invoice_weight, 3),
            "purity_summary": purity_summary,
            "grand_total": round(inv.get('grand_total', 0), 2)
        })
        
        total_sales += inv.get('grand_total', 0)
        total_weight += invoice_weight
    
    return {
        "sales_records": sales_records,
        "summary": {
            "total_sales": round(total_sales, 2),
            "total_weight": round(total_weight, 3),
            "total_invoices": len(sales_records)
        }
    }


@api_router.get("/reports/sales-history-export")
async def export_sales_history(
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    party_id: Optional[str] = None,
    search: Optional[str] = None,
    current_user: User = Depends(require_permission('reports.view'))
):
    """Export sales history report as Excel file with applied filters"""
    from fastapi.responses import StreamingResponse
    from io import BytesIO
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    
    # Get data using the main report function
    data = await get_sales_history_report(
        date_from=date_from,
        date_to=date_to,
        party_id=party_id,
        search=search,
        current_user=current_user
    )
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales History"
    
    # Add summary section
    ws.merge_cells('A1:G1')
    title_cell = ws['A1']
    title_cell.value = "Sales History Report (Finalized Invoices)"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal='center')
    
    ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    if date_from or date_to:
        ws['A3'] = f"Period: {date_from or 'Start'} to {date_to or 'End'}"
    
    # Summary row
    summary_row = 5
    ws[f'A{summary_row}'] = "Total Invoices:"
    ws[f'B{summary_row}'] = data['summary']['total_invoices']
    ws[f'C{summary_row}'] = "Total Weight:"
    ws[f'D{summary_row}'] = f"{data['summary']['total_weight']:.3f} g"
    ws[f'E{summary_row}'] = "Total Sales:"
    ws[f'F{summary_row}'] = f"{data['summary']['total_sales']:.2f} OMR"
    
    # Headers
    header_row = summary_row + 2
    headers = ["Invoice #", "Customer Name", "Phone", "Date", "Weight (g)", "Purity", "Grand Total (OMR)"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    
    # Data rows
    for row_idx, record in enumerate(data['sales_records'], header_row + 1):
        ws.cell(row=row_idx, column=1, value=record['invoice_id'])
        ws.cell(row=row_idx, column=2, value=record['customer_name'])
        ws.cell(row=row_idx, column=3, value=record['customer_phone'])
        ws.cell(row=row_idx, column=4, value=record['date'])
        ws.cell(row=row_idx, column=5, value=record['total_weight_grams'])
        ws.cell(row=row_idx, column=6, value=record['purity_summary'])
        ws.cell(row=row_idx, column=7, value=record['grand_total'])
    
    # Column widths
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 18
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"sales_history_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@api_router.get("/reports/sales-history-pdf")
async def export_sales_history_pdf(
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    party_id: Optional[str] = None,
    search: Optional[str] = None,
    current_user: User = Depends(require_permission('reports.view'))
):
    """Export sales history report as PDF"""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import inch
    from reportlab.pdfgen import canvas
    from reportlab.lib import colors
    from reportlab.platypus import Table, TableStyle
    from io import BytesIO
    from fastapi.responses import StreamingResponse
    
    # Get data using the main report function
    data = await get_sales_history_report(
        date_from=date_from,
        date_to=date_to,
        party_id=party_id,
        search=search,
        current_user=current_user
    )
    
    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    
    # Header
    c.setFont("Helvetica-Bold", 16)
    c.drawString(inch, height - inch, "Sales History Report")
    
    c.setFont("Helvetica", 10)
    date_str = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    if date_from or date_to:
        date_str += f" | Period: {date_from or 'Start'} to {date_to or 'End'}"
    c.drawString(inch, height - inch - 0.3*inch, date_str)
    
    # Summary section
    y_position = height - inch - 0.8*inch
    c.setFont("Helvetica-Bold", 12)
    c.drawString(inch, y_position, "Summary")
    y_position -= 0.3*inch
    
    c.setFont("Helvetica", 10)
    summary = data['summary']
    c.drawString(inch, y_position, f"Total Invoices: {summary['total_invoices']}")
    c.drawString(inch + 2.5*inch, y_position, f"Total Weight: {summary['total_weight']:.3f} g")
    y_position -= 0.2*inch
    c.drawString(inch, y_position, f"Total Sales: {summary['total_sales']:.2f} OMR")
    y_position -= 0.5*inch
    
    # Table header
    c.setFont("Helvetica-Bold", 12)
    c.drawString(inch, y_position, "Sales Records")
    y_position -= 0.3*inch
    
    # Create table data
    table_data = [['Invoice #', 'Customer', 'Phone', 'Date', 'Weight (g)', 'Purity', 'Total (OMR)']]
    
    # Add sales records (limit to 30 records per page for now)
    for record in data['sales_records'][:30]:
        table_data.append([
            record.get('invoice_id', '')[:15],
            record.get('customer_name', '')[:20],
            record.get('customer_phone', '')[:12],
            record.get('date', '')[:10],
            f"{record.get('total_weight_grams', 0):.2f}",
            record.get('purity_summary', ''),
            f"{record.get('grand_total', 0):.2f}"
        ])
    
    # Create and style table
    table = Table(table_data, colWidths=[1.0*inch, 1.3*inch, 0.9*inch, 0.8*inch, 0.8*inch, 0.7*inch, 0.9*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE')
    ]))
    
    # Draw table
    table.wrapOn(c, width, height)
    table.drawOn(c, inch, y_position - len(table_data) * 0.25*inch)
    
    c.save()
    buffer.seek(0)
    
    filename = f"sales_history_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@api_router.get("/reports/purchase-history")
async def get_purchase_history_report(
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    vendor_party_id: Optional[str] = None,
    search: Optional[str] = None,
    current_user: User = Depends(require_permission('reports.view'))
):
    """
    Get purchase history report showing ALL committed purchases (Paid, Partially Paid, Finalized Unpaid).
    
    Filters:
    - date_from/date_to: Date range filter
    - vendor_party_id: Filter by specific vendor (or "all" for all vendors)
    - search: Search in vendor name, phone, or description
    
    Returns table with:
    - vendor_name + phone (from parties collection)
    - date
    - weight_grams (3 decimal precision)
    - entered_purity
    - valuation_purity (converted to "22K" display)
    - amount_total
    
    Summary includes:
    - total_amount (sum of all purchases)
    - total_weight (sum of all weights)
    - total_purchases (count)
    """
    # Query for ALL COMMITTED purchases (excludes only Draft/Voided)
    query = {
        "is_deleted": False,
        "status": {"$in": ["Paid", "Partially Paid", "Finalized (Unpaid)"]}  # CRITICAL: All committed purchases
    }
    
    # Date filters
    if date_from:
        query['date'] = {"$gte": datetime.fromisoformat(date_from)}
    if date_to:
        end_dt = datetime.fromisoformat(date_to)
        if 'date' in query:
            query['date']['$lte'] = end_dt
        else:
            query['date'] = {"$lte": end_dt}
    
    # Vendor filter
    if vendor_party_id and vendor_party_id != 'all':
        query['vendor_party_id'] = vendor_party_id
    
    # Get purchases
    purchases = await db.purchases.find(query, {"_id": 0}).sort("date", -1).to_list(10000)
    
    # Process each purchase
    purchase_records = []
    total_amount = 0.0
    total_weight = 0.0
    
    for purchase in purchases:
        # Get vendor info from parties collection
        vendor_name = "Unknown Vendor"
        vendor_phone = ""
        if purchase.get('vendor_party_id'):
            vendor = await db.parties.find_one(
                {"id": purchase['vendor_party_id'], "is_deleted": False},
                {"_id": 0, "name": 1, "phone": 1}
            )
            if vendor:
                vendor_name = vendor.get('name', 'Unknown Vendor')
                vendor_phone = vendor.get('phone', '')
        
        # Apply search filter (if provided)
        if search:
            search_lower = search.lower()
            if not (
                search_lower in vendor_name.lower() or
                search_lower in vendor_phone.lower() or
                search_lower in purchase.get('description', '').lower()
            ):
                continue  # Skip this purchase if search doesn't match
        
        # Format date
        purchase_date = purchase.get('date', '')
        if isinstance(purchase_date, str):
            purchase_date = purchase_date[:10]
        elif hasattr(purchase_date, 'strftime'):
            purchase_date = purchase_date.strftime('%Y-%m-%d')
        
        # Convert valuation_purity 916 to "22K" display
        valuation_purity_display = "22K"  # 916 purity = 22K
        
        purchase_records.append({
            "vendor_name": vendor_name,
            "vendor_phone": vendor_phone,
            "date": purchase_date,
            "description": purchase.get('description', ''),
            "weight_grams": round(purchase.get('weight_grams', 0), 3),
            "entered_purity": purchase.get('entered_purity', 0),
            "valuation_purity": valuation_purity_display,
            "amount_total": round(purchase.get('amount_total', 0), 2)
        })
        
        total_amount += purchase.get('amount_total', 0)
        total_weight += purchase.get('weight_grams', 0)
    
    return {
        "purchase_records": purchase_records,
        "summary": {
            "total_amount": round(total_amount, 2),
            "total_weight": round(total_weight, 3),
            "total_purchases": len(purchase_records)
        }
    }

@api_router.get("/reports/purchase-history-export")
async def export_purchase_history(
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    vendor_party_id: Optional[str] = None,
    search: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Export purchase history report as Excel file with applied filters"""
    from fastapi.responses import StreamingResponse
    from io import BytesIO
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    
    # Get data using the main report function
    data = await get_purchase_history_report(
        date_from=date_from,
        date_to=date_to,
        vendor_party_id=vendor_party_id,
        search=search,
        current_user=current_user
    )
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Purchase History"
    
    # Add summary section
    ws.merge_cells('A1:H1')
    title_cell = ws['A1']
    title_cell.value = "Purchase History Report (All Committed Purchases)"

    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal='center')
    
    ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    if date_from or date_to:
        ws['A3'] = f"Period: {date_from or 'Start'} to {date_to or 'End'}"
    
    # Summary row
    summary_row = 5
    ws[f'A{summary_row}'] = "Total Purchases:"
    ws[f'B{summary_row}'] = data['summary']['total_purchases']
    ws[f'C{summary_row}'] = "Total Weight:"
    ws[f'D{summary_row}'] = f"{data['summary']['total_weight']:.3f} g"
    ws[f'E{summary_row}'] = "Total Amount:"
    ws[f'F{summary_row}'] = f"{data['summary']['total_amount']:.2f} OMR"
    
    # Headers
    header_row = summary_row + 2
    headers = ["Vendor Name", "Phone", "Date", "Description", "Weight (g)", "Entered Purity", "Valuation Purity", "Amount (OMR)"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    
    # Data rows
    for row_idx, record in enumerate(data['purchase_records'], header_row + 1):
        ws.cell(row=row_idx, column=1, value=record['vendor_name'])
        ws.cell(row=row_idx, column=2, value=record['vendor_phone'])
        ws.cell(row=row_idx, column=3, value=record['date'])
        ws.cell(row=row_idx, column=4, value=record['description'])
        ws.cell(row=row_idx, column=5, value=record['weight_grams'])
        ws.cell(row=row_idx, column=6, value=record['entered_purity'])
        ws.cell(row=row_idx, column=7, value=record['valuation_purity'])
        ws.cell(row=row_idx, column=8, value=record['amount_total'])
    
    # Column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 18
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"purchase_history_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )



@api_router.get("/reports/purchase-history-pdf")
async def export_purchase_history_pdf(
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    vendor_party_id: Optional[str] = None,
    search: Optional[str] = None,
    current_user: User = Depends(require_permission('reports.view'))
):
    """Export purchase history report as PDF"""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import inch
    from reportlab.pdfgen import canvas
    from reportlab.lib import colors
    from reportlab.platypus import Table, TableStyle
    from io import BytesIO
    from fastapi.responses import StreamingResponse
    
    # Get data using the main report function
    data = await get_purchase_history_report(
        date_from=date_from,
        date_to=date_to,
        vendor_party_id=vendor_party_id,
        search=search,
        current_user=current_user
    )
    
    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    
    # Header
    c.setFont("Helvetica-Bold", 16)
    c.drawString(inch, height - inch, "Purchase History Report (All Committed)")
    
    c.setFont("Helvetica", 10)
    date_str = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    if date_from or date_to:
        date_str += f" | Period: {date_from or 'Start'} to {date_to or 'End'}"
    c.drawString(inch, height - inch - 0.3*inch, date_str)
    
    # Summary section
    y_position = height - inch - 0.8*inch
    c.setFont("Helvetica-Bold", 12)
    c.drawString(inch, y_position, "Summary")
    y_position -= 0.3*inch
    
    c.setFont("Helvetica", 10)
    summary = data['summary']
    c.drawString(inch, y_position, f"Total Purchases: {summary['total_purchases']}")
    c.drawString(inch + 2.5*inch, y_position, f"Total Weight: {summary['total_weight']:.3f} g")
    y_position -= 0.2*inch
    c.drawString(inch, y_position, f"Total Amount: {summary['total_amount']:.2f} OMR")
    y_position -= 0.5*inch
    
    # Table header
    c.setFont("Helvetica-Bold", 12)
    c.drawString(inch, y_position, "Purchase Records")
    y_position -= 0.3*inch
    
    # Create table data
    table_data = [['Vendor', 'Phone', 'Date', 'Weight (g)', 'Purity', 'Amount (OMR)']]
    
    # Add purchase records (limit to 30 records per page for now)
    for record in data['purchase_records'][:30]:
        table_data.append([
            record.get('vendor_name', '')[:20],
            record.get('vendor_phone', '')[:12],
            record.get('date', '')[:10],
            f"{record.get('weight_grams', 0):.2f}",
            f"{record.get('entered_purity', '')}K",
            f"{record.get('amount_total', 0):.2f}"
        ])
    
    # Create and style table
    table = Table(table_data, colWidths=[1.5*inch, 1.0*inch, 0.9*inch, 0.9*inch, 0.8*inch, 1.0*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE')
    ]))
    
    # Draw table
    table.wrapOn(c, width, height)
    table.drawOn(c, inch, y_position - len(table_data) * 0.25*inch)
    
    c.save()
    buffer.seek(0)
    
    filename = f"purchase_history_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ============================================================================
# RETURNS REPORT ENDPOINTS
# ============================================================================

@api_router.get("/reports/returns-summary")
async def get_returns_summary_report(
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    return_type: Optional[str] = None,
    status: Optional[str] = None,
    refund_mode: Optional[str] = None,
    party_id: Optional[str] = None,
    search: Optional[str] = None,
    current_user: User = Depends(require_permission('reports.view'))
):
    """
    Get returns summary report with filters.
    
    Summary includes:
    - total_returns_count: Total number of finalized returns
    - sales_returns_amount: Total amount of sales returns (OMR)
    - purchase_returns_amount: Total amount of purchase returns (OMR)
    - total_refund_amount: Total refund amount across all returns (OMR)
    - money_refunded: Total money refunded (OMR)
    - gold_refunded: Total gold weight refunded (grams)
    
    Note: Only FINALIZED returns are included in financial totals.
    Draft returns are excluded from metrics.
    """
    # Build query - only finalized returns for summary
    query = {
        "is_deleted": False,
        "status": "finalized"  # Only finalized returns count in financial totals
    }
    
    # Date filters
    if date_from:
        query['date'] = {"$gte": datetime.fromisoformat(date_from)}
    if date_to:
        end_dt = datetime.fromisoformat(date_to)
        if 'date' in query:
            query['date']['$lte'] = end_dt
        else:
            query['date'] = {"$lte": end_dt}
    
    # Return type filter
    if return_type and return_type != 'all':
        query['return_type'] = return_type
    
    # Refund mode filter
    if refund_mode and refund_mode != 'all':
        query['refund_mode'] = refund_mode
    
    # Party filter
    if party_id and party_id != 'all':
        query['party_id'] = party_id
    
    # Get returns
    returns = await db.returns.find(query, {"_id": 0}).to_list(10000)
    
    # Calculate summary metrics
    total_returns_count = 0
    sales_returns_amount = 0.0
    purchase_returns_amount = 0.0
    total_refund_amount = 0.0
    money_refunded = 0.0
    gold_refunded = 0.0
    
    for ret in returns:
        # Apply search filter if provided
        if search:
            search_lower = search.lower()
            party_name = ret.get('party_name', '').lower()
            return_number = ret.get('return_number', '').lower()
            reason = ret.get('reason', '').lower()
            
            if not (search_lower in party_name or 
                    search_lower in return_number or 
                    search_lower in reason):
                continue  # Skip this return if search doesn't match
        
        total_returns_count += 1
        
        # Get total_amount (handle Decimal128)
        total_amt = ret.get('total_amount', 0)
        if isinstance(total_amt, Decimal128):
            total_amt = float(total_amt.to_decimal())
        
        # Categorize by return type
        if ret.get('return_type') == 'sale_return':
            sales_returns_amount += total_amt
        elif ret.get('return_type') == 'purchase_return':
            purchase_returns_amount += total_amt
        
        # Total refund amount (same as total_amount)
        total_refund_amount += total_amt
        
        # Money refunded
        refund_money = ret.get('refund_money_amount', 0)
        if isinstance(refund_money, Decimal128):
            refund_money = float(refund_money.to_decimal())
        money_refunded += refund_money
        
        # Gold refunded
        refund_gold = ret.get('refund_gold_grams', 0)
        if isinstance(refund_gold, Decimal128):
            refund_gold = float(refund_gold.to_decimal())
        gold_refunded += refund_gold
    
    return {
        "total_returns_count": total_returns_count,
        "sales_returns_amount": round(sales_returns_amount, 2),
        "purchase_returns_amount": round(purchase_returns_amount, 2),
        "total_refund_amount": round(total_refund_amount, 2),
        "money_refunded": round(money_refunded, 2),
        "gold_refunded": round(gold_refunded, 3)
    }


@api_router.get("/reports/returns-export")
async def export_returns_report(
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    return_type: Optional[str] = None,
    status: Optional[str] = None,
    refund_mode: Optional[str] = None,
    party_id: Optional[str] = None,
    search: Optional[str] = None,
    current_user: User = Depends(require_permission('reports.view'))
):
    """Export returns report as Excel file with applied filters"""
    from fastapi.responses import StreamingResponse
    from io import BytesIO
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    
    # Build query
    query = {"is_deleted": False}
    
    # Date filters
    if date_from:
        query['date'] = {"$gte": datetime.fromisoformat(date_from)}
    if date_to:
        end_dt = datetime.fromisoformat(date_to)
        if 'date' in query:
            query['date']['$lte'] = end_dt
        else:
            query['date'] = {"$lte": end_dt}
    
    # Return type filter
    if return_type and return_type != 'all':
        query['return_type'] = return_type
    
    # Status filter
    if status and status != 'all':
        query['status'] = status
    
    # Refund mode filter
    if refund_mode and refund_mode != 'all':
        query['refund_mode'] = refund_mode
    
    # Party filter
    if party_id and party_id != 'all':
        query['party_id'] = party_id
    
    # Get returns
    returns = await db.returns.find(query, {"_id": 0}).sort("date", -1).to_list(10000)
    
    # Apply search filter
    if search:
        search_lower = search.lower()
        returns = [
            ret for ret in returns
            if (search_lower in ret.get('party_name', '').lower() or
                search_lower in ret.get('return_number', '').lower() or
                search_lower in ret.get('reason', '').lower())
        ]
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Returns Report"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = [
        "Return #", "Date", "Return Type", "Party Name", "Status",
        "Refund Mode", "Refund Amount (OMR)", "Gold Weight Returned (g)",
        "Linked Invoice/Purchase #", "Payment Mode", "Notes"
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Data rows
    for row_num, ret in enumerate(returns, 2):
        # Format date
        ret_date = ret.get('date', '')
        if isinstance(ret_date, str):
            ret_date = ret_date[:10]
        elif hasattr(ret_date, 'strftime'):
            ret_date = ret_date.strftime('%Y-%m-%d')
        
        # Get refund amount
        refund_amount = ret.get('refund_money_amount', 0)
        if isinstance(refund_amount, Decimal128):
            refund_amount = float(refund_amount.to_decimal())
        
        # Get gold weight
        gold_weight = ret.get('refund_gold_grams', 0)
        if isinstance(gold_weight, Decimal128):
            gold_weight = float(gold_weight.to_decimal())
        
        # Return type display
        return_type_display = "Sales Return" if ret.get('return_type') == 'sale_return' else "Purchase Return"
        
        # Refund mode display
        refund_mode_display = ret.get('refund_mode', '').capitalize()
        
        row_data = [
            ret.get('return_number', ''),
            ret_date,
            return_type_display,
            ret.get('party_name', ''),
            ret.get('status', '').capitalize(),
            refund_mode_display,
            round(refund_amount, 2),
            round(gold_weight, 3),
            ret.get('reference_number', ret.get('reference_id', '')[:8]),
            ret.get('payment_mode', '').replace('_', ' ').title() if ret.get('payment_mode') else '',
            ret.get('notes', '')
        ]
        
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = thin_border
    
    # Adjust column widths
    column_widths = [15, 12, 15, 20, 12, 12, 18, 20, 22, 15, 30]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = width
    
    # Save to BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"returns_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@api_router.get("/reports/returns-pdf")
async def export_returns_pdf(
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    return_type: Optional[str] = None,
    status: Optional[str] = None,
    refund_mode: Optional[str] = None,
    party_id: Optional[str] = None,
    search: Optional[str] = None,
    current_user: User = Depends(require_permission('reports.view'))
):
    """Export returns report as PDF file with applied filters"""
    from fastapi.responses import StreamingResponse
    from io import BytesIO
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    
    # Build query
    query = {"is_deleted": False}
    
    # Date filters
    if date_from:
        query['date'] = {"$gte": datetime.fromisoformat(date_from)}
    if date_to:
        end_dt = datetime.fromisoformat(date_to)
        if 'date' in query:
            query['date']['$lte'] = end_dt
        else:
            query['date'] = {"$lte": end_dt}
    
    # Return type filter
    if return_type and return_type != 'all':
        query['return_type'] = return_type
    
    # Status filter
    if status and status != 'all':
        query['status'] = status
    
    # Refund mode filter
    if refund_mode and refund_mode != 'all':
        query['refund_mode'] = refund_mode
    
    # Party filter
    if party_id and party_id != 'all':
        query['party_id'] = party_id
    
    # Get returns
    returns = await db.returns.find(query, {"_id": 0}).sort("date", -1).to_list(10000)
    
    # Apply search filter
    if search:
        search_lower = search.lower()
        returns = [
            ret for ret in returns
            if (search_lower in ret.get('party_name', '').lower() or
                search_lower in ret.get('return_number', '').lower() or
                search_lower in ret.get('reason', '').lower())
        ]
    
    # Create PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#1f2937'),
        spaceAfter=30,
        alignment=1  # Center
    )
    elements.append(Paragraph("Returns Report", title_style))
    elements.append(Spacer(1, 0.2 * inch))
    
    # Filter info
    filter_info = []
    if date_from:
        filter_info.append(f"From: {date_from}")
    if date_to:
        filter_info.append(f"To: {date_to}")
    if return_type and return_type != 'all':
        filter_info.append(f"Type: {return_type}")
    if status and status != 'all':
        filter_info.append(f"Status: {status}")
    
    if filter_info:
        filter_text = " | ".join(filter_info)
        elements.append(Paragraph(f"<b>Filters:</b> {filter_text}", styles['Normal']))
        elements.append(Spacer(1, 0.2 * inch))
    
    # Table data
    table_data = [[
        "Return #", "Date", "Type", "Party", "Status",
        "Refund Mode", "Amount (OMR)", "Gold (g)"
    ]]
    
    for ret in returns:
        # Format date
        ret_date = ret.get('date', '')
        if isinstance(ret_date, str):
            ret_date = ret_date[:10]
        elif hasattr(ret_date, 'strftime'):
            ret_date = ret_date.strftime('%Y-%m-%d')
        
        # Get refund amount
        refund_amount = ret.get('refund_money_amount', 0)
        if isinstance(refund_amount, Decimal128):
            refund_amount = float(refund_amount.to_decimal())
        
        # Get gold weight
        gold_weight = ret.get('refund_gold_grams', 0)
        if isinstance(gold_weight, Decimal128):
            gold_weight = float(gold_weight.to_decimal())
        
        # Return type display (abbreviated for PDF)
        return_type_display = "Sales" if ret.get('return_type') == 'sale_return' else "Purchase"
        
        table_data.append([
            ret.get('return_number', '')[:10],
            ret_date,
            return_type_display,
            ret.get('party_name', '')[:15],
            ret.get('status', '').capitalize()[:8],
            ret.get('refund_mode', '').capitalize()[:8],
            f"{refund_amount:.2f}",
            f"{gold_weight:.3f}"
        ])
    
    # Create table
    table = Table(table_data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f3f4f6')])
    ]))
    
    elements.append(table)
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    
    filename = f"returns_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )




# Health check endpoint (no authentication required)
@api_router.get("/health")
@limiter.limit("100/minute")  # General rate limit: 100 requests per minute per IP
async def health_check(request: Request):
    """Health check endpoint for monitoring"""
    try:
        # Test database connection
        await db.command('ping')
        return {
            "status": "healthy",
            "service": "Gold Inventory Management System",
            "database": "connected",
            "timestamp": datetime.now(timezone.utc).isoformat()
        }
    except Exception as e:
        raise HTTPException(
            status_code=503,
            detail=f"Service unhealthy: {str(e)}"
        )


# ========================================
# WORKFLOW CONTROL - IMPACT SUMMARY ENDPOINTS
# ========================================

@api_router.get("/jobcards/{jobcard_id}/complete-impact")
async def get_jobcard_complete_impact(jobcard_id: str, current_user: User = Depends(require_permission('jobcards.view'))):
    """Get impact summary before completing a job card"""
    jobcard = await db.jobcards.find_one({"id": jobcard_id, "is_deleted": False})
    if not jobcard:
        raise HTTPException(status_code=404, detail="Job card not found")
    
    items = jobcard.get("items", [])
    total_weight = sum(item.get("weight_in", 0) for item in items)
    # Calculate total making charges based on making_charge_type and making_charge_value
    total_making = 0
    for item in items:
        if item.get('making_charge_value') is not None:
            making_charge_value = float(item.get('making_charge_value', 0))
            if item.get('making_charge_type') == 'per_gram':
                weight = float(item.get('weight_in', 0))
                total_making += making_charge_value * weight
            else:  # 'flat'
                total_making += making_charge_value
    
    return {
        "action": "Complete Job Card",
        "item_count": len(items),
        "total_weight": round(total_weight, 3),
        "total_making_charges": round(total_making, 2),
        "status_change": f"{jobcard.get('status', 'created')} → completed",
        "warning": "This action cannot be undone. The job card will be marked as completed.",
        "can_proceed": jobcard.get("status") in ["created", "in_progress"]
    }

@api_router.get("/jobcards/{jobcard_id}/deliver-impact")
async def get_jobcard_deliver_impact(jobcard_id: str, current_user: User = Depends(require_permission('jobcards.view'))):
    """Get impact summary before delivering a job card"""
    jobcard = await db.jobcards.find_one({"id": jobcard_id, "is_deleted": False})
    if not jobcard:
        raise HTTPException(status_code=404, detail="Job card not found")
    
    items = jobcard.get("items", [])
    total_weight = sum(item.get("weight_in", 0) for item in items)
    
    # Check if job card has been converted to invoice
    has_invoice = jobcard.get("is_invoiced", False)
    invoice_id = jobcard.get("invoice_id", None)
    
    return {
        "action": "Deliver Job Card",
        "item_count": len(items),
        "total_weight": round(total_weight, 3),
        "status_change": f"{jobcard.get('status', 'created')} → delivered",
        "warning": "This action cannot be undone. The job card will be marked as delivered and cannot be edited.",
        "can_proceed": jobcard.get("status") == "completed" and has_invoice,
        "has_invoice": has_invoice,
        "invoice_id": invoice_id,
        "invoice_required": not has_invoice
    }

@api_router.get("/jobcards/{jobcard_id}/delete-impact")
async def get_jobcard_delete_impact(jobcard_id: str, current_user: User = Depends(require_permission('jobcards.view'))):
    """Get impact summary before deleting a job card"""
    jobcard = await db.jobcards.find_one({"id": jobcard_id, "is_deleted": False})
    if not jobcard:
        raise HTTPException(status_code=404, detail="Job card not found")
    
    # Check if linked to an invoice
    linked_invoice = await db.invoices.find_one({
        "jobcard_id": jobcard_id,
        "is_deleted": False
    })
    
    items = jobcard.get("items", [])
    total_weight = sum(item.get("weight_in", 0) for item in items)
    
    can_proceed = linked_invoice is None
    blocking_reason = "Job card is linked to an invoice" if linked_invoice else None
    
    return {
        "action": "Delete Job Card",
        "item_count": len(items),
        "total_weight": round(total_weight, 3),
        "customer_name": jobcard.get("customer_name", "Walk-in Customer"),
        "status": jobcard.get("status", "created"),
        "warning": "This action cannot be undone. The job card will be permanently deleted." if can_proceed else blocking_reason,
        "can_proceed": can_proceed,
        "blocking_reason": blocking_reason,
        "linked_invoice": linked_invoice.get("invoice_number") if linked_invoice else None
    }

@api_router.get("/invoices/{invoice_id}/finalize-impact")
async def get_invoice_finalize_impact(invoice_id: str, current_user: User = Depends(require_permission('invoices.view'))):
    """Get impact summary before finalizing an invoice"""
    invoice = await db.invoices.find_one({"id": invoice_id, "is_deleted": False})
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    
    items = invoice.get("items", [])
    # FIX: Invoice items use "weight" field, not "weight_in"
    # Compute total_weight by summing (item.weight * item.qty) for all items
    total_weight = sum(item.get("weight", 0) * item.get("qty", 1) for item in items)
    
    return {
        "action": "Finalize Invoice",
        "invoice_number": invoice.get("invoice_number", "N/A"),
        "item_count": len(items),
        "total_weight": round(total_weight, 3),
        "grand_total": round(invoice.get("grand_total", 0), 2),
        "status_change": "draft → finalized",
        "warning": "This action cannot be undone. The invoice will be finalized and locked. Stock will be deducted.",
        "can_proceed": invoice.get("status") == "draft",
        "will_deduct_stock": True,
        "will_lock_jobcard": invoice.get("jobcard_id") is not None
    }

@api_router.get("/invoices/{invoice_id}/delete-impact")
async def get_invoice_delete_impact(invoice_id: str, current_user: User = Depends(require_permission('invoices.view'))):
    """Get impact summary before deleting an invoice"""
    invoice = await db.invoices.find_one({"id": invoice_id, "is_deleted": False})
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    
    items = invoice.get("items", [])
    # FIX: Invoice items use "weight" field, not "weight_in"
    # Compute total_weight by summing (item.weight * item.qty) for all items
    total_weight = sum(item.get("weight", 0) * item.get("qty", 1) for item in items)
    is_finalized = invoice.get("status") == "finalized"
    
    # Check for linked payments
    payment_count = await db.transactions.count_documents({
        "reference_type": "invoice",
        "reference_id": invoice_id,
        "is_deleted": False
    })
    
    can_proceed = not is_finalized and payment_count == 0
    blocking_reasons = []
    if is_finalized:
        blocking_reasons.append("Invoice is finalized")
    if payment_count > 0:
        blocking_reasons.append(f"{payment_count} payment(s) linked")
    
    return {
        "action": "Delete Invoice",
        "invoice_number": invoice.get("invoice_number", "N/A"),
        "item_count": len(items),
        "total_weight": round(total_weight, 3),
        "grand_total": round(invoice.get("grand_total", 0), 2),
        "status": invoice.get("status", "draft"),
        "payment_count": payment_count,
        "warning": "This action cannot be undone." if can_proceed else "Cannot delete this invoice.",
        "can_proceed": can_proceed,
        "blocking_reasons": blocking_reasons if blocking_reasons else None
    }

@api_router.get("/purchases/{purchase_id}/finalize-impact")
async def get_purchase_finalize_impact(purchase_id: str, current_user: User = Depends(require_permission('purchases.view'))):
    """Get impact summary before finalizing a purchase"""
    purchase = await db.purchases.find_one({"id": purchase_id, "is_deleted": False})
    if not purchase:
        raise HTTPException(status_code=404, detail="Purchase not found")
    
    vendor = await db.parties.find_one({"id": purchase.get("vendor_party_id")})
    vendor_name = vendor.get("name") if vendor else "Unknown"
    
    return {
        "action": "Finalize Purchase",
        "vendor_name": vendor_name,
        "weight": round(purchase.get("weight_grams", 0), 3),
        "amount_total": round(purchase.get("amount_total", 0), 2),
        "paid_amount": round(purchase.get("paid_amount_money", 0), 2),
        "balance_due": round(purchase.get("balance_due_money", 0), 2),
        "status_change": "draft → finalized",
        "warning": "This action cannot be undone. Stock will be added at 916 purity (22K). Vendor payable will be created.",
        "can_proceed": purchase.get("status") == "draft",
        "will_add_stock": True,
        "will_create_payable": True
    }

@api_router.get("/purchases/{purchase_id}/delete-impact")
async def get_purchase_delete_impact(purchase_id: str, current_user: User = Depends(require_permission('purchases.view'))):
    """Get impact summary before deleting a purchase"""
    purchase = await db.purchases.find_one({"id": purchase_id, "is_deleted": False})
    if not purchase:
        raise HTTPException(status_code=404, detail="Purchase not found")
    
    vendor = await db.parties.find_one({"id": purchase.get("vendor_party_id")})
    vendor_name = vendor.get("name") if vendor else "Unknown"
    is_finalized = purchase.get("status") == "finalized"
    
    can_proceed = not is_finalized
    blocking_reason = "Purchase is finalized and cannot be deleted" if is_finalized else None
    
    return {
        "action": "Delete Purchase",
        "vendor_name": vendor_name,
        "weight": round(purchase.get("weight_grams", 0), 3),
        "amount_total": round(purchase.get("amount_total", 0), 2),
        "status": purchase.get("status", "draft"),
        "warning": "This action cannot be undone." if can_proceed else blocking_reason,
        "can_proceed": can_proceed,
        "blocking_reason": blocking_reason
    }

@api_router.delete("/purchases/{purchase_id}")
async def delete_purchase(purchase_id: str, current_user: User = Depends(require_permission('purchases.delete'))):
    """Delete a purchase (only draft purchases can be deleted)"""
    existing = await db.purchases.find_one({"id": purchase_id, "is_deleted": False})
    if not existing:
        raise HTTPException(status_code=404, detail="Purchase not found")
    
    # CRITICAL: Only allow deleting draft purchases
    # Finalized purchases should not be deleted to maintain financial integrity
    if existing.get("status") == "finalized":
        raise HTTPException(
            status_code=400, 
            detail="Cannot delete finalized purchase. Finalized purchases are immutable to maintain financial integrity."
        )
    
    await db.purchases.update_one(
        {"id": purchase_id},
        {"$set": {"is_deleted": True}}
    )
    await create_audit_log(current_user.id, current_user.full_name, "purchases", purchase_id, "delete")
    return {"message": "Purchase deleted successfully"}

@api_router.get("/parties/{party_id}/delete-impact")
async def get_party_delete_impact(party_id: str, current_user: User = Depends(require_permission('parties.view'))):
    """Get impact summary before deleting a party"""
    party = await db.parties.find_one({"id": party_id, "is_deleted": False})
    if not party:
        raise HTTPException(status_code=404, detail="Party not found")
    
    # Check for linked records
    jobcard_count = await db.jobcards.count_documents({
        "customer_id": party_id,
        "is_deleted": False
    })
    
    invoice_count = await db.invoices.count_documents({
        "customer_id": party_id,
        "is_deleted": False
    })
    
    purchase_count = await db.purchases.count_documents({
        "vendor_party_id": party_id,
        "is_deleted": False
    })
    
    transaction_count = await db.transactions.count_documents({
        "party_id": party_id,
        "is_deleted": False
    })
    
    gold_ledger_count = await db.gold_ledger.count_documents({
        "party_id": party_id,
        "is_deleted": False
    })
    
    total_linked = jobcard_count + invoice_count + purchase_count + transaction_count + gold_ledger_count
    can_proceed = total_linked == 0
    
    blocking_reasons = []
    if jobcard_count > 0:
        blocking_reasons.append(f"{jobcard_count} job card(s)")
    if invoice_count > 0:
        blocking_reasons.append(f"{invoice_count} invoice(s)")
    if purchase_count > 0:
        blocking_reasons.append(f"{purchase_count} purchase(s)")
    if transaction_count > 0:
        blocking_reasons.append(f"{transaction_count} transaction(s)")
    if gold_ledger_count > 0:
        blocking_reasons.append(f"{gold_ledger_count} gold ledger entry/entries")
    
    return {
        "action": "Delete Party",
        "party_name": party.get("name", "Unknown"),
        "party_type": party.get("party_type", "customer"),
        "total_linked_records": total_linked,
        "linked_records": {
            "jobcards": jobcard_count,
            "invoices": invoice_count,
            "purchases": purchase_count,
            "transactions": transaction_count,
            "gold_ledger": gold_ledger_count
        },
        "warning": "This action cannot be undone." if can_proceed else f"Cannot delete party. Linked to: {', '.join(blocking_reasons)}",
        "can_proceed": can_proceed,
        "blocking_reasons": blocking_reasons if blocking_reasons else None
    }

@api_router.get("/transactions/{transaction_id}/delete-impact")
async def get_transaction_delete_impact(transaction_id: str, current_user: User = Depends(require_permission('finance.view'))):
    """Get impact summary before deleting a transaction"""
    transaction = await db.transactions.find_one({"id": transaction_id, "is_deleted": False})
    if not transaction:
        raise HTTPException(status_code=404, detail="Transaction not found")
    
    # Check if linked to invoice
    is_invoice_payment = transaction.get("reference_type") == "invoice"
    can_proceed = not is_invoice_payment
    blocking_reason = "Transaction is linked to an invoice and cannot be deleted" if is_invoice_payment else None
    
    return {
        "action": "Delete Transaction",
        "transaction_number": transaction.get("transaction_number", "N/A"),
        "transaction_type": transaction.get("transaction_type", "N/A"),
        "amount": round(transaction.get("amount", 0), 2),
        "category": transaction.get("category", "N/A"),
        "party_name": transaction.get("party_name", "N/A"),
        "is_invoice_payment": is_invoice_payment,
        "warning": "This action cannot be undone. Account balance will be adjusted." if can_proceed else blocking_reason,
        "can_proceed": can_proceed,
        "blocking_reason": blocking_reason,
        "will_affect_account": True
    }


@api_router.delete("/transactions/{transaction_id}")
@limiter.limit("1000/hour")
async def delete_transaction(
    request: Request,
    transaction_id: str,
    current_user: User = Depends(require_permission('finance.delete'))
):
    """
    Delete a transaction and reverse its account balance impact.
    
    ACCOUNTING RULES:
    - Can only delete transactions NOT linked to invoices (manual transactions)
    - Invoice payment transactions should not be deleted directly
    - Reverses the account balance change when deleted
    - Soft delete with audit trail
    """
    # Fetch transaction
    transaction = await db.transactions.find_one({"id": transaction_id, "is_deleted": False})
    if not transaction:
        raise HTTPException(status_code=404, detail="Transaction not found")
    
    # Check if linked to invoice - BLOCK deletion
    if transaction.get("reference_type") == "invoice":
        raise HTTPException(
            status_code=400,
            detail="Cannot delete invoice payment transactions. Delete the invoice or payment record instead."
        )
    
    # Get transaction details for balance reversal
    account_id = transaction.get('account_id')
    amount = transaction.get('amount', 0)
    transaction_type = transaction.get('transaction_type', 'debit')
    
    # Calculate balance reversal using CORRECT accounting rules
    # Must fetch account to get its type, then reverse using opposite transaction type
    balance_delta = 0
    if account_id:
        account = await db.accounts.find_one({"id": account_id, "is_deleted": False})
        if account:
            account_type = account.get('account_type', 'asset')
            
            # To reverse a transaction, apply opposite transaction type
            # If original was DEBIT, reverse with CREDIT calculation
            # If original was CREDIT, reverse with DEBIT calculation
            reverse_type = 'credit' if transaction_type == 'debit' else 'debit'
            balance_delta = calculate_balance_delta(account_type, reverse_type, amount)
    
    # Soft delete transaction
    await db.transactions.update_one(
        {"id": transaction_id},
        {
            "$set": {
                "is_deleted": True,
                "deleted_at": datetime.now(timezone.utc),
                "deleted_by": current_user.id
            }
        }
    )
    
    # Reverse account balance
    if account_id and balance_delta != 0:
        await db.accounts.update_one(
            {"id": account_id},
            {"$inc": {"current_balance": balance_delta}}
        )
    
    # Create audit log
    await create_audit_log(
        current_user.id,
        current_user.full_name,
        "transaction",
        transaction_id,
        "delete",
        {
            "transaction_number": transaction.get('transaction_number'),
            "amount": amount,
            "transaction_type": transaction_type,
            "account_id": account_id,
            "balance_reversed": balance_delta
        }
    )
    
    return {
        "message": "Transaction deleted successfully",
        "transaction_id": transaction_id,
        "account_balance_adjusted": bool(account_id),
        "balance_change": balance_delta
    }




# ============================================================================
# RETURNS MANAGEMENT API ENDPOINTS
# ============================================================================

@api_router.post("/returns", status_code=201)
@limiter.limit("1000/hour")
async def create_return(
    request: Request,
    return_data: dict,
    current_user: User = Depends(require_permission('returns.create'))
):
    """
    Create a new return (draft status).
    Supports both sales returns and purchase returns.
    """
    try:
        # Validate return_type
        return_type = return_data.get('return_type')
        if return_type not in ['sale_return', 'purchase_return']:
            raise HTTPException(status_code=400, detail="Invalid return_type. Must be 'sale_return' or 'purchase_return'")
        
        # Validate reference
        reference_type = return_data.get('reference_type')
        reference_id = return_data.get('reference_id')
        
        if not reference_type or not reference_id:
            raise HTTPException(status_code=400, detail="reference_type and reference_id are required")
        
        # Fetch the reference document (invoice or purchase)
        if reference_type == 'invoice':
            reference_doc = await db.invoices.find_one({"id": reference_id, "is_deleted": False})
            if not reference_doc:
                raise HTTPException(status_code=404, detail="Invoice not found")
            if reference_doc.get('status') != 'finalized':
                raise HTTPException(status_code=400, detail="Can only create return for finalized invoices")
            party_id = reference_doc.get('customer_id')
            party_name = reference_doc.get('customer_name')
            party_type = 'customer'
            reference_number = reference_doc.get('invoice_number')
        elif reference_type == 'purchase':
            reference_doc = await db.purchases.find_one({"id": reference_id, "is_deleted": False})
            if not reference_doc:
                raise HTTPException(status_code=404, detail="Purchase not found")
            if reference_doc.get('status') != 'finalized':
                raise HTTPException(status_code=400, detail="Can only create return for finalized purchases")
            party_id = reference_doc.get('vendor_party_id')
            # Fetch vendor name
            vendor = await db.parties.find_one({"id": party_id, "is_deleted": False})
            party_name = vendor.get('name') if vendor else 'Unknown Vendor'
            party_type = 'vendor'
            reference_number = reference_id  # Purchases don't have human-readable numbers yet
        else:
            raise HTTPException(status_code=400, detail="Invalid reference_type. Must be 'invoice' or 'purchase'")
        
        # Validate refund_mode
        refund_mode = return_data.get('refund_mode', 'money')
        if refund_mode not in ['money', 'gold', 'mixed']:
            raise HTTPException(status_code=400, detail="Invalid refund_mode. Must be 'money', 'gold', or 'mixed'")
        
        # Validate refund amounts
        refund_money_amount = float(return_data.get('refund_money_amount', 0))
        refund_gold_grams = float(return_data.get('refund_gold_grams', 0))
        
        if refund_mode == 'money' and refund_money_amount <= 0:
            raise HTTPException(status_code=400, detail="refund_money_amount must be greater than 0 for money refund")
        if refund_mode == 'gold' and refund_gold_grams <= 0:
            raise HTTPException(status_code=400, detail="refund_gold_grams must be greater than 0 for gold refund")
        if refund_mode == 'mixed' and (refund_money_amount <= 0 or refund_gold_grams <= 0):
            raise HTTPException(status_code=400, detail="Both refund_money_amount and refund_gold_grams must be greater than 0 for mixed refund")
        
        # Validate account if money refund
        if refund_mode in ['money', 'mixed']:
            account_id = return_data.get('account_id')
            if not account_id:
                raise HTTPException(status_code=400, detail="account_id is required for money refund")
            account = await db.accounts.find_one({"id": account_id, "is_deleted": False})
            if not account:
                raise HTTPException(status_code=404, detail="Account not found")
            account_name = account.get('name')
        else:
            account_name = None
        
        # Generate return number
        returns_count = await db.returns.count_documents({})
        return_number = f"RET-{returns_count + 1:05d}"
        
        # Calculate total weight and amount from items
        items = return_data.get('items', [])
        if not items:
            raise HTTPException(status_code=400, detail="At least one item is required")
        
        total_weight_grams = sum(float(item.get('weight_grams', 0)) for item in items)
        total_amount = sum(float(item.get('amount', 0)) for item in items)
        
        # Validate return amount doesn't exceed original invoice/purchase total
        await validate_return_against_original(
            db=db,
            reference_type=reference_type,
            reference_id=reference_id,
            reference_doc=reference_doc,
            return_items=items,
            current_return_id=None  # New return, no existing ID to exclude
        )
        
        # Create return object
        return_obj = Return(
            return_number=return_number,
            return_type=return_type,
            reference_type=reference_type,
            reference_id=reference_id,
            reference_number=reference_number,
            party_id=party_id,
            party_name=party_name,
            party_type=party_type,
            items=[ReturnItem(**item) for item in items],
            total_weight_grams=round(total_weight_grams, 3),
            total_amount=round(total_amount, 2),
            reason=return_data.get('reason'),
            refund_mode=refund_mode,
            refund_money_amount=round(refund_money_amount, 2),
            refund_gold_grams=round(refund_gold_grams, 3),
            refund_gold_purity=return_data.get('refund_gold_purity'),
            payment_mode=return_data.get('payment_mode'),
            account_id=return_data.get('account_id'),
            account_name=account_name,
            status='draft',
            created_by=current_user.id,
            notes=return_data.get('notes')
        )
        
        # Convert float values to Decimal128 for precise storage
        return_dict = return_obj.model_dump()
        return_dict = convert_return_to_decimal(return_dict)
        
        # Insert into database
        await db.returns.insert_one(return_dict)
        
        # Create audit log
        await create_audit_log(
            user_id=current_user.id,
            user_name=current_user.name,
            module="returns",
            record_id=return_obj.id,
            action="create",
            changes={"status": "draft", "return_type": return_type, "party": party_name}
        )
        
        return {"message": "Return created successfully", "return": decimal_to_float(return_obj.model_dump())}
    
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error creating return: {str(e)}")


@api_router.get("/returns")
@limiter.limit("1000/hour")
async def get_returns(
    request: Request,
    page: int = 1,
    page_size: int = 10,
    return_type: Optional[str] = None,
    party_id: Optional[str] = None,
    status: Optional[str] = None,
    refund_mode: Optional[str] = None,
    search: Optional[str] = None,
    current_user: User = Depends(require_permission('returns.view'))
):
    """
    Get all returns with pagination and filters.
    Filters: return_type, party_id, status, refund_mode, search
    """
    try:
        # Build query
        query = {"is_deleted": False}
        
        if return_type:
            query["return_type"] = return_type
        if party_id:
            query["party_id"] = party_id
        if status:
            query["status"] = status
        if refund_mode:
            query["refund_mode"] = refund_mode
        if search:
            query["$or"] = [
                {"return_number": {"$regex": search, "$options": "i"}},
                {"party_name": {"$regex": search, "$options": "i"}},
                {"reference_number": {"$regex": search, "$options": "i"}},
                {"reason": {"$regex": search, "$options": "i"}}
            ]
        
        # Count total
        total_count = await db.returns.count_documents(query)
        
        # Calculate pagination
        skip = (page - 1) * page_size
        total_pages = (total_count + page_size - 1) // page_size
        
        # Fetch returns
        cursor = db.returns.find(query).sort("created_at", -1).skip(skip).limit(page_size)
        returns = await cursor.to_list(length=page_size)
        
        return {
            "items": [decimal_to_float(ret) for ret in returns],
            "pagination": {
                "total_count": total_count,
                "page": page,
                "page_size": page_size,
                "total_pages": total_pages,
                "has_next": page < total_pages,
                "has_prev": page > 1
            }
        }
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching returns: {str(e)}")


@api_router.get("/returns/{return_id}")
@limiter.limit("1000/hour")
async def get_return_by_id(
    request: Request,
    return_id: str,
    current_user: User = Depends(require_permission('returns.view'))
):
    """Get a single return by ID"""
    return_doc = await db.returns.find_one({"id": return_id, "is_deleted": False})
    if not return_doc:
        raise HTTPException(status_code=404, detail="Return not found")
    
    return decimal_to_float(return_doc)


@api_router.patch("/returns/{return_id}")
@limiter.limit("1000/hour")
async def update_return(
    request: Request,
    return_id: str,
    return_data: dict,
    current_user: User = Depends(require_permission('returns.create'))
):
    """
    Update a return (only allowed in draft status).
    """
    try:
        # Fetch existing return
        existing_return = await db.returns.find_one({"id": return_id, "is_deleted": False})
        if not existing_return:
            raise HTTPException(status_code=404, detail="Return not found")
        
        # Check if finalized
        if existing_return.get('status') == 'finalized':
            raise HTTPException(status_code=400, detail="Cannot update finalized return")
        
        # Prepare update data
        update_fields = {}
        
        # Update items if provided
        if 'items' in return_data:
            items = [ReturnItem(**item) for item in return_data['items']]
            update_fields['items'] = [item.model_dump() for item in items]
            # Recalculate totals
            total_weight = sum(item.weight_grams for item in items)
            total_amount = sum(item.amount for item in items)
            update_fields['total_weight_grams'] = round(total_weight, 3)
            update_fields['total_amount'] = round(total_amount, 2)
            
            # Validate updated return amount doesn't exceed original
            reference_type = existing_return.get('reference_type')
            reference_id = existing_return.get('reference_id')
            
            # Fetch reference document
            if reference_type == 'invoice':
                reference_doc = await db.invoices.find_one({"id": reference_id, "is_deleted": False})
            else:  # purchase
                reference_doc = await db.purchases.find_one({"id": reference_id, "is_deleted": False})
            
            if reference_doc:
                await validate_return_against_original(
                    db=db,
                    reference_type=reference_type,
                    reference_id=reference_id,
                    reference_doc=reference_doc,
                    return_items=return_data['items'],
                    current_return_id=return_id  # Exclude current return from calculation
                )
        
        # Update other fields
        if 'reason' in return_data:
            update_fields['reason'] = return_data['reason']
        if 'refund_mode' in return_data:
            update_fields['refund_mode'] = return_data['refund_mode']
        if 'refund_money_amount' in return_data:
            update_fields['refund_money_amount'] = round(float(return_data['refund_money_amount']), 2)
        if 'refund_gold_grams' in return_data:
            update_fields['refund_gold_grams'] = round(float(return_data['refund_gold_grams']), 3)
        if 'refund_gold_purity' in return_data:
            update_fields['refund_gold_purity'] = return_data['refund_gold_purity']
        if 'payment_mode' in return_data:
            update_fields['payment_mode'] = return_data['payment_mode']
        if 'account_id' in return_data:
            account_id = return_data['account_id']
            account = await db.accounts.find_one({"id": account_id, "is_deleted": False})
            if account:
                update_fields['account_id'] = account_id
                update_fields['account_name'] = account.get('name')
        if 'notes' in return_data:
            update_fields['notes'] = return_data['notes']
        
        # Convert float values to Decimal128 for precise storage
        update_fields = convert_return_to_decimal(update_fields)
        
        # Update database
        await db.returns.update_one(
            {"id": return_id},
            {"$set": update_fields}
        )
        
        # Create audit log
        await create_audit_log(
            user_id=current_user.id,
            user_name=current_user.name,
            module="returns",
            record_id=return_id,
            action="update",
            changes=update_fields
        )
        
        # Fetch updated return
        updated_return = await db.returns.find_one({"id": return_id})
        
        return {"message": "Return updated successfully", "return": decimal_to_float(updated_return)}
    
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error updating return: {str(e)}")


@api_router.post("/returns/{return_id}/finalize")
@limiter.limit("30/minute")
async def finalize_return(
    request: Request,
    return_id: str,
    current_user: User = Depends(require_permission('returns.finalize'))
):
    """
    Finalize a return (apply stock movements, refunds, and update balances).
    Uses MongoDB multi-document transactions for true atomicity.
    All operations either succeed together or fail together with automatic rollback.
    """
    # Pre-transaction validation
    return_doc = await db.returns.find_one({"id": return_id, "is_deleted": False})
    if not return_doc:
        raise HTTPException(status_code=404, detail="Return not found")
    
    current_status = return_doc.get('status')
    
    if current_status == 'finalized':
        raise HTTPException(status_code=400, detail="Return is already finalized")
    
    if current_status == 'processing':
        raise HTTPException(status_code=409, detail="Return is currently being processed. Please try again in a moment.")
    
    # Use status lock + rollback for safety (MongoDB transactions require replica set)
    try:
        # Step 1: Atomic lock - Set status to 'processing'
        lock_result = await db.returns.update_one(
            {"id": return_id, "status": "draft", "is_deleted": False},
            {"$set": {"status": "processing", "processing_started_at": datetime.now(timezone.utc)}}
        )
                
        if lock_result.modified_count == 0:
            raise HTTPException(status_code=409, detail="Return is already being processed or was modified.")
        
        # Get return data with Decimal handling
        return_type = return_doc.get('return_type')
        reference_type = return_doc.get('reference_type')
        reference_id = return_doc.get('reference_id')
        party_id = return_doc.get('party_id')
        refund_mode = return_doc.get('refund_mode')
        
        # Convert Decimal128 to Decimal for arithmetic
        refund_money_amount = return_doc.get('refund_money_amount', 0)
        if isinstance(refund_money_amount, Decimal128):
            refund_money_amount = float(refund_money_amount.to_decimal())
        else:
            refund_money_amount = float(refund_money_amount)
        
        refund_gold_grams = return_doc.get('refund_gold_grams', 0)
        if isinstance(refund_gold_grams, Decimal128):
            refund_gold_grams = float(refund_gold_grams.to_decimal())
        else:
            refund_gold_grams = float(refund_gold_grams)
        
        stock_movement_ids = []
        transaction_id = None
        gold_ledger_id = None
        
        # ========================================================================
        # SALES RETURN WORKFLOW
        # ========================================================================
        if return_type == 'sale_return':
            # 1. Create stock movements (Stock IN - returned goods back to inventory)
            for item in return_doc.get('items', []):
                weight_grams = item.get('weight_grams', 0)
                # Handle Decimal128
                if isinstance(weight_grams, Decimal128):
                    weight_grams = float(weight_grams.to_decimal())
                
                if weight_grams > 0:
                    movement_id = str(uuid.uuid4())
                    stock_movement = StockMovement(
                        id=movement_id,
                        type="IN",
                        category=item.get('description'),
                        weight=round(weight_grams, 3),
                        purity=item.get('purity'),
                        date=datetime.now(timezone.utc),
                        purpose=f"Sales Return - {return_doc.get('return_number')}",
                        reference_type="return",
                        reference_id=return_id,
                        notes=return_doc.get('reason'),
                        created_by=current_user.id
                    )
                    await db.stock_movements.insert_one(stock_movement.model_dump())
                    stock_movement_ids.append(movement_id)
                    
                    # Update inventory header stock
                    await db.inventory_headers.update_one(
                        {"name": item.get('description')},
                        {
                            "$inc": {
                                "current_qty": item.get('qty', 0),
                                "current_weight": round(weight_grams, 3)
                            }
                        }
                    )
            
            # 2. Create money refund transactions
            if refund_mode in ['money', 'mixed'] and refund_money_amount > 0:
                account_id = return_doc.get('account_id')
                account = await db.accounts.find_one({"id": account_id})
                if not account:
                    raise HTTPException(status_code=400, detail="Account not found for money refund")
            
            # 2a. Transaction 1: Debit Cash/Bank account (money going out)
            transactions_count = await db.transactions.count_documents({})
            transaction_number = f"TXN-{transactions_count + 1:05d}"
            
            transaction_id = str(uuid.uuid4())
            transaction = Transaction(
                id=transaction_id,
                transaction_number=transaction_number,
                date=datetime.now(timezone.utc),
                transaction_type="debit",  # Money going out to customer
                mode=return_doc.get('payment_mode', 'cash'),
                account_id=account_id,
                account_name=account.get('name'),
                party_id=party_id,
                party_name=return_doc.get('party_name'),
                amount=round(refund_money_amount, 2),
                category="sales_return",
                notes=f"Sales Return Refund - {return_doc.get('return_number')}",
                reference_type="return",
                reference_id=return_id,
                created_by=current_user.id
            )
            await db.transactions.insert_one(transaction.model_dump())
            
            # Update Cash/Bank account balance (debit = decrease balance for asset accounts)
            await db.accounts.update_one(
                {"id": account_id},
                {"$inc": {"current_balance": -round(refund_money_amount, 2)}}
            )
            
            # 2b. Transaction 2: Debit Sales Income account (reduce revenue)
            # Find Sales Income account
            sales_income_account = await db.accounts.find_one({
                "is_deleted": False,
                "account_type": "income",
                "$or": [
                    {"name": {"$regex": "sales income", "$options": "i"}},
                    {"name": {"$regex": "^sales$", "$options": "i"}}
                ]
            })
            
            if sales_income_account:
                transactions_count = await db.transactions.count_documents({})
                income_transaction_number = f"TXN-{transactions_count + 1:05d}"
                income_transaction_id = str(uuid.uuid4())
                
                income_transaction = Transaction(
                    id=income_transaction_id,
                    transaction_number=income_transaction_number,
                    date=datetime.now(timezone.utc),
                    transaction_type="debit",  # Debit income = reduce revenue
                    mode="adjustment",
                    account_id=sales_income_account.get('id'),
                    account_name=sales_income_account.get('name'),
                    party_id=party_id,
                    party_name=return_doc.get('party_name'),
                    amount=round(refund_money_amount, 2),
                    category="sales_return",
                    notes=f"Sales Return Revenue Adjustment - {return_doc.get('return_number')}",
                    reference_type="return",
                    reference_id=return_id,
                    created_by=current_user.id
                )
                await db.transactions.insert_one(income_transaction.model_dump())
                
                # Update Sales Income account balance (debit income = decrease balance)
                # For income accounts: credits increase (+), debits decrease (-)
                await db.accounts.update_one(
                    {"id": sales_income_account.get('id')},
                    {"$inc": {"current_balance": -round(refund_money_amount, 2)}}
                )
        
        # 3. Create gold refund (GoldLedgerEntry - OUT)
        if refund_mode in ['gold', 'mixed'] and refund_gold_grams > 0:
            gold_ledger_id = str(uuid.uuid4())
            gold_entry = GoldLedgerEntry(
                id=gold_ledger_id,
                party_id=party_id,
                date=datetime.now(timezone.utc),
                type="OUT",  # Shop gives gold to customer
                weight_grams=round(refund_gold_grams, 3),
                purity_entered=return_doc.get('refund_gold_purity', 916),
                purpose="sales_return",
                reference_type="return",
                reference_id=return_id,
                notes=f"Sales Return Gold Refund - {return_doc.get('return_number')}",
                created_by=current_user.id
            )
            await db.gold_ledger.insert_one(gold_entry.model_dump())
        
        # 4. Update invoice (adjust paid_amount and balance_due)
        if reference_type == 'invoice':
            invoice = await db.invoices.find_one({"id": reference_id})
            if invoice:
                # Reduce paid amount by refund amount (as we're returning money)
                new_paid = invoice.get('paid_amount', 0) - refund_money_amount
                new_balance = invoice.get('grand_total', 0) - new_paid
                
                await db.invoices.update_one(
                    {"id": reference_id},
                    {
                        "$set": {
                            "paid_amount": round(max(0, new_paid), 2),
                            "balance_due": round(max(0, new_balance), 2),
                            "payment_status": "unpaid" if new_balance > 0 else "paid"
                        }
                    }
                )
        
        # 5. Update customer outstanding (if saved customer)
        if party_id:
            party = await db.parties.find_one({"id": party_id})
            if party and party.get('party_type') == 'customer':
                # Increase outstanding (customer owes less due to refund)
                current_outstanding = party.get('outstanding_balance', 0)
                # For sales return, we're reducing what customer paid, so increase outstanding
                new_outstanding = current_outstanding + refund_money_amount
                await db.parties.update_one(
                    {"id": party_id},
                    {"$set": {"outstanding_balance": round(new_outstanding, 2)}}
                )
        
        # ========================================================================
        # PURCHASE RETURN WORKFLOW
        # ========================================================================
        elif return_type == 'purchase_return':
            # 1. Create stock movements (Stock OUT - returned to vendor)
            for item in return_doc.get('items', []):
                if item.get('weight_grams', 0) > 0:
                    movement_id = str(uuid.uuid4())
                    stock_movement = StockMovement(
                        id=movement_id,
                        type="OUT",
                        category=item.get('description'),
                        weight=round(item.get('weight_grams'), 3),
                        purity=item.get('purity'),
                        date=datetime.now(timezone.utc),
                        purpose=f"Purchase Return - {return_doc.get('return_number')}",
                        reference_type="return",
                        reference_id=return_id,
                        notes=return_doc.get('reason'),
                        created_by=current_user.id
                    )
                    await db.stock_movements.insert_one(stock_movement.model_dump())
                    stock_movement_ids.append(movement_id)
                    
                    # Update inventory header stock (decrease)
                    await db.inventory_headers.update_one(
                        {"name": item.get('description')},
                        {
                            "$inc": {
                                "current_qty": -item.get('qty', 0),
                                "current_weight": -round(item.get('weight_grams'), 3)
                            }
                        }
                    )
            
            # 2. Create money refund (Transaction - Credit - vendor refunds us)
            if refund_mode in ['money', 'mixed'] and refund_money_amount > 0:
                account_id = return_doc.get('account_id')
                account = await db.accounts.find_one({"id": account_id})
                if not account:
                    raise HTTPException(status_code=400, detail="Account not found for money refund")
                
                # Generate transaction number
                transactions_count = await db.transactions.count_documents({})
                transaction_number = f"TXN-{transactions_count + 1:05d}"
                
                transaction_id = str(uuid.uuid4())
                transaction = Transaction(
                    id=transaction_id,
                    transaction_number=transaction_number,
                    date=datetime.now(timezone.utc),
                    transaction_type="credit",  # Money coming in from vendor
                    mode=return_doc.get('payment_mode', 'cash'),
                    account_id=account_id,
                    account_name=account.get('name'),
                    party_id=party_id,
                    party_name=return_doc.get('party_name'),
                    amount=round(refund_money_amount, 2),
                    category="purchase_return",
                    notes=f"Purchase Return Refund - {return_doc.get('return_number')}",
                    reference_type="return",
                    reference_id=return_id,
                    created_by=current_user.id
                )
                await db.transactions.insert_one(transaction.model_dump())
                
                # Update account balance (credit = increase balance)
                await db.accounts.update_one(
                    {"id": account_id},
                    {"$inc": {"current_balance": round(refund_money_amount, 2)}}
                )
            
            # 3. Create gold refund (GoldLedgerEntry - IN - vendor returns gold to us)
            if refund_mode in ['gold', 'mixed'] and refund_gold_grams > 0:
                gold_ledger_id = str(uuid.uuid4())
                gold_entry = GoldLedgerEntry(
                    id=gold_ledger_id,
                    party_id=party_id,
                    date=datetime.now(timezone.utc),
                    type="IN",  # Vendor gives gold back to shop
                    weight_grams=round(refund_gold_grams, 3),
                    purity_entered=return_doc.get('refund_gold_purity', 916),
                    purpose="purchase_return",
                    reference_type="return",
                    reference_id=return_id,
                    notes=f"Purchase Return Gold Refund - {return_doc.get('return_number')}",
                    created_by=current_user.id
                )
                await db.gold_ledger.insert_one(gold_entry.model_dump())
            
            # 4. Update purchase (adjust balance_due_money)
            if reference_type == 'purchase':
                purchase = await db.purchases.find_one({"id": reference_id})
                if purchase:
                    # Reduce balance due by refund amount
                    new_balance = purchase.get('balance_due_money', 0) - refund_money_amount
                    
                    await db.purchases.update_one(
                        {"id": reference_id},
                        {"$set": {"balance_due_money": round(max(0, new_balance), 2)}}
                    )
            
            # 5. Update vendor payable
            if party_id:
                party = await db.parties.find_one({"id": party_id})
                if party and party.get('party_type') == 'vendor':
                    # Decrease outstanding (we owe vendor less due to return)
                    current_outstanding = party.get('outstanding_balance', 0)
                    new_outstanding = current_outstanding - refund_money_amount
                    await db.parties.update_one(
                        {"id": party_id},
                        {"$set": {"outstanding_balance": round(new_outstanding, 2)}}
                    )
        
        # ========================================================================
        # UPDATE RETURN STATUS TO FINALIZED
        # ========================================================================
        await db.returns.update_one(
            {"id": return_id},
            {
                "$set": {
                    "status": "finalized",
                    "finalized_at": datetime.now(timezone.utc),
                    "finalized_by": current_user.id,
                    "stock_movement_ids": stock_movement_ids,
                    "transaction_id": transaction_id,
                    "gold_ledger_id": gold_ledger_id
                }
            }
        )
        
        # Create audit log
        await create_audit_log(
            user_id=current_user.id,
            user_name=current_user.name,
            module="returns",
            record_id=return_id,
            action="finalize",
            changes={
                "status": "finalized",
                "stock_movements_created": len(stock_movement_ids),
                "transaction_created": transaction_id is not None,
                "gold_ledger_created": gold_ledger_id is not None
            }
        )
        
        # Fetch updated return
        updated_return = await db.returns.find_one({"id": return_id})
        
        return {
            "message": "Return finalized successfully",
            "return": decimal_to_float(updated_return),
            "details": {
                "stock_movements_created": len(stock_movement_ids),
                "transaction_created": transaction_id is not None,
                "gold_ledger_created": gold_ledger_id is not None
            }
        }
    
    except HTTPException as he:
        # Rollback: Reset status to draft if processing lock was acquired
        try:
            await db.returns.update_one(
                {"id": return_id, "status": "processing"},
                {"$set": {"status": "draft"}, "$unset": {"processing_started_at": ""}}
            )
        except:
            pass  # Best effort rollback
        raise he
    except Exception as e:
        # CRITICAL ROLLBACK - Finalization failed mid-process, must revert all changes
        try:
            # 1. Rollback return status to draft
            await db.returns.update_one(
                {"id": return_id},
                {
                    "$set": {"status": "draft"},
                    "$unset": {
                        "processing_started_at": "",
                        "stock_movement_ids": "",
                        "transaction_id": "",
                        "gold_ledger_id": ""
                    }
                }
            )
            
            # 2. Delete any stock movements created
            if stock_movement_ids:
                for movement_id in stock_movement_ids:
                    await db.stock_movements.delete_one({"id": movement_id})
            
            # 3. Delete transaction if created
            if transaction_id:
                transaction = await db.transactions.find_one({"id": transaction_id})
                if transaction:
                    # Revert account balance
                    account_id = transaction.get('account_id')
                    amount = transaction.get('amount', 0)
                    transaction_type = transaction.get('transaction_type')
                    if account_id:
                        # Reverse the balance change
                        balance_change = amount if transaction_type == 'debit' else -amount
                        await db.accounts.update_one(
                            {"id": account_id},
                            {"$inc": {"current_balance": balance_change}}
                        )
                    # Delete transaction
                    await db.transactions.delete_one({"id": transaction_id})
            
            # 4. Delete gold ledger entry if created
            if gold_ledger_id:
                await db.gold_ledger.delete_one({"id": gold_ledger_id})
            
            # 5. Revert inventory header changes
            return_doc = await db.returns.find_one({"id": return_id})
            if return_doc:
                return_type = return_doc.get('return_type')
                for item in return_doc.get('items', []):
                    if item.get('weight_grams', 0) > 0:
                        # Reverse the inventory change
                        qty_change = item.get('qty', 0)
                        weight_change = round(item.get('weight_grams'), 3)
                        
                        if return_type == 'sale_return':
                            # We added stock, now subtract it back
                            await db.inventory_headers.update_one(
                                {"name": item.get('description')},
                                {
                                    "$inc": {
                                        "current_qty": -qty_change,
                                        "current_weight": -weight_change
                                    }
                                }
                            )
                        else:  # purchase_return
                            # We removed stock, now add it back
                            await db.inventory_headers.update_one(
                                {"name": item.get('description')},
                                {
                                    "$inc": {
                                        "current_qty": qty_change,
                                        "current_weight": weight_change
                                    }
                                }
                            )
            
            # 6. Create audit log for rollback
            await create_audit_log(
                user_id=current_user.id,
                user_name=current_user.name,
                module="returns",
                record_id=return_id,
                action="finalize_rollback",
                changes={
                    "error": str(e),
                    "rollback_completed": True,
                    "stock_movements_deleted": len(stock_movement_ids),
                    "transaction_deleted": transaction_id is not None,
                    "gold_ledger_deleted": gold_ledger_id is not None
                }
            )
        except Exception as rollback_error:
            # Even rollback failed - log critical error
            print(f"CRITICAL: Rollback failed for return {return_id}: {str(rollback_error)}")
        
        raise HTTPException(
            status_code=500,
            detail=f"Error finalizing return: {str(e)}. Changes have been rolled back."
        )


@api_router.delete("/returns/{return_id}")
@limiter.limit("30/minute")
async def delete_return(
    request: Request,
    return_id: str,
    current_user: User = Depends(require_permission('returns.delete'))
):
    """
    Soft delete a return (only allowed in draft status).
    """
    try:
        # Fetch return
        return_doc = await db.returns.find_one({"id": return_id, "is_deleted": False})
        if not return_doc:
            raise HTTPException(status_code=404, detail="Return not found")
        
        # Check if finalized
        if return_doc.get('status') == 'finalized':
            raise HTTPException(status_code=400, detail="Cannot delete finalized return. Finalized returns are immutable.")
        
        # Soft delete
        await db.returns.update_one(
            {"id": return_id},
            {
                "$set": {
                    "is_deleted": True,
                    "deleted_at": datetime.now(timezone.utc),
                    "deleted_by": current_user.id
                }
            }
        )
        
        # Create audit log
        await create_audit_log(
            user_id=current_user.id,
            user_name=current_user.name,
            module="returns",
            record_id=return_id,
            action="delete",
            changes={"is_deleted": True}
        )
        
        return {"message": "Return deleted successfully"}
    
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error deleting return: {str(e)}")


@api_router.get("/returns/{return_id}/finalize-impact")
@limiter.limit("1000/hour")
async def get_return_finalize_impact(
    request: Request,
    return_id: str,
    current_user: User = Depends(require_permission('returns.view'))
):
    """Get impact summary before finalizing a return"""
    return_doc = await db.returns.find_one({"id": return_id, "is_deleted": False})
    if not return_doc:
        raise HTTPException(status_code=404, detail="Return not found")
    
    if return_doc.get('status') == 'finalized':
        return {
            "can_proceed": False,
            "message": "Return is already finalized"
        }
    
    return_type = return_doc.get('return_type')
    refund_mode = return_doc.get('refund_mode')
    
    impact = {
        "action": "Finalize Return",
        "return_number": return_doc.get('return_number'),
        "return_type": return_type,
        "party_name": return_doc.get('party_name'),
        "refund_mode": refund_mode,
        "can_proceed": True
    }
    
    # Add specific impacts based on return type
    if return_type == 'sale_return':
        impact["impacts"] = [
            f"✅ Stock IN: {return_doc.get('total_weight_grams', 0):.3f}g returned to inventory",
            f"💰 Customer Refund: {return_doc.get('refund_money_amount', 0):.2f} OMR" if refund_mode in ['money', 'mixed'] else None,
            f"🪙 Gold Refund: {return_doc.get('refund_gold_grams', 0):.3f}g to customer" if refund_mode in ['gold', 'mixed'] else None,
            "📊 Customer outstanding will be updated"
        ]
    else:  # purchase_return
        impact["impacts"] = [
            f"❌ Stock OUT: {return_doc.get('total_weight_grams', 0):.3f}g returned to vendor",
            f"💰 Vendor Refund: {return_doc.get('refund_money_amount', 0):.2f} OMR received" if refund_mode in ['money', 'mixed'] else None,
            f"🪙 Gold Refund: {return_doc.get('refund_gold_grams', 0):.3f}g from vendor" if refund_mode in ['gold', 'mixed'] else None,
            "📊 Vendor payable will be updated"
        ]
    
    # Filter out None values
    impact["impacts"] = [i for i in impact["impacts"] if i]
    
    impact["warning"] = "This action cannot be undone. All stock movements, transactions, and ledger entries will be created."
    
    return impact



app.include_router(api_router)



# 1. HTTPS Redirect (Innermost)
# app.add_middleware(HTTPSRedirectMiddleware)  <-- Comment this out

# 2. Security Headers
# app.add_middleware(SecurityHeadersMiddleware)

# 3. Input Sanitization
# temporarily disabling the Input Sanitization middleware
# app.add_middleware(InputSanitizationMiddleware)

# 4. CSRF Protection
# (You can comment this out if you still have issues, but moving it 'above' CORS usually fixes it)
# app.add_middleware(CSRFProtectionMiddleware)

# 5. CORS Middleware (MUST BE LAST/OUTERMOST)
# This ensures CORS headers are added to ALL responses, even 403 errors.
from fastapi.middleware.cors import CORSMiddleware

# Configure CORS origins from environment variable
cors_origins_str = os.environ.get('CORS_ORIGINS', 'http://localhost:3000')
cors_origins = [origin.strip() for origin in cors_origins_str.split(',')]

# Add common localhost variations if not already present
common_origins = ["http://localhost:3000", "http://127.0.0.1:3000", "http://localhost:8001", "http://127.0.0.1:8001"]
for origin in common_origins:
    if origin not in cors_origins:
        cors_origins.append(origin)

app.add_middleware(
    CORSMiddleware,
    allow_origins=cors_origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
    expose_headers=["*"],
)


logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("startup")
async def startup_db_init():
    """Initialize database with default users on startup"""
    try:
        from init_db import initialize_database
        await initialize_database()
    except Exception as e:
        logger.warning(f"Database initialization warning: {e}")

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()